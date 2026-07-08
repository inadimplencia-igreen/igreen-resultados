import streamlit as st
import pandas as pd
from pymongo import MongoClient
from datetime import datetime, date
import io
import base64
import re
import warnings
warnings.filterwarnings('ignore')

st.set_page_config(page_title="Inadimplência Performance", page_icon="logo.png", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
* { font-family: 'Inter', sans-serif !important; }

/* ── BASE — tema claro corporativo ── */
.stApp { background-color: #eef6ee !important; --text-color: #1a3a1a; --text-muted: #5a8a5a; --bg-card: #ffffff; --border-color: #c8e0c8; }

/* ── SIDEBAR ── */
[data-testid="stSidebar"] {
    background: #f5fbf5 !important;
    border-right: 1px solid #d0e8d0 !important;
}

/* ── Esconde label Menu do radio ── */
[data-testid="stSidebar"] .stRadio > div > p { display: none !important; }

/* ── SIDEBAR NAV — cards padronizados ── */
[data-testid="stSidebar"] .stRadio > div {
    gap: 3px !important;
    padding: 0 8px !important;
}
[data-testid="stSidebar"] .stRadio label {
    color: #2d4a2d !important;
    font-size: 13px !important;
    font-weight: 500 !important;
    padding: 10px 16px !important;
    white-space: nowrap !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
    display: flex !important;
    align-items: center !important;
    border-radius: 8px !important;
    margin: 1px 0 !important;
    width: 100% !important;
    box-sizing: border-box !important;
    transition: all 0.15s !important;
    background: #ffffff !important;
    border: 1px solid #d8ead8 !important;
    box-shadow: 0 1px 2px rgba(0,0,0,0.04) !important;
    min-height: 40px !important;
}
[data-testid="stSidebar"] .stRadio label:hover {
    color: #1a3a1a !important;
    background: #edf7ed !important;
    border-color: #2e7d32 !important;
}
[data-testid="stSidebar"] .stRadio [data-baseweb="radio"] > div:first-child { display: none !important; }
[data-testid="stSidebar"] .stRadio [data-testid="stMarkdownContainer"] p {
    color: inherit !important;
    white-space: nowrap !important;
    font-size: 13px !important;
    margin: 0 !important;
}
[data-testid="stSidebar"] .stRadio label[data-checked="true"] {
    color: #ffffff !important;
    background: #2e7d32 !important;
    border-color: #2e7d32 !important;
    font-weight: 600 !important;
    box-shadow: 0 2px 6px rgba(46,125,50,0.2) !important;
}

/* ── METRICS ── */
[data-testid="stMetric"] {
    background: #ffffff !important;
    border: 1px solid #e0e8e0 !important;
    border-radius: 10px !important;
    padding: 16px 20px !important;
    border-top: 3px solid #2e7d32 !important;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06) !important;
}
[data-testid="stMetricValue"] { color: #1a2e1a !important; font-size: 16px !important; font-weight: 700 !important; white-space: nowrap !important; overflow: visible !important; }
[data-testid="stMetricLabel"] { color: #5a8a5a !important; font-size: 10px !important; text-transform: uppercase; letter-spacing: 1.5px; font-weight: 600; }

/* ── BUTTONS ── */
.stButton > button {
    background: #f0f7f0 !important;
    color: #2e7d32 !important;
    border: 1px solid #c8e0c8 !important;
    border-radius: 6px !important;
    font-weight: 500 !important;
    font-size: 12px !important;
    padding: 6px 14px !important;
    transition: all 0.15s !important;
    box-shadow: none !important;
}
.stButton > button:hover {
    background: #2e7d32 !important;
    color: #ffffff !important;
    border-color: #2e7d32 !important;
}

/* ── TYPOGRAPHY ── */
h1 { color: #1a2e1a !important; font-size: 20px !important; font-weight: 700 !important; }
h2 { color: #2d4a2d !important; font-size: 16px !important; font-weight: 600 !important; }
h3 { color: #5a8a5a !important; font-size: 10px !important; font-weight: 600 !important; text-transform: uppercase; letter-spacing: 2px; }
p  { color: #1a3a1a !important; font-size: 13px; }
hr { border: none !important; border-top: 1px solid #e0e8e0 !important; margin: 14px 0 !important; }

/* ── INPUTS ── */
.stTextInput input, .stNumberInput input, .stTextArea textarea {
    background: #ffffff !important;
    border: 1px solid #c8e0c8 !important;
    color: #1a2e1a !important;
    border-radius: 8px !important;
    font-size: 13px !important;
}
.stTextInput input:focus, .stNumberInput input:focus, .stTextArea textarea:focus {
    border-color: #2e7d32 !important;
    box-shadow: 0 0 0 2px rgba(46,125,50,0.12) !important;
}
.stTextInput input::placeholder { color: #9ab89a !important; }

/* ── SELECTS ── */
.stSelectbox > div > div {
    background: #ffffff !important;
    border: 1px solid #c8e0c8 !important;
    color: #1a2e1a !important;
    border-radius: 8px !important;
    font-size: 13px !important;
}
.stSelectbox > div > div > div {
    color: #1a2e1a !important;
    white-space: nowrap !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
    max-width: calc(100% - 32px) !important;
}
[data-baseweb="select"] { background: #ffffff !important; }
[data-baseweb="select"] > div { background: #ffffff !important; color: #1a2e1a !important; }
[data-baseweb="select"] input { opacity: 0 !important; position: absolute !important; width: 0 !important; height: 0 !important; pointer-events: none !important; }
[data-baseweb="select"] svg { display: none !important; }
[data-baseweb="popover"] { background: #ffffff !important; border: 1px solid #c8e0c8 !important; border-radius: 8px !important; z-index: 99999 !important; }
[data-baseweb="menu"] { background: #ffffff !important; }
[role="option"] { background: #ffffff !important; color: #1a2e1a !important; padding: 8px 14px !important; }
[role="option"]:hover, [aria-selected="true"][role="option"] { background: #e8f5e8 !important; color: #1a2e1a !important; }

/* ── TABS ── */
.stTabs [data-baseweb="tab-list"] {
    background: #f0f7f0 !important;
    border-radius: 8px !important;
    padding: 4px !important;
    gap: 3px !important;
    border: 1px solid #c8e0c8 !important;
}
.stTabs [data-baseweb="tab"] { color: #5a8a5a !important; border-radius: 6px !important; font-size: 12px !important; font-weight: 500 !important; padding: 7px 14px !important; }
.stTabs [aria-selected="true"] { background: #2e7d32 !important; color: #ffffff !important; }

/* ── CHECKBOXES ── */
.stCheckbox label { color: #2d4a2d !important; font-size: 13px !important; }

/* ── FILE UPLOADER ── */
[data-testid="stFileUploader"] > div {
    background: #f8fdf8 !important;
    border: 1.5px dashed #c8e0c8 !important;
    border-radius: 10px !important;
}
[data-testid="stFileUploader"] * { color: #5a8a5a !important; }

/* ── ALERTS ── */
.stSuccess > div { background: #f0faf0 !important; border: 1px solid #c8e0c8 !important; color: #2e7d32 !important; border-radius: 8px !important; font-size: 13px !important; border-left: 3px solid #2e7d32 !important; }
.stWarning > div { background: #fffbf0 !important; border: 1px solid #f0e0a0 !important; color: #8a6a00 !important; border-radius: 8px !important; font-size: 13px !important; border-left: 3px solid #f0c000 !important; }
.stError > div { background: #fff5f5 !important; border: 1px solid #f0c0c0 !important; color: #c62828 !important; border-radius: 8px !important; font-size: 13px !important; border-left: 3px solid #c62828 !important; }
.stInfo > div { background: #f0f8ff !important; border: 1px solid #b0d0f0 !important; color: #1565c0 !important; border-radius: 8px !important; font-size: 13px !important; border-left: 3px solid #1565c0 !important; }

/* ── DATA TABLE ── */
[data-testid="stDataFrame"] { border: 1px solid #e0e8e0 !important; border-radius: 10px !important; background: #ffffff !important; }

/* ── EXPANDER ── */
.streamlit-expanderHeader { background: #f8fdf8 !important; border: 1px solid #c8e0c8 !important; border-radius: 8px !important; color: #2d4a2d !important; font-size: 13px !important; }
.streamlit-expanderContent { background: #ffffff !important; border: 1px solid #c8e0c8 !important; border-top: none !important; border-radius: 0 0 8px 8px !important; }

/* ── SCROLLBAR ── */
::-webkit-scrollbar { width: 4px; height: 4px; }
::-webkit-scrollbar-track { background: #f0f7f0; }
::-webkit-scrollbar-thumb { background: #c8e0c8; border-radius: 2px; }
::-webkit-scrollbar-thumb:hover { background: #2e7d32; }

/* ── Esconde ícones arrow do expander ── */
.streamlit-expanderHeader svg { display: none !important; }
.streamlit-expanderHeader [data-testid="stExpanderToggleIcon"] { display: none !important; }
details summary svg { display: none !important; }

/* ── HIDE STREAMLIT CHROME ── */
[data-testid="stSidebarCollapseButton"] { display: none !important; }
[data-testid="collapsedControl"] { display: none !important; }
button[data-testid="baseButton-header"] { display: none !important; }
.st-emotion-cache-1egp75k { display: none !important; }
[data-testid="stSidebarContent"] > div:first-child > div:first-child { display: none !important; }
button[kind="header"] { display: none !important; }
#MainMenu { visibility: hidden !important; }
header[data-testid="stHeader"] { display: none !important; }
footer { display: none !important; }
[data-testid="stToolbar"] { display: none !important; }
[data-testid="stDecoration"] { display: none !important; }
.viewerBadge_container__1QSob { display: none !important; }
div[data-testid="stStatusWidget"] { display: none !important; }

/* ── SIDEBAR SEMPRE VISÍVEL ── */
[data-testid="stSidebar"] { display: flex !important; visibility: visible !important; opacity: 1 !important; width: 260px !important; min-width: 260px !important; transform: none !important; }
[data-testid="stSidebarCollapsedControl"] { display: none !important; }
section[data-testid="stSidebar"] { display: flex !important; }

/* ── LAYOUT ── */
.block-container { padding: 2rem 2rem 2rem !important; max-width: 1200px !important; }
div[data-testid="stVerticalBlock"] label { color: #3a5a3a !important; font-size: 12px !important; }

/* ── FORNECEDORAS sem quebra ── */
.forn-nome { white-space: nowrap !important; overflow: hidden !important; text-overflow: ellipsis !important; max-width: 160px !important; display: inline-block !important; }
</style>
""", unsafe_allow_html=True)

LOGO_B64 = "data:image/webp;base64,UklGRkYXAABXRUJQVlA4IDoXAADw3ACdASorAyoCPm02mkmkIyKhIRHYiIANiWdu4XZ4q+d/YD9AKWV6xPLvANaXYBbv63dj7CjMeT3YnnUdE+dv/k+t79Qewf+qPUu82n7YesJ/wvXp/oPSM6p30ZPLk9qn9r/SZzkT9hO4HaTsFdo32uUFdpPACer2h2JXgh/Q60qeb/3vMD+2b8+OnPnKMUfcmSWxS69yZJbFLr3JklsUuvcmSWxS69yZJbFLr3JklsUuvcmSaQMHI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HIP8hPqfByOQf5CfU+Dkcg/yE+p8HI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HIP8hPqJDhX1Pg5HIP8hPqfByOQf5CfU+Dkcg/yE+p7vd44HbcHI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg4S2SrzhSc8tyKs8p5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HHTUHFqJp0Mmk7brirtASPN+Qn1Pg5HINDOIvLcoPiyayfU+Dkcg/yDwwQ8/qvaIHhdCxa7UetvyE+p8HI5B9eemgRga/S/bjCzL7qefFSfui9Ztu//aISTMbklescqTJTQu7i0lzcTYU7UkHrzhCR5vyE+p7vsmKx4VuTfnIP8hPqfByOOmZikgqldwMfUNzTOBeafG/mn0GLlPnY4eJHfkJ9T4OQ6pewnOp8HI5B/kJ9RBkV2TQc06e3QyLlM1kij7hJZ6iskXQkThVSm8WEB5HF7XdqIseLILaUNp1Pg5HIP8hPqfByOQf5CfKmZGDuQK7mm24Rpc2txUg9ZA4omotF55fCTfAHWVJerL74KScKuTP1qOBTwOy5W0++WNjkH+Qn1Pg5HIP8hPqfBw/gWsoN0UTPCvB8SH0l5voMSAkkggzqfzJxddUeEmZM0XZMnU4POQq+PbAz5d+h7Y71j2uchX+GSfKLmbLZASLBI835CfU+Dkcg/yE+p8Jh9zfbOICMLAl8HKcqdpre9mCZ1dcokC/ZQbVRmgaeZB5zR07wUwnoGoBEm2wtHkH+Qn1Pg5HIP8hPqfByYwKWBRlWTCfbi8MHtqarr7DLNROf/yKXN8kctW2FJvqI6ran6Tazlb4P8hPqfByOQf5CfU+DkxgUeJ04T7jbhdanAT0cACAALBWNlFeYdrw2Nz8HJBSTTHdVy+qRuainYau+ytgLHmHnf/O2VBVDX86VjKGqfNHm/IT6nwcjkH+Qn1Pg5HegUsJ0DWzljJMfgCulAgPVQD7tJ8wGGi6c4k87pgNagiir0F5Ruq2YHWQHg/e/tnP5YmAs08laTYYCZ0WFrgn1Pg5HIP8hPqfByOQf5DZWM4yhae5Y7EyvMSXE+FGbVaEWJs38KNiCbViMsYQEWtbwrsXuIwqclcwMUTE4vVU6KwAdrQ1DTnjoEK+6j/WPaZ4nVB0MymvLYWjyD/IT6nwcjkH+Qn1Pg5MYHHNyo64kswWvtw3HhMGQIpQbjGn+gz4KvAk4VbXfKt5zv//5qCYR2/AcyOUZTnKwtjpU6fLY26Zq/uCm5WkyjdJ/UT6nwcjkH+Qn1Pg5HIP8oKnxAe+VW9NI3lZNnw+atFzTpr+G23CB+v1ZjXux1w4tmxH7Iapyciea0ALhqYdlRyOQf5CfU+DkchBYWyH3Zj+LGWFBJH+CkBVaBjd849NCm0CaJucKHTKcHCjB5n7yDkcg/yE+p8HI5B/kJ9T5XVfSdRxPwOa9ng4TGVD2s2CHMWC1DJkkyxfMxR7HKdq/6F/C1FMImeEvg5HIP8hPqfByOQf5CfU7K/y9kgwcT1TnNPRyplc5QHr5xG7uH8zl20fQoCsn65fILMzmmP8cqUvtDen1Pg5HIP8hPqfByOQf5CcxhhTPFChP0lu8isbgG7r/buSfEky46TTR+ZzMypX5CfU+Dkcg/yE+p8HI5B9j9Vgo1xv9BZFQt/27kiP5lHYc086i6nvrtoYKZuSlSNZYQEjzfkJ9T4ORyD/IT6nwcNnIOFlwJJFTH6PNKLLPva//aISTMbko7n3AuXQI7hfLidBU9KQnyn1f9LuSPN+Qn1Pg5HIP8hPqfByOQfgL+aWxeQNtSs0OsGCfU+Dkcg/yE+p8HI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HIP8hPqfByOQf5CfU+Dkcg/yE+p8HI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HIP8hPqfByOQf5CfU+Dkcg/yE+p8HI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HIP8hPqfByOQf5CfU+Dkcg/yE+p8HI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HIP8hPqfByOQf5CfU+Dkcg/yE+p8HI5B/kJ9T4ORyD/IT6nwcjkH+Qn1EAAA/v/Ba9THJIzmnUSVZEgeVDPuL3enkDK8eokqyJA8qGfcXu9PIGV49RJVkSB5UM+4vd6eQMrx6iSrIkDyoZ9xe708gZXj1ElWRIHlQz7i93p5AyvHqJFZa+O0xaYf/QFkV2PTk+RCqIiW7PXAGVx5ZpNHFQLbz/PhdsVosmLLo8Urhnzum/njXjDVDf5SoXdosh4LTbPvXCdW1ccKV4687K/mBOVrBEnkfNEp9O8yFMjR7Fdtp4NdxUjD0Nw/k6fjkxkrPuwSrrPKAc8RN/FtIpqjkmzn6/pjvUz3Enxmh2HQ4iHvBgVB5X9H+cdUyKI33JhxBvhLu4mN9z57hHo8MxU5EY54L9FcSrFG8eppWmBdhXp+mYqTU/zYjpcsoU5rDEOz16f7YiEQiEB8gt/IJDU7dEpcOBf9ELkvXsn9n6Hfnq+2ltfUXeauNPaFOZ8SvpNCc82w2NiNwQqbOvenGNs7pOEvkT3XVYZCzmf58u+iSM4Cb/JrGePxv2J8WNc7zeYX1hS1O4QOaCfIavBCGpzT2IszQyPtOe59/K0mOh/E4Yw2ToJ8JiG9hhE8X+R2xNlRZVjIlrBXv+adtm0Gn/B/E+8vzgNunRdvtVFyXvV0vCw/ZohAvkDSta0MdjZROS9VrzXhIV4N44wY9cwQGnwmE3nZORI5oYWDV6WWKVk+Y2W7EY7/Mfq3ZaktH8WVtvmK0JJSVfN4AcKRtVWgDed2B00OQFePmxudFLiZjKHUnOQ1oS1phHI3KEn/Z+WKMtYZEyKzEDtr7kqpcTxZhe/XCXOFl30A1vjMEylLWqm3zjc5f773os9HWaJL77rP+F/K8z+UJ8iOfkRGewe8vq35h/nWybbtSlQHFR/BqmjzkT5rdPSkyynpzG3v9P/953I3m0s5rYcGQ8UEqx8viZIOX4x1Nvrlr3pacf25ZTaK2fVI229yS3Nllx+cs4U5ewqj/udeCnz2EyNlmhvcmma4p2Cag4D574GextSi8RL591k7aJ3eFcUYcldl20PO88Pp3HVvl2R//O9QMiSoGe/oo5DEs7khUjtn29oB0o0qzzLB7JfdSZO3iadLoElYTlHwXjEknnyAr6L0ISW+b+5l9JzfT9QCRJP8aK5Yp1jxg4yLVnBbz54BJTnpzrH3KhMdb/AUrAXRwNCu/yuNVJQZ5lHzB+abtfXu+nMF+ldfTzp6PXNyBvDFlzlkrEdmyGMpM2N93TuY+qmGmxw+670GUaoeQ/aIBBTjl3HgXLZaVOadXUHBKiOZZo4lB8yFGXECw62W+IpSn2XAjUNXkirDoWWJOTFaMV2OVCjqUWqhaWy8yaLYln4d/lkf/wpWuhpSQAGdu8Tyx4czQNCVGw2b1CpkBS5NE5n/59Mv3PQrJ/x6tR3OTLzpoH/fqehcnSKO46zd35h8GVEYtxT6OG1GOOoPKJJ+ADQhuTb0r7XdNiixSr2q8EWqMLVZbFS+Iih9UNY5bKsYdgvS+HqgLMDSKkQ7EM3i03LULV3Ds7zHa37eDxvlU+CmD9pf54YB6km+IokM3WDd5cF8z0sMUqSluZo7upZ88OfL2vIiJ4lKL4KV4jOlM7R/a6Q0hSONgHfcODaiLYBBqdzGibYz3g4QrQ0G/GGblQq/HNlCJ2e3yRke5FZwfZGiqEPJ1S6xrrI5V3yQix3w5IaunTWwlt7td5FnJLXFXiDAYXxXPYHJyIe1deXBQNJijhej2oAUB44T1JAChHQinM1kbO8GM57NivWfHE5mkBd023gyuZ6fYKLlqoL7SDVWyiTiDL0uLmgpA83dJTng+TAlyo4EgGWOy2KJEGTbWALf8eH1jTFxI6NbzRtEHmNeT4+Kgz0TKxp7Np4fO150jjWRcSUfL6kG3/8Ag5gUuxLBNo9BJvA4RFWi/ZBgY2E16JAO0TjgV6x4LMyYqer9Fqk73KIQQcJo8DBKz4FZCfautquuR7p/xSc9asvgViUHLCVQ66wcGAmcRlhpPlGsPxC5bP8sF9oCFgGOWuou5zVY0epjQuwtn9XP2c19dIHP2l2nMVshHecyFRMdZdX70WpUfIM4kMitXwJBiIfjefC6JEcfP6XutHw9FWF/Rt8YSf+GiHt3wWdRFCN2J2U5b3jUxy2EbmQXCa4wte8g8lxRbP8GVDSEIggvUgvFGxTjmJ104wL8/Lhv5+1cSaIiV+J6xuiSITVkz2CB902XuhLkLNUa70MiNlD5ltyATjhFHSZ3mkAaYNPY2bhsdm9NxYhZVaCL9qjk0HnY5sLn0Xab/nOeZtBXxXRT55iYAervHxTqIUktipkDFl4Wy8yfXTSNGGjgquI4KAZOR+buoAnrmXPDP8RA8XqVbMk9IriccQScd/KAKRV4CnPb9Q5PaIV9z3urT7cYeKESK4or/z3R0/79xo01FZoa8ZfNu7j5xuu3C1n7PetqFN2y2l/g8LHjeALR5c5pvqR6/zI/X8UaZ6nj/59VWsV1A0oSlhzV1y6K8AfxMY80IO8N+ebOZ8JvFwIUxSpOym4qTKTiQbwcTm90cioBJn//f/pf9xb/2Y6l3QOaAP2jjhVlX2ocbrnSB6AZDsaFeu9Y/cbqocX2wcsNWOvwCfQj8HiHDJIb/bC484oGWWuq7JoDdiqsjtalx3sVPQzxBuyvkEk/fCILTWgbsvW7spWct3neeD+zGdKEgoeCICovCU/R34qzEL1uVOQaLjeZ0GUoOnoryWQ+c6ITxO35IDh+qMr2Pi2hWtUS5nlgXH/tTFu0LndEs0mw0V+6mT9cZcWvI4uZ5uRyBOWITynzmFKGfYmtcCv1yxu5chaU92vIOu4Dmd+AcQgkfmzCQX935zcNk90hzR3cK2WKKnlhgZEiLgoV8/Midnlfq7juUbamV15AcQqQqAdJBFebKva4e4QIiQA0R5SCDsxn8qeNK0RyJCtszmmnSfQzenSD200SL0kBBqui91eTeBFB2/mXdQoZAyfZ1SRSRZO3jcJSuBltnLjnrLPGMAUjjplqSRysZhWbk/VoK3vg15rmDTJJqRvC0V30DkotzSM4qU8YpnZwEnUiUVjVsOJgGZjsqNi+FARghVhwdxmlsAd1aPqB1EJ9NqCdeTLuU5yztWWMATz32Vvc3VdCMOUW70EGdcy59JjR6L83vIsHptjcl3HDzr3ovDtV/uLkHyeUbiCgF4avSz4/5yJ3uLFrAUdUGFxdc+WIT5cLuK16nWp33VuJ8VTbD6SiU3osiyUeUnvttihtwtpU3UVmZuEUQqV7MbkX6k4TN6JqVwR1ln4ZIXTRW7VhPcLVvyXQ7kg2qW0+b98+n4Gg+L5eXs+2NhO20zTGnAsBb99AygxbjQmTVCm25NDWFIhoMa04Z9bXN89EqO2M2X3jWmnXXZ1m2EYmiv6gCBfftD0un3QSkD2u7xjakw2I9hwEad+oQdP0oUIB44s0VjE/ZQmKrOQs0WiavE4xVfFOYJQdAwtcGHKxKOlec3LBhnhHQvM5W4FTVor1O/LcI9fs5fZ/SZ7WE4hNNfNL/EPWT8nH50ODv9n7ScRYO0cvFw6LFT/YAroFLseGDru7jBMz0fisqnLL1lQtopT6w9GDuizE+hEVQd9qVN6uN7YjudPYEvchw3hvgMxHGWN+tSPju4E8MazGH+r6QMFSLCZGCM94PChlHfUzb7WTJ7j4ua0doYUGJdGAxXGqcWYdDZ0Kk4TrDXWkxzpNNrlgyxkK+8tMXyqClfL8LL266E23yIM603GtG7TNUiH9m03aEH4YNjOO8H1bBBxbh6jyO5D//gDLBVjRTrMP3KKOIBFoCipYdULnpZRbPvl58PldwURixOQ8utLtFxOaBLMjJ1FPOly0csbmoOqpCQwzN1s6RX+jUn4FFkh3JkjBLetI+TrU8PKi2tGC6qAhENCiMImkCg/oD7Yspi9Se0/7JGmUBFU4N96T1eHKoeI8kF0sHCjXd/ZZoKJDHssOcgWZzklYpP6CpwGuTMVtuSnzvkoBPhtk8YnIKR8GwzMWpVFvB44nSNzoDJ3NcJUh79AglfccmReXjZOBg+F/kJAkADhw57QADQGElQ6OItrrDnOWMvMCk925i5CNgRrnhugCgRXrXI5YP9uoaDVEovMJzh26gSmnuG8s3cdYVrPBtRt9eldx2IHWpXO8n+7/LM2WH3YIs6LoHdygmAW3Kkk/vjUvVj/qMOaXjiAON6Oq3lEqkCKx4UgyoULMxPgWevNBm6osmtQZYOt93NHEVI5u85IfjKkJF9DB+DVgWSoYubRUZtzpbL7I5R1RurzcS/REwnnCRlRMQc/qabrbpLlo6KLArK0vQ8kekHjO4Rxa8maH6djlrA482uwPVhAyCwY0L4JNcoDt9mbaUpKRB01kJ7Uq3o2E9V9BXIqDhKQ/9szos87vr/ChzzpqnDM79D1FMui4T4KRX/iZQAW9jKyNIPXPzLQM/pHG5fyKS1xTvZ4KFukCf5MCsyWw9LxQJDSYJhMjBLpdQSNi06HVBHj/ZY5mAzZo10VcOwSGN76k1g7d5jYv999BV+e1Ogx9SLPhDLN0PdadiozEAuqAAAkwYuddZ3ZvB9CDIrohL89ypfdWbcTiWcpFMdlpSSEvPUYhYUA2dgxdQTKPPfQ+K/LuxMNuN5tWL/PED5ouexw3X3PqeimOr3bwxPvh9z4KZxP6akCOTy6AzPsklkgJk3eugTs881gEhiVhPNAPrO2BLsFXTj6yg5iRa5NlgXI+a9ITIAIL/cYMfdudyw0NK6mniI0Twiuussv+D5z0kyt5kHhJ6vdX5rPLhG7wJvbKrX036jvapmbxQL2+QthagbUQtbzYy7rnvHu7sUe1f0UlCrRM0AEYR3szCi2kmR6NSPHFdRBqG3YWDRbuUdovmzbXMuOqwuGFvUbXKq1PvqipGXd0yYlHflQFhllVr0/+SHdYJk7X1Uor14ksnWk+IOan5xlI2j3/RlKzN3JGr5mvLAhJOHqYpkc3LUA/ExeVbt/QHoeL4+Hb6u+7jCZ3v9EJc7CrGTmXB3pn9uatGHqB4dWWzJTtwqKygrUWLuYLiEpgDJK9p/y39ORq9t4B8u3/0iD1zAiW37RBJEj+3IJ3k6XiOzj9UAx2rKs2rbYXvLHgwnhSbAhdrvJCNw6rNSLKAZAelPnZ2lQIceY17VNW3PBL8GJttheN77lSu7ykybM3+lMJimCmWViqXsEPIfGpgCrsIeMTsVqHx8hBxPGYlcaOND5ITi+mBtEau/SpqAxEBLnX6cbggOxdgcxldZ0HU91lHsj4wbB+glZTzD0txBq0/0uwLMDBInV3Vp2Z7Lojsch8ctus4U4nV4nNGk/sRnTws3C8dgvppMrHSCSmDjcBpATeezwyOXVy4OkzG365HpR4/vx9EeZnG4Xcug/pES0hIufbKaqpDeDPLEdTLKouiqrf5tM47VV30RhHCXZzFIueoBi16XhCHI/PRkMEqMxFsfPaP+qmT/YiFRF6GwAsNcjy4/+s+PA7ITB0DsE4srhSUpkbK+91WtkapOIvnN8l/0BfOb53/QIBzfJf9AX8pTEiAAA="

def _get_usuarios():
    def _s(key, fallback):
        try: return st.secrets["usuarios"][key]
        except: return fallback
    return {
        "tamires": {"senha": _s("tamires","9cd2r11QvOqD8a"), "equipe":"tamires","role":"admin",  "nome":"Tamires"},
        "luciano": {"senha": _s("luciano","TCLemDjWSGv!yz"), "equipe":"luciano","role":"gestor", "nome":"Luciano"},
        "deborah": {"senha": _s("deborah","L4f10IJo5bGJ3O"), "equipe":"deborah","role":"gestor", "nome":"Déborah"},
        "veloso":  {"senha": _s("veloso", "U2B!niJH7W96rL"), "equipe":None,    "role":"diretor","nome":"Veloso"},
        "moyara":  {"senha": _s("moyara", "ug8omeP4Cvt3nl"), "equipe":None,    "role":"diretor_upload","nome":"Moyara"},
        "gabriel": {"senha": _s("gabriel","gabriel123"),   "equipe":"metcool","role":"gestor", "nome":"Gabriel"},
    }

USUARIOS = _get_usuarios()
EQUIPES = {
    "luciano":{"nome":"Luciano","cor":"#2daf5c"},
    "deborah":{"nome":"Déborah","cor":"#a855f7"},
    "tamires":{"nome":"Tamires","cor":"#f97316"},
    "metcool":{"nome":"Meet Call","cor":"#3b82f6"},
}
MESES_NOMES = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
               "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
CRITERIOS_PADRAO = [
    {"id":"c1","num":"1º","nome":"Abertura e Identificação","peso":5,"itens":["Saudação adequada","Identificação do operador e da empresa","Sem conversas paralelas fora do mudo"],"obrigatorio":False},
    {"id":"c2","num":"2º","nome":"Comunicação e Postura","peso":5,"itens":["Clareza na fala e respeito com o cliente","Tom respeitoso, sem ironia ou pressão","Escuta ativa — não interromper"],"obrigatorio":False},
    {"id":"c3","num":"3º","nome":"Diagnóstico da Dívida","peso":10,"itens":["Questionar o motivo da inadimplência","Recorda do contrato? Recebeu boleto? Tem acesso ao app? Previsão de pagamento?"],"obrigatorio":False},
    {"id":"c4","num":"4º","nome":"Negociação","peso":40,"itens":["Argumentação de benefícios do pagamento pontual","! Obrigatório: perguntar sobre dúvidas em boletos","! Obrigatório: perguntar sobre acesso ao app","! Obrigatório: falar sobre iGreen Club (mín. 2)"],"obrigatorio":True},
    {"id":"c5","num":"5º","nome":"Conformidade","peso":20,"itens":["Questionar o motivo do cancelamento","Não ameaçar ou constranger"],"obrigatorio":False},
    {"id":"c6","num":"6º","nome":"Registros e Procedimentos","peso":10,"itens":["Registro correto no sistema","Classificação adequada da ligação"],"obrigatorio":False},
    {"id":"c7","num":"7º","nome":"Encerramento","peso":10,"itens":["Esclarecimento do acordo fechado","Agradecimento e cordialidade"],"obrigatorio":False},
]
ERROS_CRITICOS_PADRAO = [
    {"id":"e1","nome":"Informação incorreta","desc":"Passou informação incorreta, incompleta ou errada ao cliente"},
    {"id":"e2","nome":"Postura ríspida","desc":"Agiu de forma ríspida ou ameaçadora"},
    {"id":"e3","nome":"Linguagem agressiva","desc":"Usar linguajar agressivo com o cliente"},
    {"id":"e4","nome":"Retenção de ligação","desc":"Segurar a ligação até dar o tempo legível para cota"},
    {"id":"e5","nome":"Contra-argumentação indevida","desc":"Cliente reclama do desconto e você oferta conta única"},
]
CRITERIOS_CHAT_PADRAO = [
    {"id":"cc1","num":"1º","nome":"Abertura e Identificação","peso":10,"itens":["Saudação adequada","Identificação correta do operador e da empresa"],"obrigatorio":False},
    {"id":"cc2","num":"2º","nome":"Comunicação e Postura","peso":15,"itens":["Clareza na comunicação","Tom respeitoso (sem pressão excessiva, ironia)","Ausência de gírias, vícios de linguagem ou informalidade excessiva","Não ameaçar ou constranger"],"obrigatorio":False},
    {"id":"cc3","num":"3º","nome":"Negociação/Conformidade","peso":40,"itens":["Apresentação clara de propostas","Seguir script de Retenção (incluir benefícios e disponibilidade das faturas dentro do aplicativo)","Escrita correta"],"obrigatorio":True},
    {"id":"cc4","num":"4º","nome":"Registro e Procedimentos","peso":20,"itens":["Registro correto no sistema","Classificação adequada do atendimento (tabulação)"],"obrigatorio":False},
    {"id":"cc5","num":"5º","nome":"Encerramento","peso":15,"itens":["Esclarecimento do acordo fechado","Agradecimento e cordialidade","Encerramento de atendimento"],"obrigatorio":False},
]
ERROS_CRITICOS_CHAT_PADRAO = [
    {"id":"ec1","nome":"Informação incorreta","desc":"Passou informação incorreta ao cliente"},
    {"id":"ec2","nome":"Postura ríspida","desc":"Agiu de forma ríspida ou ameaçadora"},
    {"id":"ec3","nome":"Linguagem agressiva","desc":"Linguagem agressiva por escrito"},
    {"id":"ec4","nome":"Ignorou mensagem","desc":"Ignorou mensagem/deixou sem resposta (Encerramento sem conclusão)"},
    {"id":"ec5","nome":"Desconto fora da regra","desc":"Ofereceu desconto fora da regra"},
    {"id":"ec6","nome":"Boleto inválido","desc":"Enviou boleto inválido e finalizou sem verificar"},
]
FAIXAS_PONTOS = [(0,70,0),(71,80,300),(81,90,500),(91,95,700),(96,99,1000),(100,100,1100)]
SEMANAS_MONITORIA = [
    "1ª Semana — 1ª Monitoria","1ª Semana — 2ª Monitoria",
    "2ª Semana — 1ª Monitoria","2ª Semana — 2ª Monitoria",
    "3ª Semana — 1ª Monitoria","3ª Semana — 2ª Monitoria",
    "4ª Semana — 1ª Monitoria","4ª Semana — 2ª Monitoria",
]
OPERADORES_PADRAO = {
    "luciano":[("Jennifer Silveira",True),("Paulo Roberto",False),("Samires Barros",False),("Maycow Gabriel",False),("Otaides Junior",False),("Heverton Tavares",False),("Camila Nara",False),("Caua Alves",False),("Eduarda Sanqueta",False),("Jheniffer Santos",False),("Ketie Silva",False),("Emanuel Cardoso",False),("Victória Silva",False),("Grasielli Santos",False),("Laura Silva",False),("Michelle Batista",False),("Lorenzzo Pereira",False),("Diogo Oliveira",False),("Maria Paulino",False),("Gabrielle Martins",False),("Marcos Martins",False)],
    "deborah":[("Mikael Dias",False),("Amanda Eduarda",False),("Larissa Barcelos",False),("Nicole Amaral",False),("Sara Rocha",False),("Isabelly Araujo",False),("Silye Paula",False)],
    "tamires":[("Danilo Rodrigues",True),("Raiane Pereira",False),("Wynara Dos Reis",False),("Esteffany Souza",False),("André Gomes",False),("Wanessa Cardoso",False),("Larisse Garcia",False),("Arthur Alves",False)],
    "metcool":[("Leilson Gomes",False),("Hannah Vitoria",False),("Kesia Lima",False),("Renata Ribeiro",False),("Thayna Guerreiro Ferreira",False),("Kimberlyn da Silva",False),("Vitor Eder",False),("Laiza Teixeira",False),("Glaucio Fernandes",False),("Thais de Fatima",False),("Mayara Leal",False),("Ivone Coutinho",False),("Aline Cristine",False),("Anderson Soares da Silva",False),("Michelle Pereira",False),("Maria Ferreira",False),("Mariana Matias",False),("Jessica Faria Albertino Miranda Vieira",False),("Lorrane Moura",False),("Jennifer Edjane",False),("Haissa Batista",False),("Bruna de Barros Santanna",False)],
}
FORNECEDORAS_TODAS = ["COTESA/MOVE","ULTRA","VANTAGE","FARO","BOM FUTURO","SUNCLICK","ATUA","GEDISA","SUNNE","SOLATIO","EDP","FIT","GV","COMERC"]
FORNECEDORAS_POR_GESTOR = {
    "luciano": ["COMERC"],
    "tamires": ["VANTAGE","BOM FUTURO","COTESA/MOVE","SUNCLICK","FARO","ULTRA","GEDISA"],
    "deborah": ["SUNNE","SOLATIO","EDP","FIT","GV"],
    "metcool": ["COMERC"],
}
OPERADORES_MEETCALL = ['Leilson Gomes', 'Hannah Vitoria', 'Kesia Lima', 'Renata Ribeiro', 'Thayna Guerreiro Ferreira', 'Kimberlyn da Silva', 'Vitor Eder', 'Laiza Teixeira', 'Glaucio Fernandes', 'Thais de Fatima', 'Mayara Leal', 'Ivone Coutinho', 'Aline Cristine', 'Anderson Soares da Silva', 'Michelle Pereira', 'Maria Ferreira', 'Mariana Matias', 'Jessica Faria Albertino Miranda Vieira', 'Lorrane Moura', 'Jennifer Edjane', 'Haissa Batista', 'Bruna de Barros Santanna']
CORES_FORN = {"COTESA/MOVE":"#1b5e20","ULTRA":"#0d47a1","VANTAGE":"#e65100","FARO":"#b71c1c","BOM FUTURO":"#4a148c","SUNCLICK":"#004d40","ATUA":"#37474f","GEDISA":"#006064","SUNNE":"#f57f17","SOLATIO":"#4527a0","EDP":"#0277bd","FIT":"#2e7d32","GV":"#558b2f","COMERC":"#37474f"}

# ── MONGODB ────────────────────────────────────
@st.cache_resource
def get_db():
    client = MongoClient(
        st.secrets["mongo"]["uri"],
        serverSelectionTimeoutMS=5000,
        connectTimeoutMS=5000,
        socketTimeoutMS=15000,
        maxPoolSize=20,
        minPoolSize=2,
        retryWrites=True,
        w="majority"
    )
    return client[st.secrets["mongo"]["db"]]

def get_criterios():
    try:
        doc = get_db().configuracoes.find_one({"_id":"criterios_monitoria"})
        if doc and doc.get("criterios"): return doc["criterios"]
    except: pass
    return CRITERIOS_PADRAO

def salvar_criterios(c):
    get_db().configuracoes.update_one({"_id":"criterios_monitoria"},{"$set":{"_id":"criterios_monitoria","criterios":c,"atualizadoEm":datetime.now()}},upsert=True)

def get_criterios_chat():
    try:
        doc = get_db().configuracoes.find_one({"_id":"criterios_chat_monitoria"})
        if doc and doc.get("criterios"): return doc["criterios"]
    except: pass
    return CRITERIOS_CHAT_PADRAO

def salvar_criterios_chat(c):
    get_db().configuracoes.update_one({"_id":"criterios_chat_monitoria"},{"$set":{"_id":"criterios_chat_monitoria","criterios":c,"atualizadoEm":datetime.now()}},upsert=True)

def get_erros_criticos_chat():
    try:
        doc = get_db().configuracoes.find_one({"_id":"erros_criticos_chat_monitoria"})
        if doc and doc.get("erros"): return doc["erros"]
    except: pass
    return ERROS_CRITICOS_CHAT_PADRAO

def salvar_erros_criticos_chat(e):
    get_db().configuracoes.update_one({"_id":"erros_criticos_chat_monitoria"},{"$set":{"_id":"erros_criticos_chat_monitoria","erros":e,"atualizadoEm":datetime.now()}},upsert=True)

def get_erros_criticos():
    try:
        doc = get_db().configuracoes.find_one({"_id":"erros_criticos_monitoria"})
        if doc and doc.get("erros"): return doc["erros"]
    except: pass
    return ERROS_CRITICOS_PADRAO

def salvar_erros_criticos(e):
    get_db().configuracoes.update_one({"_id":"erros_criticos_monitoria"},{"$set":{"_id":"erros_criticos_monitoria","erros":e,"atualizadoEm":datetime.now()}},upsert=True)

def migrar_meetcall_para_luciano():
    try:
        db = get_db()
        for nome in OPERADORES_MEETCALL:
            oid_luc = re.sub(r'[^a-z0-9]','-',nome.lower().strip())
            oid_luc = re.sub(r'-+','-',oid_luc).strip('-')
            oid_luc = f"luc-{oid_luc}"[:40]
            if not db.operadores.find_one({"_id":oid_luc}):
                db.operadores.insert_one({"_id":oid_luc,"equipeId":"luciano","nome":nome,"pleno":False,"meetcall":True,"criadoEm":datetime.now()})
            oid_mc = re.sub(r'[^a-z0-9]','-',nome.lower().strip())
            oid_mc = re.sub(r'-+','-',oid_mc).strip('-')
            oid_mc = f"mc-{oid_mc}"[:40]
            existentes = list(db.operadores.find({"equipeId":"metcool","nome":nome}))
            if len(existentes) > 1:
                for ex in existentes[1:]: db.operadores.delete_one({"_id":ex["_id"]})
            if not db.operadores.find_one({"_id":oid_mc}):
                db.operadores.insert_one({"_id":oid_mc,"equipeId":"metcool","nome":nome,"pleno":False,"meetcall":True,"criadoEm":datetime.now()})
    except: pass

def corrigir_ids_operadores():
    db = get_db()
    for op in list(db.operadores.find({})):
        eq = op.get("equipeId",""); nome = op.get("nome","")
        if not nome: continue
        nid = re.sub(r'[^a-z0-9]','-',nome.lower().strip())
        nid = re.sub(r'-+','-',nid).strip('-')
        idc = f"{eq[:3]}-{nid}"[:40]
        if op["_id"] != idc:
            if not db.operadores.find_one({"_id":idc}):
                db.operadores.insert_one({"_id":idc,"equipeId":eq,"nome":nome,"pleno":op.get("pleno",False),"criadoEm":op.get("criadoEm",datetime.now())})
            db.operadores.delete_one({"_id":op["_id"]})

@st.cache_data(ttl=3600)
def buscar_operadores(eq):
    ops = list(get_db().operadores.find({"equipeId":eq}).sort("nome",1))
    # Deduplica por nome
    vistos = set()
    unicos = []
    for op in ops:
        nome_norm = op.get("nome","").strip().lower()
        if nome_norm not in vistos:
            vistos.add(nome_norm)
            unicos.append(op)
    # Luciano: nunca retorna operadores Meet Call
    if eq == "luciano":
        unicos = [op for op in unicos if op.get("nome","") not in OPERADORES_MEETCALL]
    return unicos

def salvar_operador(eq, nome, pleno=False):
    oid = re.sub(r'[^a-z0-9]','-',nome.lower().strip())
    oid = re.sub(r'-+','-',oid).strip('-')
    oid = f"{eq[:3]}-{oid}"[:40]
    if not get_db().operadores.find_one({"_id":oid}):
        get_db().operadores.insert_one({"_id":oid,"equipeId":eq,"nome":nome,"pleno":pleno,"criadoEm":datetime.now()})
    return oid

def excluir_operador(oid): get_db().operadores.delete_one({"_id":oid})
def atualizar_operador(oid, nome, pleno): get_db().operadores.update_one({"_id":oid},{"$set":{"nome":nome,"pleno":pleno}})

def salvar_meta_operador(ma, eq, oid, v):
    did = f"meta_op__{ma}__{eq}__{oid}"
    get_db().metas.update_one({"_id":did},{"$set":{"_id":did,"mesAno":ma,"equipeId":eq,"opId":oid,"valor":v}},upsert=True)

@st.cache_data(ttl=300)
def buscar_metas_equipe(ma, eq):
    return {d["opId"]:d.get("valor",0) for d in get_db().metas.find({"mesAno":ma,"equipeId":eq}) if "opId" in d}

def salvar_meta_gestora(ma, eq, meta, tpct):
    did = f"meta_gest__{ma}__{eq}"
    get_db().metas.update_one({"_id":did},{"$set":{"_id":did,"mesAno":ma,"equipeId":eq,"metaGestora":meta,"targetPct":tpct,"tipo":"gestora"}},upsert=True)

def buscar_meta_gestora(ma, eq):
    return get_db().metas.find_one({"_id":f"meta_gest__{ma}__{eq}"}) or {"metaGestora":0,"targetPct":125}

def salvar_lancamento_meetcall(ma, total_ligacoes, rec_geral, rec_geral_total=0):
    get_db().metas.update_one(
        {"_id":f"meetcall__{ma}"},
        {"$set":{"_id":f"meetcall__{ma}","mesAno":ma,"totalLigacoes":total_ligacoes,"recGeral":rec_geral,"recGeralTotal":rec_geral_total,"atualizadoEm":datetime.now()}},
        upsert=True)

def buscar_lancamento_meetcall(ma):
    return get_db().metas.find_one({"_id":f"meetcall__{ma}"}) or {}

def criar_lancamento(ma, eq, data_ref, label, agentes, total, sem_int, dt, td, rec_geral=0):
    ts = datetime.now().strftime("%Y%m%d%H%M%S%f")
    get_db().lancamentos.insert_one({"_id":f"lanc__{ma}__{eq}__{ts}","mesAno":ma,"equipeId":eq,"dataRef":data_ref,"label":label,"agentes":agentes,"totalEquipe":total,"semInteracao":sem_int,"diasTrabalhados":dt,"totalDias":td,"recGeral":rec_geral,"criadoEm":datetime.now()})

@st.cache_data(ttl=300)
def buscar_lancamentos(ma, eq):
    novos = list(get_db().lancamentos.find({"mesAno":ma,"equipeId":eq}).sort("criadoEm",-1))
    antigos = []
    for d in get_db().resultados.find({"mesAno":ma,"equipeId":eq}):
        antigos.append({"_id":d["_id"],"mesAno":d["mesAno"],"equipeId":d["equipeId"],"label":d.get("semanaId","Registro anterior"),"dataRef":d.get("atualizadoEm",""),"agentes":d.get("agentes",{}),"totalEquipe":d.get("totalEquipe",0),"valorGeral":d.get("valorGeral",0),"semInteracao":d.get("semInteracao",0),"diasTrabalhados":d.get("diasTrabalhados",0),"totalDias":d.get("totalDias",22),"criadoEm":d.get("atualizadoEm",datetime.now())})
    todos = novos + antigos
    def _sort_key(x):
        v = x.get("criadoEm", "")
        if hasattr(v, 'isoformat'): return v.isoformat()
        return str(v or "")
    todos.sort(key=_sort_key, reverse=True)
    return todos

def excluir_lancamento(did):
    if get_db().lancamentos.delete_one({"_id":did}).deleted_count == 0:
        get_db().resultados.delete_one({"_id":did})

def salvar_monitoria(eq, oid, onome, prot, obs, crits, erros, nota, ma, semana=None, tipo="ligacao"):
    ts = datetime.now().strftime("%Y%m%d%H%M%S%f")
    get_db().monitorias.insert_one({"_id":f"mon__{eq}__{oid}__{ts}","equipeId":eq,"opId":oid,"opNome":onome,"protocolo":prot,"observacao":obs,"criterios":crits,"errosCriticos":erros,"nota":nota,"mesAno":ma,"semana_mon":semana,"tipo":tipo,"criadoEm":datetime.now()})

def buscar_monitorias_operador(oid):
    # Buscar monitorias pelo opId — se tiver vinculadoA, busca pelo original também
    op_doc = get_db().operadores.find_one({"_id": oid})
    ids_busca = [oid]
    if op_doc and op_doc.get('vinculadoA'):
        ids_busca.append(op_doc['vinculadoA'])
    return list(get_db().monitorias.find({"opId": {"$in": ids_busca}}).sort("criadoEm",-1))

def buscar_monitorias_equipe(eq, ma=None):
    f = {"equipeId":eq}
    if ma: f["mesAno"] = ma
    return list(get_db().monitorias.find(f).sort("criadoEm",-1))

def excluir_monitoria(did): get_db().monitorias.delete_one({"_id":did})

def salvar_processamento(ma, eq, df, usuario_nome=""):
    # Calcular métricas antes de salvar
    df_num = df.copy()
    df_num["valor"] = pd.to_numeric(df_num.get("valor", pd.Series(dtype=float)), errors="coerce").fillna(0)
    elig = df_num[df_num["elegibilidade"]=="Elegível"] if "elegibilidade" in df_num.columns else df_num
    valor_elig = float(elig["valor"].sum())
    boletos_elig = len(elig)
    clientes_elig = int(elig["uc_cpf"].nunique()) if "uc_cpf" in elig.columns else 0
    forns = sorted(df_num["fornecedora"].dropna().unique().tolist()) if "fornecedora" in df_num.columns else []

    # Calcular breakdown por fornecedora (e por UF se disponível)
    por_fornecedora = {}
    if "fornecedora" in elig.columns:
        for forn, grp in elig.groupby("fornecedora"):
            por_fornecedora[str(forn)] = {
                "valor": float(grp["valor"].sum()),
                "boletos": len(grp),
                "clientes": int(grp["uc_cpf"].nunique()) if "uc_cpf" in grp.columns else 0
            }
            # Se tiver UF, salvar breakdown por UF dentro da fornecedora
            if "uf" in elig.columns:
                por_uf = {}
                for uf, grp_uf in grp.groupby("uf"):
                    if pd.notna(uf) and str(uf).strip():
                        por_uf[str(uf).strip().upper()] = {
                            "valor": float(grp_uf["valor"].sum()),
                            "boletos": len(grp_uf)
                        }
                if por_uf:
                    por_fornecedora[str(forn)]["porUF"] = por_uf

    # Salvar métricas + breakdown por fornecedora (sem registros completos)
    get_db().processamentos.update_one(
        {"_id":f"proc__{ma}__{eq}"},
        {"$set":{
            "_id":f"proc__{ma}__{eq}",
            "mesAno":ma,
            "equipeId":eq,
            "usuarioNome":usuario_nome,
            "valorElegivel":valor_elig,
            "boletosElegiveis":boletos_elig,
            "clientesElegiveis":clientes_elig,
            "fornecedoras":forns,
            "porFornecedora":por_fornecedora,
            "totalRegistros":len(df_num),
            "atualizadoEm":datetime.now()
        }},
        upsert=True)
    try: salvar_historico_processamento(ma,eq,usuario_nome,df)
    except: pass

def buscar_ultimo_processamento(ma, eq):
    doc = get_db().processamentos.find_one(
        {"mesAno":ma,"equipeId":eq},
        {"_id":1,"mesAno":1,"equipeId":1,"valorElegivel":1,"boletosElegiveis":1,
         "clientesElegiveis":1,"fornecedoras":1,"porFornecedora":1,"totalRegistros":1,
         "usuarioNome":1,"atualizadoEm":1,"registros":1}
    )
    if not doc: return {}
    # Se não tem porFornecedora mas tem registros (base antiga), calcular
    if not doc.get("porFornecedora") and doc.get("registros"):
        try:
            df = pd.DataFrame(doc["registros"])
            df["valor"] = pd.to_numeric(df.get("valor", pd.Series(dtype=float)), errors="coerce").fillna(0)
            elig = df[df["elegibilidade"]=="Elegível"] if "elegibilidade" in df.columns else df
            doc["valorElegivel"] = float(elig["valor"].sum())
            doc["boletosElegiveis"] = len(elig)
            doc["clientesElegiveis"] = int(elig["uc_cpf"].nunique()) if "uc_cpf" in elig.columns else 0
            if "fornecedora" in elig.columns:
                pf = {}
                for forn, grp in elig.groupby("fornecedora"):
                    pf[str(forn)] = {"valor": float(grp["valor"].sum()), "boletos": len(grp)}
                doc["porFornecedora"] = pf
        except: pass
    return doc

def buscar_historico_processamentos(ma, eq): return list(get_db().processamentos.find({"mesAno":ma,"equipeId":eq}).sort("criadoEm",-1))
def excluir_processamento(did): get_db().processamentos.delete_one({"_id":did})

def buscar_processamentos(ma=None, eq=None):
    f = {}
    if ma: f["mesAno"]=ma
    if eq: f["equipeId"]=eq
    frames = []
    for d in get_db().processamentos.find(f):
        if d.get("registros"):
            df = pd.DataFrame(d["registros"])
            df["_equipe"]=d["equipeId"]; df["_mes_ano"]=d["mesAno"]
            frames.append(df)
    return pd.concat(frames,ignore_index=True) if frames else pd.DataFrame()

def listar_meses_processados(): return sorted(get_db().processamentos.distinct("mesAno"),reverse=True)

def salvar_senha_usuario(uid, nova_senha):
    get_db().usuarios_senhas.update_one(
        {"_id": uid},
        {"$set": {"_id": uid, "senha": nova_senha, "atualizadoEm": datetime.now()}},
        upsert=True)
    buscar_senha_usuario.clear()

@st.cache_data(ttl=3600)
def buscar_senha_usuario(uid):
    try:
        doc = get_db().usuarios_senhas.find_one({"_id": uid})
        if doc and doc.get("senha"):
            return doc["senha"]
    except: pass
    u = USUARIOS.get(uid)
    if u: return u.get("senha")
    return None

def salvar_historico_processamento(mes_ano, equipe_id, usuario_nome, df):
    doc_id = f"hist_proc__{mes_ano}__{equipe_id}"
    df_num = df.copy()
    df_num["valor"] = pd.to_numeric(df_num.get("valor", pd.Series(dtype=float)), errors="coerce").fillna(0)
    elig = df_num[df_num["elegibilidade"]=="Elegível"] if "elegibilidade" in df_num.columns else df_num
    fornecedoras = sorted(df_num["fornecedora"].dropna().unique().tolist()) if "fornecedora" in df_num.columns else []
    get_db().historico_processamentos.update_one(
        {"_id": doc_id},
        {"$set": {
            "_id": doc_id,
            "mesAno": mes_ano,
            "equipeId": equipe_id,
            "usuarioNome": usuario_nome,
            "fornecedoras": fornecedoras,
            "totalBoletos": len(df_num),
            "boletosElegiveis": len(elig),
            "valorElegivel": float(elig["valor"].sum()),
            "valorTotal": float(df_num["valor"].sum()),
            "criadoEm": datetime.now()
        }},
        upsert=True)

def buscar_historico_geral(mes_ano=None, equipe_id=None):
    filtro = {}
    if mes_ano: filtro["mesAno"] = mes_ano
    if equipe_id: filtro["equipeId"] = equipe_id

    # Buscar histórico permanente (leve)
    try:
        novos = list(get_db().historico_processamentos.find(filtro))
    except: novos = []
    ids_ja_vistos = set(f"{h['mesAno']}__{h['equipeId']}" for h in novos)

    # Buscar processamentos SEM registros (só metadados)
    antigos = []
    try:
        procs = list(get_db().processamentos.find(
            filtro,
            {"_id":1,"mesAno":1,"equipeId":1,"usuarioNome":1,"atualizadoEm":1,
             "valorElegivel":1,"boletosElegiveis":1,"fornecedoras":1}
        ))
        for p in procs:
            chave = f"{p['mesAno']}__{p['equipeId']}"
            if chave in ids_ja_vistos: continue
            antigos.append({
                "_id": p["_id"],
                "mesAno": p.get("mesAno",""),
                "equipeId": p.get("equipeId",""),
                "usuarioNome": p.get("usuarioNome", EQUIPES.get(p.get("equipeId",""),{}).get("nome","—")),
                "fornecedoras": p.get("fornecedoras",[]),
                "totalBoletos": p.get("boletosElegiveis",0),
                "boletosElegiveis": p.get("boletosElegiveis",0),
                "valorElegivel": float(p.get("valorElegivel",0)),
                "criadoEm": p.get("atualizadoEm", datetime.now()),
            })
    except: pass

    todos = novos + antigos
    def _sort_key2(x):
        v = x.get("criadoEm", "")
        if hasattr(v, 'isoformat'): return v.isoformat()
        return str(v or "")
    todos.sort(key=_sort_key2, reverse=True)
    return todos

def salvar_inadimplencia(ma, eq, dados):
    did = f"inadimp__{ma}__{eq}"
    get_db().inadimplencia.update_one({"_id":did},{"$set":{"_id":did,"mesAno":ma,"equipeId":eq,"dados":dados,"atualizadoEm":datetime.now()}},upsert=True)

def processar_base_inadimplencia(arquivo, eq, ma):
    """Processa base de inadimplência e calcula faixas D30/D31-60/D61-90/D90+."""
    import unicodedata
    def norm(s): return unicodedata.normalize('NFKD',str(s).upper().strip()).encode('ascii','ignore').decode()

    try:
        df = ler_arquivo(arquivo)
    except Exception as e:
        return None, f"Erro ao ler arquivo: {e}"

    # Mapear colunas
    cols_norm = {norm(str(c)): c for c in df.columns}
    col_forn = cols_norm.get("FORNECEDORA") or cols_norm.get("FORNECEDOR")
    if not col_forn:
        for c in df.columns:
            if "FORN" in norm(str(c)): col_forn=c; break

    # Mapear colunas pelo nome normalizado (sem espaços, sem acentos)
    # Data Vencimento — exclui "Origi" para não pegar "Data Vencimento Origi"
    col_dvenc = None
    for c in df.columns:
        cn = norm(str(c))
        if "VENCIMENTO" in cn and "ORIG" not in cn:
            col_dvenc = c; break

    # Data Pagamento
    col_dpag = None
    for c in df.columns:
        cn = norm(str(c))
        if "PAGAMENTO" in cn:
            col_dpag = c; break

    # Valor A Pagar — prioriza "Valor A Pagar" sobre outros
    col_val = None
    for c in df.columns:
        cn = norm(str(c))
        if "VALORAPAGAR" in cn or "VALOR A PAGAR" in cn or "VALOR_A_PAGAR" in cn:
            col_val = c; break
    if not col_val:
        for c in df.columns:
            cn = norm(str(c))
            if "VALOR" in cn and c not in [col_forn, col_dvenc, col_dpag]:
                col_val = c; break

    # Debug — mostrar o que foi mapeado
    debug_msg = f"Mapeamento: forn='{col_forn}' | venc='{col_dvenc}' | pag='{col_dpag}' | val='{col_val}'"
    if not col_forn or not col_dvenc or not col_val:
        return None, f"Colunas não encontradas. {debug_msg}. Colunas disponíveis: {list(df.columns)}"

    # Data de referência = último dia do mês selecionado
    import calendar
    try:
        # Formato "Junho-2026"
        partes = ma.split("-")
        if len(partes) == 2:
            mes_nome, ano_str = partes[0].strip(), partes[1].strip()
            if mes_nome.isdigit():
                ano, mes = int(ano_str), int(mes_nome)
            else:
                meses_map = {"Janeiro":1,"Fevereiro":2,"Março":3,"Abril":4,"Maio":5,"Junho":6,
                             "Julho":7,"Agosto":8,"Setembro":9,"Outubro":10,"Novembro":11,"Dezembro":12}
                mes = meses_map.get(mes_nome, 1)
                ano = int(ano_str)
        else:
            return None, f"Formato de mês inválido: {ma}"
    except Exception as e:
        return None, f"Erro ao processar mês {ma}: {e}"
    from datetime import date as _date
    hoje = pd.Timestamp(_date.today())
    ultimo_dia = calendar.monthrange(ano, mes)[1]
    data_ref_mes = pd.Timestamp(ano, mes, ultimo_dia)
    # Se o mês ainda não fechou, usar hoje como data de referência
    data_ref = hoje if hoje < data_ref_mes else data_ref_mes
    # Janela: dia 01 do mesmo mês do ano anterior até data de referência
    data_inicio = pd.Timestamp(ano - 1, mes, 1)

    # Converter datas
    df["_dvenc"] = pd.to_datetime(df[col_dvenc], errors="coerce")
    df["_dpag"] = pd.to_datetime(df[col_dpag], errors="coerce") if col_dpag else pd.NaT
    # Converter valor — aceita vírgula como decimal
    def _conv_valor_inad(s):
        import re
        s = str(s).strip().replace("R$","").replace(" ","")
        # Se tem vírgula como decimal (ex: 1.234,56) -> remover ponto milhar e trocar vírgula
        if re.search(r',\d{1,2}$', s):
            s = s.replace(".","").replace(",",".")
        # Se já é float com ponto (ex: 471.52) -> usar direto, sem remover ponto
        return s
    df["_valor"] = pd.to_numeric(df[col_val].apply(_conv_valor_inad), errors="coerce").fillna(0)
    df["_forn"] = df[col_forn].astype(str).str.strip().str.upper()

    # Detectar coluna Status (PAGO/VENCIDO) se existir
    col_status = None
    for c in df.columns:
        cn = norm(str(c))
        if cn == "STATUS":
            col_status = c; break

    # Aplicar janela de 1 ano
    mask_venc_janela = (df["_dvenc"] >= data_inicio) & (df["_dvenc"] <= data_ref)
    mask_pag_janela = df["_dpag"].notna() & (df["_dpag"] >= data_inicio) & (df["_dpag"] <= data_ref)
    df = df[mask_venc_janela | mask_pag_janela].copy()

    if df.empty:
        return None, "Nenhum registro encontrado dentro da janela de 1 ano."

    # Calcular faixa pela data de vencimento
    df["_dias"] = (data_ref - df["_dvenc"]).dt.days
    def calc_faixa(dias):
        if pd.isna(dias): return "D90+"
        if dias <= 30: return "D0-30"
        if dias <= 60: return "D31-60"
        if dias <= 90: return "D61-90"
        return "D90+"
    df["_faixa"] = df["_dias"].apply(calc_faixa)

    # Pago = pagou ATÉ a data de referência
    df["_pago"] = df["_dpag"].notna() & (df["_dpag"] <= data_ref)

    # Calcular resultado por fornecedora e faixa
    FAIXAS = ["D0-30", "D31-60", "D61-90", "D90+"]
    resultado = {}
    for forn in sorted(df["_forn"].unique()):
        df_f = df[df["_forn"]==forn]
        total_geral = float(df_f["_valor"].sum())
        resultado[forn] = {}
        for faixa in FAIXAS:
            df_fx = df_f[df_f["_faixa"]==faixa]
            pagos = float(df_fx[df_fx["_pago"]]["_valor"].sum())
            vencidos = float(df_fx[~df_fx["_pago"]]["_valor"].sum())
            total_faixa = pagos + vencidos
            # %Faixa = Vencidos / Total da faixa
            pct_faixa = (vencidos/total_faixa*100) if total_faixa>0 else 0
            # %Geral = Vencidos da faixa / Total geral (pagos + vencidos de todas as faixas)
            pct_inad = (vencidos/total_geral*100) if total_geral>0 else 0
            resultado[forn][faixa] = {
                "pagos": pagos,
                "vencidos": vencidos,
                "pct_faixa": pct_faixa,
                "pct_inad": pct_inad
            }

    return resultado, None

def buscar_inadimplencia(ma, eq): return get_db().inadimplencia.find_one({"_id":f"inadimp__{ma}__{eq}"})
def listar_meses_inadimplencia(): return sorted(get_db().inadimplencia.distinct("mesAno"),reverse=True)

# ── HELPERS ────────────────────────────────────
def fmt_brl(v):
    if v is None or v=="": return "R$ 0,00"
    try: return "R$ "+f"{float(v):_.2f}".replace(".",",").replace("_",".")
    except: return "R$ 0,00"

def parse_brl(s):
    """Converte string no formato R$ 1.234.567,89 para float."""
    if s is None: return 0.0
    s = str(s).strip().replace("R$","").replace(" ","").strip()
    # Formato BR: ponto como milhar, vírgula como decimal
    if "," in s:
        s = s.replace(".","").replace(",",".")
    else:
        s = s.replace(".","")
    try: return float(s)
    except: return 0.0

def fmt_brl_td(v):
    if not v or float(v)==0: return "—"
    return "R$ "+f"{float(v):_.2f}".replace(".",",").replace("_",".")

def calc_projecao(v, dt, td):
    if not dt or dt<=0: return 0
    return (v/dt)*td

def calc_variacao(atual, ant):
    if not ant or ant==0: return None
    return ((atual-ant)/ant)*100

def cor_pct(p):
    if p>=80: return "#2daf5c"
    if p>=50: return "#f0a500"
    return "#e03c3c"

def status_pct(p):
    if p>=80: return "Ótimo"
    if p>=50: return "Regular"
    return "Abaixo"

def calc_pontos(media):
    import math
    m=math.floor(media+0.5)
    if m<=70: return 0
    if m<=80: return 300
    if m<=90: return 500
    if m<=95: return 700
    if m<=99: return 1000
    if m>=100: return 1100
    return 0

def calc_media_operador(oid, ma=None):
    monts = buscar_monitorias_operador(oid)
    if ma: monts=[m for m in monts if m.get("mesAno")==ma]
    if not monts: return 0,0
    notas=[m["nota"] for m in monts if "nota" in m]
    if not notas: return 0,0
    media=sum(notas)/len(notas)
    return round(media,1),len(notas)

def get_status_media(media):
    if media==0:   return "Zerada","#e53935","#ffebee"
    if media>=91:  return "Excelente","#2e7d32","#e8f5e9"
    if media>=81:  return "Bom","#1565c0","#e3f2fd"
    if media>=71:  return "Regular","#f57f17","#fff8e1"
    return "Em desenvolvimento","#6d4c41","#efebe9"

def get_iniciais(nome):
    p=nome.strip().split()
    if len(p)>=2: return (p[0][0]+p[1][0]).upper()
    return nome[:2].upper()

CORES_INICIAIS=["#1565c0","#2e7d32","#6a1b9a","#bf360c","#00695c","#4527a0","#ad1457","#0277bd","#558b2f","#4e342e"]
def get_cor_inicial(nome): return CORES_INICIAIS[sum(ord(c) for c in nome)%len(CORES_INICIAIS)]

def get_todos_meses_ano(ano=None):
    if not ano: ano=datetime.now().year
    return [f"{m}-{ano}" for m in MESES_NOMES]

def get_anos_disponiveis():
    hoje=datetime.now()
    return [str(hoje.year),str(hoje.year-1)]

def aging_faixa(dias):
    if pd.isna(dias): return "ND"
    if dias<=30: return "D0-30"
    if dias<=60: return "D31-60"
    if dias<=90: return "D61-90"
    return "D90+"

def header_page(titulo, sub=""):
    st.markdown(f"""
    <div style="background:#ffffff;border:1px solid #c8e0c8;
                border-radius:12px;padding:22px 28px;margin-bottom:24px;
                border-left:4px solid #2e7d32;box-shadow:0 2px 8px rgba(0,0,0,0.06)">
        <h1 style="margin:0;color:#1a2e1a;font-size:20px;font-weight:700">{titulo}</h1>
        {"<p style='color:#5a8a5a;margin:4px 0 0;font-size:12px;text-transform:uppercase;letter-spacing:1px'>"+sub+"</p>" if sub else ""}
    </div>""",unsafe_allow_html=True)

def seletor_equipe(default=None, key_suffix=""):
    u=st.session_state.usuario
    if u["role"] in ["gestor","diretor"]:
        return u["equipe"]
    if u["role"] in ["admin","diretor_upload"]:
        eq_opts=list(EQUIPES.keys())
        eq_labels=[f"Equipe {EQUIPES[e]['nome']}" for e in eq_opts]
        default_idx=eq_opts.index(default) if default and default in eq_opts else 0
        import traceback
        caller=traceback.extract_stack()[-2].name
        sel=st.selectbox("Gerenciando equipe:",eq_labels,index=default_idx,key=f"admin_eq_{caller}{key_suffix}")
        return eq_opts[eq_labels.index(sel)]
    return u["equipe"]

def importar_excel_operadores(arquivo, ops):
    import unicodedata
    def norm_nome(s):
        s = unicodedata.normalize('NFKD', str(s).strip()).encode('ascii','ignore').decode().upper()
        return re.sub(r'\s+', ' ', s).strip()
    def limpar_valor(v):
        s = str(v).strip()
        try:
            f = float(v)
            if f > 0: return f
        except: pass
        s = s.replace('R$','').replace(' ','').strip()
        if ',' in s:
            s = s.replace('.','').replace(',','.')
        else:
            s = s.replace('.','')
        try: return float(s)
        except: return 0.0

    try: df = pd.read_excel(arquivo, header=0)
    except Exception as e: return None, f"Erro ao ler arquivo: {e}"

    col_nome = None
    for c in df.columns:
        cn = norm_nome(str(c))
        if any(x in cn for x in ['OPERADOR','NOME','AGENTE','COLABORADOR','ATENDENTE']):
            col_nome = c; break
    if not col_nome:
        for c in df.columns:
            if df[c].dtype == object:
                col_nome = c; break
    if not col_nome: return None, "Coluna de nome não encontrada"

    col_val = None
    for c in df.columns:
        if c == col_nome: continue
        cn = norm_nome(str(c))
        if any(x in cn for x in ['VALOR','RECEBIDO','PAGO','RECEBIMENTO','RECEBI']):
            col_val = c; break
    if not col_val:
        for c in df.columns:
            if c == col_nome: continue
            try:
                vals = df[c].apply(limpar_valor)
                if vals.sum() > 0: col_val = c; break
            except: pass
    if not col_val: return None, "Coluna de valor não encontrada"

    col_lig = None
    for c in df.columns:
        if c in [col_nome, col_val]: continue
        cn = norm_nome(str(c))
        if any(x in cn for x in ['LIGAC','LIG','CALL','ATEND']):
            col_lig = c; break

    ops_norm = {norm_nome(op['nome']): op for op in ops}
    resultados = []
    for _, row in df.iterrows():
        nome_excel = str(row[col_nome]).strip()
        if not nome_excel or nome_excel.lower() in ['nan','none','']: continue
        valor = limpar_valor(row[col_val])
        ligacoes = int(limpar_valor(row[col_lig])) if col_lig and col_lig in row else 0
        nome_norm = norm_nome(nome_excel)
        primeiro_nome = nome_norm.split()[0] if nome_norm.split() else ''
        match = ops_norm.get(nome_norm)
        status = 'exato'
        if not match:
            matches_pn = [op for n, op in ops_norm.items() if n.startswith(primeiro_nome + ' ') or n == primeiro_nome]
            if len(matches_pn) == 1:
                match = matches_pn[0]; status = 'primeiro_nome'
            elif len(matches_pn) > 1:
                partes = nome_norm.split()
                if len(partes) >= 2:
                    dois_nomes = ' '.join(partes[:2])
                    matches_2n = [op for n, op in ops_norm.items() if n.startswith(dois_nomes)]
                    if len(matches_2n) == 1:
                        match = matches_2n[0]; status = 'dois_nomes'
                    else:
                        status = 'ambiguo'
                else:
                    status = 'ambiguo'
        resultados.append({'nome_excel': nome_excel,'op': match,'valor': valor,'ligacoes': ligacoes,'status': status,'col_lig': col_lig is not None})

    return resultados, None

def get_val_op(ag, oid, onome):
    for k,v in ag.items():
        if isinstance(v,dict) and v.get("nome","").strip().lower()==onome.strip().lower():
            return float(v.get("valorRecebido",0))
    if oid in ag:
        v=ag[oid]
        return float(v.get("valorRecebido",0) if isinstance(v,dict) else v)
    return 0.0

# ── CORREÇÃO PRINCIPAL: normalizar_cpf com zfill(11) ──
def parse_data_inteligente(series):
    """Detecta formato de data automaticamente e converte corretamente."""
    s = series.dropna().astype(str)
    if s.empty: return pd.to_datetime(series, errors="coerce").dt.normalize()
    amostra = s.iloc[0].strip()
    # Formato ISO: YYYY-MM-DD ou YYYY/MM/DD
    if len(amostra) >= 10 and amostra[4] in ["-","/"]:
        return pd.to_datetime(series, dayfirst=False, errors="coerce").dt.normalize()
    # Formato BR: DD/MM/YYYY ou DD-MM-YYYY
    if len(amostra) >= 10 and amostra[2] in ["/","-"]:
        return pd.to_datetime(series, dayfirst=True, errors="coerce").dt.normalize()
    # Fallback
    return pd.to_datetime(series, dayfirst=True, errors="coerce").dt.normalize()

def normalizar_cpf(s):
    s=str(s).strip()
    # Remove formatação primeiro
    s=s.replace(".","").replace("-","").replace("/","").replace(" ","")
    # Remove .0 do final (Excel converte para float)
    if s.endswith(".0"): s=s[:-2]
    # Converte notação científica (ex: 1.23457E+10)
    if "E" in s.upper() or "e" in s:
        try:
            f=float(s)
            if f>0: s=str(int(round(f)))
        except: pass
    # Zfill só se tiver entre 8 e 11 dígitos (CPF válido tem 11)
    if s.isdigit() and 8<=len(s)<=11:
        s=s.zfill(11)
    return s

def gerar_pdf_monitoria(onome, prot, obs, crits, erros, nota, media, n_mon, ma):
    pontos=calc_pontos(media)
    L=[]
    L.append(f"""<!DOCTYPE html><html><head><meta charset='utf-8'><style>
body{{font-family:'Segoe UI',Arial,sans-serif;background:#fff;color:#1a1a1a;margin:0;padding:0}}
.hdr{{background:#0a2414;color:#fff;padding:32px 40px}}.logo{{font-size:24px;font-weight:800;color:#2daf5c}}
.body{{padding:32px 40px}}
.irow{{display:flex;gap:32px;margin-bottom:24px;background:#f8fdf9;border-radius:10px;padding:16px 20px;border-left:4px solid #2daf5c}}
.lbl{{font-size:10px;text-transform:uppercase;letter-spacing:1px;color:#5a9a70;font-weight:600}}
.val{{font-size:15px;font-weight:700;color:#0a2414;margin-top:2px}}
table{{width:100%;border-collapse:collapse;margin-bottom:24px;font-size:13px}}
thead th{{background:#0a2414;color:#fff;padding:10px 14px;text-align:left}}
tbody tr:nth-child(even){{background:#f0f9f3}}
tbody td{{padding:10px 14px;border-bottom:1px solid #e0ede5}}
.ok{{color:#1a6b35;font-weight:700}}.no{{color:#c0392b;font-weight:700}}
.nbox{{background:#0a2414;color:#fff;border-radius:12px;padding:24px;text-align:center;margin-bottom:24px}}
.nnum{{font-size:48px;font-weight:800;color:#2daf5c}}
.mbox{{background:#f0f9f3;border:1px solid #c3e6cb;border-radius:10px;padding:16px 20px;margin-bottom:24px;display:flex;gap:32px}}
.crit{{background:#fdf0f0;border:1px solid #f5c6cb;border-radius:10px;padding:16px 20px;margin-bottom:24px}}
.obs{{background:#f8fdf9;border:1px solid #c3e6cb;border-radius:10px;padding:16px 20px;margin-bottom:24px}}
.foot{{background:#f0f9f3;padding:16px 40px;text-align:center;font-size:11px;color:#5a9a70;border-top:2px solid #2daf5c}}
</style></head><body>
<div class='hdr'><div class='logo'>iGREEN ENERGY</div><div style='font-size:13px;color:#5a9a70;margin-top:4px'>Relatório de Monitoria</div></div>
<div class='body'>
<div class='irow'>
  <div><div class='lbl'>Operador</div><div class='val'>{onome}</div></div>
  <div><div class='lbl'>Protocolo</div><div class='val'>{prot}</div></div>
  <div><div class='lbl'>Mês</div><div class='val'>{ma.replace('-',' ')}</div></div>
  <div><div class='lbl'>Data</div><div class='val'>{datetime.now().strftime('%d/%m/%Y')}</div></div>
</div>""")
    if erros:
        L.append("<div class='crit'><strong>MONITORIA ZERADA — Erro Crítico</strong><br>")
        for e in erros: L.append(f"• {e['nome']}: {e['desc']}<br>")
        L.append("</div>")
    L.append("<table><thead><tr><th>#</th><th>Critério</th><th>Peso</th><th>Resultado</th></tr></thead><tbody>")
    for c in crits:
        p="<span class='ok'>Passou</span>" if c["passou"] else "<span class='no'>Não passou</span>"
        L.append(f"<tr><td>{c['num']}</td><td>{c['nome']}</td><td>{c['peso']}</td><td>{p}</td></tr>")
    L.append("</tbody></table>")
    L.append(f"<div class='nbox'><div style='font-size:13px;color:#5a9a70'>Nota desta Monitoria</div><div class='nnum'>{nota:.0f}%</div></div>")
    L.append(f"<div class='mbox'><div><div class='lbl'>Média ({n_mon} monitorias)</div><div style='font-size:24px;font-weight:800'>{media:.2f}%</div></div><div><div class='lbl'>Pontuação</div><div style='font-size:24px;font-weight:800;color:#1a6b35'>{pontos} pts</div></div></div>")
    if obs: L.append(f"<div class='obs'><strong>Observações:</strong><br>{obs}</div>")
    L.append(f"</div><div class='foot'>iGreen Energy · {datetime.now().strftime('%d/%m/%Y às %H:%M')}</div></body></html>")
    return "".join(L)

# ── LOGIN ──────────────────────────────────────
def tela_login():
    c1,c2,c3=st.columns([1,1.2,1])
    with c2:
        st.markdown("""<div style="background:#003318;border-radius:16px;padding:40px 32px;box-shadow:0 8px 32px rgba(0,0,0,0.3);border:1px solid #005a25">
        <div style="text-align:center;padding:0 0 28px">
            <img src="data:image/webp;base64,UklGRkYXAABXRUJQVlA4IDoXAADw3ACdASorAyoCPm02mkmkIyKhIRHYiIANiWdu4XZ4q+d/YD9AKWV6xPLvANaXYBbv63dj7CjMeT3YnnUdE+dv/k+t79Qewf+qPUu82n7YesJ/wvXp/oPSM6p30ZPLk9qn9r/SZzkT9hO4HaTsFdo32uUFdpPACer2h2JXgh/Q60qeb/3vMD+2b8+OnPnKMUfcmSWxS69yZJbFLr3JklsUuvcmSWxS69yZJbFLr3JklsUuvcmSaQMHI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HIP8hPqfByOQf5CfU+Dkcg/yE+p8HI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HIP8hPqJDhX1Pg5HIP8hPqfByOQf5CfU+Dkcg/yE+p7vd44HbcHI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg4S2SrzhSc8tyKs8p5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HHTUHFqJp0Mmk7brirtASPN+Qn1Pg5HINDOIvLcoPiyayfU+Dkcg/yDwwQ8/qvaIHhdCxa7UetvyE+p8HI5B9eemgRga/S/bjCzL7qefFSfui9Ztu//aISTMbklescqTJTQu7i0lzcTYU7UkHrzhCR5vyE+p7vsmKx4VuTfnIP8hPqfByOOmZikgqldwMfUNzTOBeafG/mn0GLlPnY4eJHfkJ9T4OQ6pewnOp8HI5B/kJ9RBkV2TQc06e3QyLlM1kij7hJZ6iskXQkThVSm8WEB5HF7XdqIseLILaUNp1Pg5HIP8hPqfByOQf5CfKmZGDuQK7mm24Rpc2txUg9ZA4omotF55fCTfAHWVJerL74KScKuTP1qOBTwOy5W0++WNjkH+Qn1Pg5HIP8hPqfBw/gWsoN0UTPCvB8SH0l5voMSAkkggzqfzJxddUeEmZM0XZMnU4POQq+PbAz5d+h7Y71j2uchX+GSfKLmbLZASLBI835CfU+Dkcg/yE+p8Jh9zfbOICMLAl8HKcqdpre9mCZ1dcokC/ZQbVRmgaeZB5zR07wUwnoGoBEm2wtHkH+Qn1Pg5HIP8hPqfByYwKWBRlWTCfbi8MHtqarr7DLNROf/yKXN8kctW2FJvqI6ran6Tazlb4P8hPqfByOQf5CfU+DkxgUeJ04T7jbhdanAT0cACAALBWNlFeYdrw2Nz8HJBSTTHdVy+qRuainYau+ytgLHmHnf/O2VBVDX86VjKGqfNHm/IT6nwcjkH+Qn1Pg5HegUsJ0DWzljJMfgCulAgPVQD7tJ8wGGi6c4k87pgNagiir0F5Ruq2YHWQHg/e/tnP5YmAs08laTYYCZ0WFrgn1Pg5HIP8hPqfByOQf5DZWM4yhae5Y7EyvMSXE+FGbVaEWJs38KNiCbViMsYQEWtbwrsXuIwqclcwMUTE4vVU6KwAdrQ1DTnjoEK+6j/WPaZ4nVB0MymvLYWjyD/IT6nwcjkH+Qn1Pg5MYHHNyo64kswWvtw3HhMGQIpQbjGn+gz4KvAk4VbXfKt5zv//5qCYR2/AcyOUZTnKwtjpU6fLY26Zq/uCm5WkyjdJ/UT6nwcjkH+Qn1Pg5HIP8oKnxAe+VW9NI3lZNnw+atFzTpr+G23CB+v1ZjXux1w4tmxH7Iapyciea0ALhqYdlRyOQf5CfU+DkchBYWyH3Zj+LGWFBJH+CkBVaBjd849NCm0CaJucKHTKcHCjB5n7yDkcg/yE+p8HI5B/kJ9T5XVfSdRxPwOa9ng4TGVD2s2CHMWC1DJkkyxfMxR7HKdq/6F/C1FMImeEvg5HIP8hPqfByOQf5CfU7K/y9kgwcT1TnNPRyplc5QHr5xG7uH8zl20fQoCsn65fILMzmmP8cqUvtDen1Pg5HIP8hPqfByOQf5CcxhhTPFChP0lu8isbgG7r/buSfEky46TTR+ZzMypX5CfU+Dkcg/yE+p8HI5B9j9Vgo1xv9BZFQt/27kiP5lHYc086i6nvrtoYKZuSlSNZYQEjzfkJ9T4ORyD/IT6nwcNnIOFlwJJFTH6PNKLLPva//aISTMbko7n3AuXQI7hfLidBU9KQnyn1f9LuSPN+Qn1Pg5HIP8hPqfByOQfgL+aWxeQNtSs0OsGCfU+Dkcg/yE+p8HI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HIP8hPqfByOQf5CfU+Dkcg/yE+p8HI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HIP8hPqfByOQf5CfU+Dkcg/yE+p8HI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HIP8hPqfByOQf5CfU+Dkcg/yE+p8HI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HIP8hPqfByOQf5CfU+Dkcg/yE+p8HI5B/kJ9T4ORyD/IT6nwcjkH+Qn1EAAA/v/Ba9THJIzmnUSVZEgeVDPuL3enkDK8eokqyJA8qGfcXu9PIGV49RJVkSB5UM+4vd6eQMrx6iSrIkDyoZ9xe708gZXj1ElWRIHlQz7i93p5AyvHqJFZa+O0xaYf/QFkV2PTk+RCqIiW7PXAGVx5ZpNHFQLbz/PhdsVosmLLo8Urhnzum/njXjDVDf5SoXdosh4LTbPvXCdW1ccKV4687K/mBOVrBEnkfNEp9O8yFMjR7Fdtp4NdxUjD0Nw/k6fjkxkrPuwSrrPKAc8RN/FtIpqjkmzn6/pjvUz3Enxmh2HQ4iHvBgVB5X9H+cdUyKI33JhxBvhLu4mN9z57hHo8MxU5EY54L9FcSrFG8eppWmBdhXp+mYqTU/zYjpcsoU5rDEOz16f7YiEQiEB8gt/IJDU7dEpcOBf9ELkvXsn9n6Hfnq+2ltfUXeauNPaFOZ8SvpNCc82w2NiNwQqbOvenGNs7pOEvkT3XVYZCzmf58u+iSM4Cb/JrGePxv2J8WNc7zeYX1hS1O4QOaCfIavBCGpzT2IszQyPtOe59/K0mOh/E4Yw2ToJ8JiG9hhE8X+R2xNlRZVjIlrBXv+adtm0Gn/B/E+8vzgNunRdvtVFyXvV0vCw/ZohAvkDSta0MdjZROS9VrzXhIV4N44wY9cwQGnwmE3nZORI5oYWDV6WWKVk+Y2W7EY7/Mfq3ZaktH8WVtvmK0JJSVfN4AcKRtVWgDed2B00OQFePmxudFLiZjKHUnOQ1oS1phHI3KEn/Z+WKMtYZEyKzEDtr7kqpcTxZhe/XCXOFl30A1vjMEylLWqm3zjc5f773os9HWaJL77rP+F/K8z+UJ8iOfkRGewe8vq35h/nWybbtSlQHFR/BqmjzkT5rdPSkyynpzG3v9P/953I3m0s5rYcGQ8UEqx8viZIOX4x1Nvrlr3pacf25ZTaK2fVI229yS3Nllx+cs4U5ewqj/udeCnz2EyNlmhvcmma4p2Cag4D574GextSi8RL591k7aJ3eFcUYcldl20PO88Pp3HVvl2R//O9QMiSoGe/oo5DEs7khUjtn29oB0o0qzzLB7JfdSZO3iadLoElYTlHwXjEknnyAr6L0ISW+b+5l9JzfT9QCRJP8aK5Yp1jxg4yLVnBbz54BJTnpzrH3KhMdb/AUrAXRwNCu/yuNVJQZ5lHzB+abtfXu+nMF+ldfTzp6PXNyBvDFlzlkrEdmyGMpM2N93TuY+qmGmxw+670GUaoeQ/aIBBTjl3HgXLZaVOadXUHBKiOZZo4lB8yFGXECw62W+IpSn2XAjUNXkirDoWWJOTFaMV2OVCjqUWqhaWy8yaLYln4d/lkf/wpWuhpSQAGdu8Tyx4czQNCVGw2b1CpkBS5NE5n/59Mv3PQrJ/x6tR3OTLzpoH/fqehcnSKO46zd35h8GVEYtxT6OG1GOOoPKJJ+ADQhuTb0r7XdNiixSr2q8EWqMLVZbFS+Iih9UNY5bKsYdgvS+HqgLMDSKkQ7EM3i03LULV3Ds7zHa37eDxvlU+CmD9pf54YB6km+IokM3WDd5cF8z0sMUqSluZo7upZ88OfL2vIiJ4lKL4KV4jOlM7R/a6Q0hSONgHfcODaiLYBBqdzGibYz3g4QrQ0G/GGblQq/HNlCJ2e3yRke5FZwfZGiqEPJ1S6xrrI5V3yQix3w5IaunTWwlt7td5FnJLXFXiDAYXxXPYHJyIe1deXBQNJijhej2oAUB44T1JAChHQinM1kbO8GM57NivWfHE5mkBd023gyuZ6fYKLlqoL7SDVWyiTiDL0uLmgpA83dJTng+TAlyo4EgGWOy2KJEGTbWALf8eH1jTFxI6NbzRtEHmNeT4+Kgz0TKxp7Np4fO150jjWRcSUfL6kG3/8Ag5gUuxLBNo9BJvA4RFWi/ZBgY2E16JAO0TjgV6x4LMyYqer9Fqk73KIQQcJo8DBKz4FZCfautquuR7p/xSc9asvgViUHLCVQ66wcGAmcRlhpPlGsPxC5bP8sF9oCFgGOWuou5zVY0epjQuwtn9XP2c19dIHP2l2nMVshHecyFRMdZdX70WpUfIM4kMitXwJBiIfjefC6JEcfP6XutHw9FWF/Rt8YSf+GiHt3wWdRFCN2J2U5b3jUxy2EbmQXCa4wte8g8lxRbP8GVDSEIggvUgvFGxTjmJ104wL8/Lhv5+1cSaIiV+J6xuiSITVkz2CB902XuhLkLNUa70MiNlD5ltyATjhFHSZ3mkAaYNPY2bhsdm9NxYhZVaCL9qjk0HnY5sLn0Xab/nOeZtBXxXRT55iYAervHxTqIUktipkDFl4Wy8yfXTSNGGjgquI4KAZOR+buoAnrmXPDP8RA8XqVbMk9IriccQScd/KAKRV4CnPb9Q5PaIV9z3urT7cYeKESK4or/z3R0/79xo01FZoa8ZfNu7j5xuu3C1n7PetqFN2y2l/g8LHjeALR5c5pvqR6/zI/X8UaZ6nj/59VWsV1A0oSlhzV1y6K8AfxMY80IO8N+ebOZ8JvFwIUxSpOym4qTKTiQbwcTm90cioBJn//f/pf9xb/2Y6l3QOaAP2jjhVlX2ocbrnSB6AZDsaFeu9Y/cbqocX2wcsNWOvwCfQj8HiHDJIb/bC484oGWWuq7JoDdiqsjtalx3sVPQzxBuyvkEk/fCILTWgbsvW7spWct3neeD+zGdKEgoeCICovCU/R34qzEL1uVOQaLjeZ0GUoOnoryWQ+c6ITxO35IDh+qMr2Pi2hWtUS5nlgXH/tTFu0LndEs0mw0V+6mT9cZcWvI4uZ5uRyBOWITynzmFKGfYmtcCv1yxu5chaU92vIOu4Dmd+AcQgkfmzCQX935zcNk90hzR3cK2WKKnlhgZEiLgoV8/Midnlfq7juUbamV15AcQqQqAdJBFebKva4e4QIiQA0R5SCDsxn8qeNK0RyJCtszmmnSfQzenSD200SL0kBBqui91eTeBFB2/mXdQoZAyfZ1SRSRZO3jcJSuBltnLjnrLPGMAUjjplqSRysZhWbk/VoK3vg15rmDTJJqRvC0V30DkotzSM4qU8YpnZwEnUiUVjVsOJgGZjsqNi+FARghVhwdxmlsAd1aPqB1EJ9NqCdeTLuU5yztWWMATz32Vvc3VdCMOUW70EGdcy59JjR6L83vIsHptjcl3HDzr3ovDtV/uLkHyeUbiCgF4avSz4/5yJ3uLFrAUdUGFxdc+WIT5cLuK16nWp33VuJ8VTbD6SiU3osiyUeUnvttihtwtpU3UVmZuEUQqV7MbkX6k4TN6JqVwR1ln4ZIXTRW7VhPcLVvyXQ7kg2qW0+b98+n4Gg+L5eXs+2NhO20zTGnAsBb99AygxbjQmTVCm25NDWFIhoMa04Z9bXN89EqO2M2X3jWmnXXZ1m2EYmiv6gCBfftD0un3QSkD2u7xjakw2I9hwEad+oQdP0oUIB44s0VjE/ZQmKrOQs0WiavE4xVfFOYJQdAwtcGHKxKOlec3LBhnhHQvM5W4FTVor1O/LcI9fs5fZ/SZ7WE4hNNfNL/EPWT8nH50ODv9n7ScRYO0cvFw6LFT/YAroFLseGDru7jBMz0fisqnLL1lQtopT6w9GDuizE+hEVQd9qVN6uN7YjudPYEvchw3hvgMxHGWN+tSPju4E8MazGH+r6QMFSLCZGCM94PChlHfUzb7WTJ7j4ua0doYUGJdGAxXGqcWYdDZ0Kk4TrDXWkxzpNNrlgyxkK+8tMXyqClfL8LL266E23yIM603GtG7TNUiH9m03aEH4YNjOO8H1bBBxbh6jyO5D//gDLBVjRTrMP3KKOIBFoCipYdULnpZRbPvl58PldwURixOQ8utLtFxOaBLMjJ1FPOly0csbmoOqpCQwzN1s6RX+jUn4FFkh3JkjBLetI+TrU8PKi2tGC6qAhENCiMImkCg/oD7Yspi9Se0/7JGmUBFU4N96T1eHKoeI8kF0sHCjXd/ZZoKJDHssOcgWZzklYpP6CpwGuTMVtuSnzvkoBPhtk8YnIKR8GwzMWpVFvB44nSNzoDJ3NcJUh79AglfccmReXjZOBg+F/kJAkADhw57QADQGElQ6OItrrDnOWMvMCk925i5CNgRrnhugCgRXrXI5YP9uoaDVEovMJzh26gSmnuG8s3cdYVrPBtRt9eldx2IHWpXO8n+7/LM2WH3YIs6LoHdygmAW3Kkk/vjUvVj/qMOaXjiAON6Oq3lEqkCKx4UgyoULMxPgWevNBm6osmtQZYOt93NHEVI5u85IfjKkJF9DB+DVgWSoYubRUZtzpbL7I5R1RurzcS/REwnnCRlRMQc/qabrbpLlo6KLArK0vQ8kekHjO4Rxa8maH6djlrA482uwPVhAyCwY0L4JNcoDt9mbaUpKRB01kJ7Uq3o2E9V9BXIqDhKQ/9szos87vr/ChzzpqnDM79D1FMui4T4KRX/iZQAW9jKyNIPXPzLQM/pHG5fyKS1xTvZ4KFukCf5MCsyWw9LxQJDSYJhMjBLpdQSNi06HVBHj/ZY5mAzZo10VcOwSGN76k1g7d5jYv999BV+e1Ogx9SLPhDLN0PdadiozEAuqAAAkwYuddZ3ZvB9CDIrohL89ypfdWbcTiWcpFMdlpSSEvPUYhYUA2dgxdQTKPPfQ+K/LuxMNuN5tWL/PED5ouexw3X3PqeimOr3bwxPvh9z4KZxP6akCOTy6AzPsklkgJk3eugTs881gEhiVhPNAPrO2BLsFXTj6yg5iRa5NlgXI+a9ITIAIL/cYMfdudyw0NK6mniI0Twiuussv+D5z0kyt5kHhJ6vdX5rPLhG7wJvbKrX036jvapmbxQL2+QthagbUQtbzYy7rnvHu7sUe1f0UlCrRM0AEYR3szCi2kmR6NSPHFdRBqG3YWDRbuUdovmzbXMuOqwuGFvUbXKq1PvqipGXd0yYlHflQFhllVr0/+SHdYJk7X1Uor14ksnWk+IOan5xlI2j3/RlKzN3JGr5mvLAhJOHqYpkc3LUA/ExeVbt/QHoeL4+Hb6u+7jCZ3v9EJc7CrGTmXB3pn9uatGHqB4dWWzJTtwqKygrUWLuYLiEpgDJK9p/y39ORq9t4B8u3/0iD1zAiW37RBJEj+3IJ3k6XiOzj9UAx2rKs2rbYXvLHgwnhSbAhdrvJCNw6rNSLKAZAelPnZ2lQIceY17VNW3PBL8GJttheN77lSu7ykybM3+lMJimCmWViqXsEPIfGpgCrsIeMTsVqHx8hBxPGYlcaOND5ITi+mBtEau/SpqAxEBLnX6cbggOxdgcxldZ0HU91lHsj4wbB+glZTzD0txBq0/0uwLMDBInV3Vp2Z7Lojsch8ctus4U4nV4nNGk/sRnTws3C8dgvppMrHSCSmDjcBpATeezwyOXVy4OkzG365HpR4/vx9EeZnG4Xcug/pES0hIufbKaqpDeDPLEdTLKouiqrf5tM47VV30RhHCXZzFIueoBi16XhCHI/PRkMEqMxFsfPaP+qmT/YiFRF6GwAsNcjy4/+s+PA7ITB0DsE4srhSUpkbK+91WtkapOIvnN8l/0BfOb53/QIBzfJf9AX8pTEiAAA=" style="width:72px;height:72px;border-radius:16px;object-fit:cover;margin-bottom:14px;box-shadow:0 4px 16px rgba(0,0,0,0.2)"/>
            <div style="font-size:24px;font-weight:800;color:#ffffff;margin-bottom:4px">iGreen Performance</div>
            <div style="width:36px;height:2px;background:#00c853;margin:6px auto 10px"></div>
            <p style="color:#5a9a70;font-size:11px;text-transform:uppercase;letter-spacing:2px;margin:0">Painel de Gestão de Inadimplência</p>
        </div>""",unsafe_allow_html=True)
        st.markdown("<p style='font-size:11px;text-transform:uppercase;letter-spacing:1px;color:#5a9a70;margin-bottom:4px'>USUÁRIO</p>",unsafe_allow_html=True)
        usuario=st.text_input("u",placeholder="seu usuário",label_visibility="collapsed")
        st.markdown("<p style='font-size:11px;text-transform:uppercase;letter-spacing:1px;color:#5a9a70;margin-bottom:4px;margin-top:12px'>SENHA</p>",unsafe_allow_html=True)
        senha=st.text_input("s",type="password",placeholder="••••••••",label_visibility="collapsed")
        st.markdown("<div style='height:8px'></div>",unsafe_allow_html=True)
        if st.button("Entrar",use_container_width=True):
            uid = usuario.lower().strip()
            u = USUARIOS.get(uid)
            if u:
                senha_correta = buscar_senha_usuario(uid) or u.get("senha")
                if senha_correta and senha.strip() == senha_correta:
                    st.session_state.usuario={"id":uid,**u}; st.rerun()
                else: st.error("Usuário ou senha incorretos.")
            else: st.error("Usuário ou senha incorretos.")
        st.markdown('<p style="text-align:center;color:#1a4d2e;font-size:11px;margin-top:24px">iGreen Energy © 2026</p>',unsafe_allow_html=True)

# ── SIDEBAR ────────────────────────────────────
def _mini_operadores(u):
    eq = u.get("equipe")
    if not eq:
        st.info("Disponível apenas para gestores.")
        return
    ops = buscar_operadores(eq)
    st.markdown(f"<p style='font-size:11px;color:#5a9a70;margin-bottom:8px'>{len(ops)} operadores — Equipe {EQUIPES[eq]['nome']}</p>",unsafe_allow_html=True)
    nn = st.text_input("Nome do novo operador",placeholder="Nome completo",key="mc_op_nome")
    np = st.checkbox("Pleno",key="mc_op_pleno")
    if st.button("Adicionar",use_container_width=True,key="mc_op_add"):
        if nn.strip(): salvar_operador(eq,nn.strip(),np); st.success(f"{nn} adicionado!"); st.rerun()
        else: st.error("Digite o nome.")
    if ops:
        st.markdown("---")
        for op in ops:
            c1,c2,c3 = st.columns([3,1,1])
            with c1: ne = st.text_input("",value=op["nome"],key=f"mc_n_{op['_id']}",label_visibility="collapsed")
            with c2:
                if st.button("V",key=f"mc_s_{op['_id']}"):
                    atualizar_operador(op["_id"],ne,op.get("pleno",False)); st.rerun()
            with c3:
                if st.button("X",key=f"mc_d_{op['_id']}"):
                    excluir_operador(op["_id"]); st.rerun()

def _mini_criterios():
    crits = get_criterios()
    ce = []
    for i,c in enumerate(crits):
        with st.expander(f"{c['num']} {c['nome']}",expanded=False):
            nm = st.text_input("Nome",value=c["nome"],key=f"mcc_n_{i}")
            ps = st.number_input("Peso",min_value=1,max_value=100,value=int(c["peso"]),key=f"mcc_p_{i}")
            ob = st.checkbox("Obrigatório",value=c.get("obrigatorio",False),key=f"mcc_o_{i}")
            it = st.text_area("Itens",value="\n".join(c.get("itens",[])),height=80,key=f"mcc_i_{i}")
            ce.append({"id":c["id"],"num":c["num"],"nome":nm,"peso":ps,"obrigatorio":ob,"itens":[x.strip() for x in it.split("\n") if x.strip()]})
    if st.button("Salvar Critérios",use_container_width=True,key="mc_crit_save"):
        salvar_criterios(ce); st.success("Salvo!"); st.rerun()

DARK_CSS = """
.stApp { background-color: #0d1117 !important; --text-color: #e6edf3; --text-muted: #8b949e; --bg-card: #161b22; --border-color: #30363d; }
[data-testid="stSidebar"] { background: #161b22 !important; border-right: 1px solid #30363d !important; }
[data-testid="stSidebar"] .stRadio label { color: #e6edf3 !important; background: #21262d !important; border: 1px solid #30363d !important; font-weight: 500 !important; }
[data-testid="stSidebar"] .stRadio label:hover { background: #2d333b !important; color: #ffffff !important; border-color: #3fb950 !important; }
[data-testid="stSidebar"] .stRadio label[data-checked="true"] { background: #238636 !important; color: #ffffff !important; border-color: #3fb950 !important; font-weight: 600 !important; }
[data-testid="stMetric"] { background: #161b22 !important; border: 1px solid #30363d !important; border-top: 2px solid #3fb950 !important; }
[data-testid="stMetricValue"] { color: #e6edf3 !important; font-size: 16px !important; white-space: nowrap !important; overflow: visible !important; }
[data-testid="stMetricLabel"] { color: #7ee787 !important; }
.stButton > button { background: #21262d !important; color: #7ee787 !important; border: 1px solid #30363d !important; }
.stButton > button:hover { background: #238636 !important; color: #ffffff !important; border-color: #3fb950 !important; }
h1 { color: #e6edf3 !important; } h2 { color: #c9d1d9 !important; } p { color: #8b949e !important; }
hr { border-top: 1px solid #30363d !important; }
.stTextInput input, .stNumberInput input, .stTextArea textarea { background: #0d1117 !important; border: 1px solid #30363d !important; color: #e6edf3 !important; }
.stSelectbox > div > div { background: #161b22 !important; border: 1px solid #30363d !important; color: #e6edf3 !important; }
[data-baseweb="select"] { background: #161b22 !important; }
[data-baseweb="select"] > div { background: #161b22 !important; color: #e6edf3 !important; }
[data-baseweb="popover"] { background: #161b22 !important; border: 1px solid #30363d !important; }
[role="option"] { background: #161b22 !important; color: #e6edf3 !important; }
[role="option"]:hover { background: #21262d !important; }
.stTabs [data-baseweb="tab-list"] { background: #161b22 !important; border: 1px solid #30363d !important; }
.stTabs [data-baseweb="tab"] { color: #8b949e !important; }
.stTabs [aria-selected="true"] { background: #238636 !important; color: #ffffff !important; }
.stCheckbox label { color: #c9d1d9 !important; }
[data-testid="stFileUploader"] > div { background: #161b22 !important; border: 1.5px dashed #30363d !important; }
.stSuccess > div { background: #0f2117 !important; border-left: 3px solid #3fb950 !important; color: #7ee787 !important; }
.stError > div { background: #2d1117 !important; border-left: 3px solid #f85149 !important; color: #ffa198 !important; }
.stWarning > div { background: #271a00 !important; border-left: 3px solid #d29922 !important; color: #e3b341 !important; }
.stInfo > div { background: #0c2d6b !important; border-left: 3px solid #388bfd !important; color: #79c0ff !important; }
[data-testid="stDataFrame"] { border: 1px solid #30363d !important; background: #161b22 !important; }
[data-testid="stDataFrame"] * { color: #e6edf3 !important; }
.streamlit-expanderHeader { background: #161b22 !important; border: 1px solid #30363d !important; color: #c9d1d9 !important; }
.streamlit-expanderContent { background: #0d1117 !important; border: 1px solid #30363d !important; }
div[data-testid="stVerticalBlock"] label { color: #8b949e !important; }
.block-container { background: #0d1117 !important; }
"""

def render_sidebar():
    u=st.session_state.usuario
    role_label='Administrador' if u['role']=='admin' else 'Diretoria' if u['role']=='diretor' else 'Gestor'
    with st.sidebar:
        st.markdown(
            f"<div style='padding:16px 12px 8px'>"
            f"<div style='display:flex;align-items:center;gap:10px;margin-bottom:16px'>"
            f"<div style='width:34px;height:34px;background:#2e7d32;border-radius:8px;"
            f"display:flex;align-items:center;justify-content:center;font-weight:900;font-size:16px;color:#fff'><img src='data:image/webp;base64,UklGRkYXAABXRUJQVlA4IDoXAADw3ACdASorAyoCPm02mkmkIyKhIRHYiIANiWdu4XZ4q+d/YD9AKWV6xPLvANaXYBbv63dj7CjMeT3YnnUdE+dv/k+t79Qewf+qPUu82n7YesJ/wvXp/oPSM6p30ZPLk9qn9r/SZzkT9hO4HaTsFdo32uUFdpPACer2h2JXgh/Q60qeb/3vMD+2b8+OnPnKMUfcmSWxS69yZJbFLr3JklsUuvcmSWxS69yZJbFLr3JklsUuvcmSaQMHI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HIP8hPqfByOQf5CfU+Dkcg/yE+p8HI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HIP8hPqJDhX1Pg5HIP8hPqfByOQf5CfU+Dkcg/yE+p7vd44HbcHI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg4S2SrzhSc8tyKs8p5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HHTUHFqJp0Mmk7brirtASPN+Qn1Pg5HINDOIvLcoPiyayfU+Dkcg/yDwwQ8/qvaIHhdCxa7UetvyE+p8HI5B9eemgRga/S/bjCzL7qefFSfui9Ztu//aISTMbklescqTJTQu7i0lzcTYU7UkHrzhCR5vyE+p7vsmKx4VuTfnIP8hPqfByOOmZikgqldwMfUNzTOBeafG/mn0GLlPnY4eJHfkJ9T4OQ6pewnOp8HI5B/kJ9RBkV2TQc06e3QyLlM1kij7hJZ6iskXQkThVSm8WEB5HF7XdqIseLILaUNp1Pg5HIP8hPqfByOQf5CfKmZGDuQK7mm24Rpc2txUg9ZA4omotF55fCTfAHWVJerL74KScKuTP1qOBTwOy5W0++WNjkH+Qn1Pg5HIP8hPqfBw/gWsoN0UTPCvB8SH0l5voMSAkkggzqfzJxddUeEmZM0XZMnU4POQq+PbAz5d+h7Y71j2uchX+GSfKLmbLZASLBI835CfU+Dkcg/yE+p8Jh9zfbOICMLAl8HKcqdpre9mCZ1dcokC/ZQbVRmgaeZB5zR07wUwnoGoBEm2wtHkH+Qn1Pg5HIP8hPqfByYwKWBRlWTCfbi8MHtqarr7DLNROf/yKXN8kctW2FJvqI6ran6Tazlb4P8hPqfByOQf5CfU+DkxgUeJ04T7jbhdanAT0cACAALBWNlFeYdrw2Nz8HJBSTTHdVy+qRuainYau+ytgLHmHnf/O2VBVDX86VjKGqfNHm/IT6nwcjkH+Qn1Pg5HegUsJ0DWzljJMfgCulAgPVQD7tJ8wGGi6c4k87pgNagiir0F5Ruq2YHWQHg/e/tnP5YmAs08laTYYCZ0WFrgn1Pg5HIP8hPqfByOQf5DZWM4yhae5Y7EyvMSXE+FGbVaEWJs38KNiCbViMsYQEWtbwrsXuIwqclcwMUTE4vVU6KwAdrQ1DTnjoEK+6j/WPaZ4nVB0MymvLYWjyD/IT6nwcjkH+Qn1Pg5MYHHNyo64kswWvtw3HhMGQIpQbjGn+gz4KvAk4VbXfKt5zv//5qCYR2/AcyOUZTnKwtjpU6fLY26Zq/uCm5WkyjdJ/UT6nwcjkH+Qn1Pg5HIP8oKnxAe+VW9NI3lZNnw+atFzTpr+G23CB+v1ZjXux1w4tmxH7Iapyciea0ALhqYdlRyOQf5CfU+DkchBYWyH3Zj+LGWFBJH+CkBVaBjd849NCm0CaJucKHTKcHCjB5n7yDkcg/yE+p8HI5B/kJ9T5XVfSdRxPwOa9ng4TGVD2s2CHMWC1DJkkyxfMxR7HKdq/6F/C1FMImeEvg5HIP8hPqfByOQf5CfU7K/y9kgwcT1TnNPRyplc5QHr5xG7uH8zl20fQoCsn65fILMzmmP8cqUvtDen1Pg5HIP8hPqfByOQf5CcxhhTPFChP0lu8isbgG7r/buSfEky46TTR+ZzMypX5CfU+Dkcg/yE+p8HI5B9j9Vgo1xv9BZFQt/27kiP5lHYc086i6nvrtoYKZuSlSNZYQEjzfkJ9T4ORyD/IT6nwcNnIOFlwJJFTH6PNKLLPva//aISTMbko7n3AuXQI7hfLidBU9KQnyn1f9LuSPN+Qn1Pg5HIP8hPqfByOQfgL+aWxeQNtSs0OsGCfU+Dkcg/yE+p8HI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HIP8hPqfByOQf5CfU+Dkcg/yE+p8HI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HIP8hPqfByOQf5CfU+Dkcg/yE+p8HI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HIP8hPqfByOQf5CfU+Dkcg/yE+p8HI5B/kJ9T4ORyD/IT6nwcjkH+Qn1Pg5HIP8hPqfByOQf5CfU+Dkcg/yE+p8HI5B/kJ9T4ORyD/IT6nwcjkH+Qn1EAAA/v/Ba9THJIzmnUSVZEgeVDPuL3enkDK8eokqyJA8qGfcXu9PIGV49RJVkSB5UM+4vd6eQMrx6iSrIkDyoZ9xe708gZXj1ElWRIHlQz7i93p5AyvHqJFZa+O0xaYf/QFkV2PTk+RCqIiW7PXAGVx5ZpNHFQLbz/PhdsVosmLLo8Urhnzum/njXjDVDf5SoXdosh4LTbPvXCdW1ccKV4687K/mBOVrBEnkfNEp9O8yFMjR7Fdtp4NdxUjD0Nw/k6fjkxkrPuwSrrPKAc8RN/FtIpqjkmzn6/pjvUz3Enxmh2HQ4iHvBgVB5X9H+cdUyKI33JhxBvhLu4mN9z57hHo8MxU5EY54L9FcSrFG8eppWmBdhXp+mYqTU/zYjpcsoU5rDEOz16f7YiEQiEB8gt/IJDU7dEpcOBf9ELkvXsn9n6Hfnq+2ltfUXeauNPaFOZ8SvpNCc82w2NiNwQqbOvenGNs7pOEvkT3XVYZCzmf58u+iSM4Cb/JrGePxv2J8WNc7zeYX1hS1O4QOaCfIavBCGpzT2IszQyPtOe59/K0mOh/E4Yw2ToJ8JiG9hhE8X+R2xNlRZVjIlrBXv+adtm0Gn/B/E+8vzgNunRdvtVFyXvV0vCw/ZohAvkDSta0MdjZROS9VrzXhIV4N44wY9cwQGnwmE3nZORI5oYWDV6WWKVk+Y2W7EY7/Mfq3ZaktH8WVtvmK0JJSVfN4AcKRtVWgDed2B00OQFePmxudFLiZjKHUnOQ1oS1phHI3KEn/Z+WKMtYZEyKzEDtr7kqpcTxZhe/XCXOFl30A1vjMEylLWqm3zjc5f773os9HWaJL77rP+F/K8z+UJ8iOfkRGewe8vq35h/nWybbtSlQHFR/BqmjzkT5rdPSkyynpzG3v9P/953I3m0s5rYcGQ8UEqx8viZIOX4x1Nvrlr3pacf25ZTaK2fVI229yS3Nllx+cs4U5ewqj/udeCnz2EyNlmhvcmma4p2Cag4D574GextSi8RL591k7aJ3eFcUYcldl20PO88Pp3HVvl2R//O9QMiSoGe/oo5DEs7khUjtn29oB0o0qzzLB7JfdSZO3iadLoElYTlHwXjEknnyAr6L0ISW+b+5l9JzfT9QCRJP8aK5Yp1jxg4yLVnBbz54BJTnpzrH3KhMdb/AUrAXRwNCu/yuNVJQZ5lHzB+abtfXu+nMF+ldfTzp6PXNyBvDFlzlkrEdmyGMpM2N93TuY+qmGmxw+670GUaoeQ/aIBBTjl3HgXLZaVOadXUHBKiOZZo4lB8yFGXECw62W+IpSn2XAjUNXkirDoWWJOTFaMV2OVCjqUWqhaWy8yaLYln4d/lkf/wpWuhpSQAGdu8Tyx4czQNCVGw2b1CpkBS5NE5n/59Mv3PQrJ/x6tR3OTLzpoH/fqehcnSKO46zd35h8GVEYtxT6OG1GOOoPKJJ+ADQhuTb0r7XdNiixSr2q8EWqMLVZbFS+Iih9UNY5bKsYdgvS+HqgLMDSKkQ7EM3i03LULV3Ds7zHa37eDxvlU+CmD9pf54YB6km+IokM3WDd5cF8z0sMUqSluZo7upZ88OfL2vIiJ4lKL4KV4jOlM7R/a6Q0hSONgHfcODaiLYBBqdzGibYz3g4QrQ0G/GGblQq/HNlCJ2e3yRke5FZwfZGiqEPJ1S6xrrI5V3yQix3w5IaunTWwlt7td5FnJLXFXiDAYXxXPYHJyIe1deXBQNJijhej2oAUB44T1JAChHQinM1kbO8GM57NivWfHE5mkBd023gyuZ6fYKLlqoL7SDVWyiTiDL0uLmgpA83dJTng+TAlyo4EgGWOy2KJEGTbWALf8eH1jTFxI6NbzRtEHmNeT4+Kgz0TKxp7Np4fO150jjWRcSUfL6kG3/8Ag5gUuxLBNo9BJvA4RFWi/ZBgY2E16JAO0TjgV6x4LMyYqer9Fqk73KIQQcJo8DBKz4FZCfautquuR7p/xSc9asvgViUHLCVQ66wcGAmcRlhpPlGsPxC5bP8sF9oCFgGOWuou5zVY0epjQuwtn9XP2c19dIHP2l2nMVshHecyFRMdZdX70WpUfIM4kMitXwJBiIfjefC6JEcfP6XutHw9FWF/Rt8YSf+GiHt3wWdRFCN2J2U5b3jUxy2EbmQXCa4wte8g8lxRbP8GVDSEIggvUgvFGxTjmJ104wL8/Lhv5+1cSaIiV+J6xuiSITVkz2CB902XuhLkLNUa70MiNlD5ltyATjhFHSZ3mkAaYNPY2bhsdm9NxYhZVaCL9qjk0HnY5sLn0Xab/nOeZtBXxXRT55iYAervHxTqIUktipkDFl4Wy8yfXTSNGGjgquI4KAZOR+buoAnrmXPDP8RA8XqVbMk9IriccQScd/KAKRV4CnPb9Q5PaIV9z3urT7cYeKESK4or/z3R0/79xo01FZoa8ZfNu7j5xuu3C1n7PetqFN2y2l/g8LHjeALR5c5pvqR6/zI/X8UaZ6nj/59VWsV1A0oSlhzV1y6K8AfxMY80IO8N+ebOZ8JvFwIUxSpOym4qTKTiQbwcTm90cioBJn//f/pf9xb/2Y6l3QOaAP2jjhVlX2ocbrnSB6AZDsaFeu9Y/cbqocX2wcsNWOvwCfQj8HiHDJIb/bC484oGWWuq7JoDdiqsjtalx3sVPQzxBuyvkEk/fCILTWgbsvW7spWct3neeD+zGdKEgoeCICovCU/R34qzEL1uVOQaLjeZ0GUoOnoryWQ+c6ITxO35IDh+qMr2Pi2hWtUS5nlgXH/tTFu0LndEs0mw0V+6mT9cZcWvI4uZ5uRyBOWITynzmFKGfYmtcCv1yxu5chaU92vIOu4Dmd+AcQgkfmzCQX935zcNk90hzR3cK2WKKnlhgZEiLgoV8/Midnlfq7juUbamV15AcQqQqAdJBFebKva4e4QIiQA0R5SCDsxn8qeNK0RyJCtszmmnSfQzenSD200SL0kBBqui91eTeBFB2/mXdQoZAyfZ1SRSRZO3jcJSuBltnLjnrLPGMAUjjplqSRysZhWbk/VoK3vg15rmDTJJqRvC0V30DkotzSM4qU8YpnZwEnUiUVjVsOJgGZjsqNi+FARghVhwdxmlsAd1aPqB1EJ9NqCdeTLuU5yztWWMATz32Vvc3VdCMOUW70EGdcy59JjR6L83vIsHptjcl3HDzr3ovDtV/uLkHyeUbiCgF4avSz4/5yJ3uLFrAUdUGFxdc+WIT5cLuK16nWp33VuJ8VTbD6SiU3osiyUeUnvttihtwtpU3UVmZuEUQqV7MbkX6k4TN6JqVwR1ln4ZIXTRW7VhPcLVvyXQ7kg2qW0+b98+n4Gg+L5eXs+2NhO20zTGnAsBb99AygxbjQmTVCm25NDWFIhoMa04Z9bXN89EqO2M2X3jWmnXXZ1m2EYmiv6gCBfftD0un3QSkD2u7xjakw2I9hwEad+oQdP0oUIB44s0VjE/ZQmKrOQs0WiavE4xVfFOYJQdAwtcGHKxKOlec3LBhnhHQvM5W4FTVor1O/LcI9fs5fZ/SZ7WE4hNNfNL/EPWT8nH50ODv9n7ScRYO0cvFw6LFT/YAroFLseGDru7jBMz0fisqnLL1lQtopT6w9GDuizE+hEVQd9qVN6uN7YjudPYEvchw3hvgMxHGWN+tSPju4E8MazGH+r6QMFSLCZGCM94PChlHfUzb7WTJ7j4ua0doYUGJdGAxXGqcWYdDZ0Kk4TrDXWkxzpNNrlgyxkK+8tMXyqClfL8LL266E23yIM603GtG7TNUiH9m03aEH4YNjOO8H1bBBxbh6jyO5D//gDLBVjRTrMP3KKOIBFoCipYdULnpZRbPvl58PldwURixOQ8utLtFxOaBLMjJ1FPOly0csbmoOqpCQwzN1s6RX+jUn4FFkh3JkjBLetI+TrU8PKi2tGC6qAhENCiMImkCg/oD7Yspi9Se0/7JGmUBFU4N96T1eHKoeI8kF0sHCjXd/ZZoKJDHssOcgWZzklYpP6CpwGuTMVtuSnzvkoBPhtk8YnIKR8GwzMWpVFvB44nSNzoDJ3NcJUh79AglfccmReXjZOBg+F/kJAkADhw57QADQGElQ6OItrrDnOWMvMCk925i5CNgRrnhugCgRXrXI5YP9uoaDVEovMJzh26gSmnuG8s3cdYVrPBtRt9eldx2IHWpXO8n+7/LM2WH3YIs6LoHdygmAW3Kkk/vjUvVj/qMOaXjiAON6Oq3lEqkCKx4UgyoULMxPgWevNBm6osmtQZYOt93NHEVI5u85IfjKkJF9DB+DVgWSoYubRUZtzpbL7I5R1RurzcS/REwnnCRlRMQc/qabrbpLlo6KLArK0vQ8kekHjO4Rxa8maH6djlrA482uwPVhAyCwY0L4JNcoDt9mbaUpKRB01kJ7Uq3o2E9V9BXIqDhKQ/9szos87vr/ChzzpqnDM79D1FMui4T4KRX/iZQAW9jKyNIPXPzLQM/pHG5fyKS1xTvZ4KFukCf5MCsyWw9LxQJDSYJhMjBLpdQSNi06HVBHj/ZY5mAzZo10VcOwSGN76k1g7d5jYv999BV+e1Ogx9SLPhDLN0PdadiozEAuqAAAkwYuddZ3ZvB9CDIrohL89ypfdWbcTiWcpFMdlpSSEvPUYhYUA2dgxdQTKPPfQ+K/LuxMNuN5tWL/PED5ouexw3X3PqeimOr3bwxPvh9z4KZxP6akCOTy6AzPsklkgJk3eugTs881gEhiVhPNAPrO2BLsFXTj6yg5iRa5NlgXI+a9ITIAIL/cYMfdudyw0NK6mniI0Twiuussv+D5z0kyt5kHhJ6vdX5rPLhG7wJvbKrX036jvapmbxQL2+QthagbUQtbzYy7rnvHu7sUe1f0UlCrRM0AEYR3szCi2kmR6NSPHFdRBqG3YWDRbuUdovmzbXMuOqwuGFvUbXKq1PvqipGXd0yYlHflQFhllVr0/+SHdYJk7X1Uor14ksnWk+IOan5xlI2j3/RlKzN3JGr5mvLAhJOHqYpkc3LUA/ExeVbt/QHoeL4+Hb6u+7jCZ3v9EJc7CrGTmXB3pn9uatGHqB4dWWzJTtwqKygrUWLuYLiEpgDJK9p/y39ORq9t4B8u3/0iD1zAiW37RBJEj+3IJ3k6XiOzj9UAx2rKs2rbYXvLHgwnhSbAhdrvJCNw6rNSLKAZAelPnZ2lQIceY17VNW3PBL8GJttheN77lSu7ykybM3+lMJimCmWViqXsEPIfGpgCrsIeMTsVqHx8hBxPGYlcaOND5ITi+mBtEau/SpqAxEBLnX6cbggOxdgcxldZ0HU91lHsj4wbB+glZTzD0txBq0/0uwLMDBInV3Vp2Z7Lojsch8ctus4U4nV4nNGk/sRnTws3C8dgvppMrHSCSmDjcBpATeezwyOXVy4OkzG365HpR4/vx9EeZnG4Xcug/pES0hIufbKaqpDeDPLEdTLKouiqrf5tM47VV30RhHCXZzFIueoBi16XhCHI/PRkMEqMxFsfPaP+qmT/YiFRF6GwAsNcjy4/+s+PA7ITB0DsE4srhSUpkbK+91WtkapOIvnN8l/0BfOb53/QIBzfJf9AX8pTEiAAA=' style='width:28px;height:28px;border-radius:6px;object-fit:cover'/></div>"
            f"<div><span style='color:#2e7d32;font-weight:700;font-size:15px'>iGreen</span>"
            f"<span style='color:#1a2e1a;font-weight:700;font-size:15px'> Performance</span></div>"
            f"</div></div>",
            unsafe_allow_html=True)
        st.markdown(
            f"<div style='margin:0 8px 12px;background:#f0f7f0;border:1px solid #c8e0c8;"
            f"border-radius:8px;padding:10px 12px'>"
            f"<div style='color:#2e7d32;font-weight:700;font-size:14px'>{u['nome']}</div>"
            f"<div style='color:#5a8a5a;font-size:11px'>{role_label}</div>"
            f"</div>",
            unsafe_allow_html=True)
        anos=get_anos_disponiveis()
        ano=st.selectbox('Ano',anos,label_visibility='collapsed')
        meses=get_todos_meses_ano(int(ano))
        mes_labels=[m.split('-')[0] for m in meses]
        mes_sel=st.selectbox('Mês',mes_labels,index=datetime.now().month-1,label_visibility='collapsed')
        mes_ano=f'{mes_sel}-{ano}'
        st.markdown("<div style='height:8px'></div>",unsafe_allow_html=True)
        if u['role']=='diretor':
            pags=['Quadro de Resultados','Visualização RCA','Análise dos Operadores','Monitorias','Análise de Inadimplência','Metas','Minha Conta']
        elif u['role']=='diretor_upload':
            pags=['Quadro de Resultados','Upload de Bases','Minha Conta']
        elif u['role']=='admin':
            pags=['Quadro de Resultados','Lançamento','Visualização RCA','Análise dos Operadores','Monitorias','Upload de Bases','Análise de Inadimplência','Metas','Minha Conta']
        elif u.get('equipe')=='metcool':
            pags=['Meet Call','Análise dos Operadores','Monitorias','Minha Conta']
        else:
            pags=['Quadro de Resultados','Lançamento','Meet Call','Análise dos Operadores','Monitorias','Upload de Bases','Análise de Inadimplência','Metas','Minha Conta']
        pag=st.radio('Menu',pags,label_visibility='collapsed',index=0)
        st.markdown("<div style='height:8px'></div>",unsafe_allow_html=True)
        if st.button('Sair',use_container_width=True,key='btn_sair'):
            del st.session_state.usuario; st.rerun()
    return mes_ano,pag

# ── OPERADORES ─────────────────────────────────
def pagina_operadores():
    u=st.session_state.usuario
    header_page("Operadores","Gerencie os operadores da equipe")
    eq=seletor_equipe(u["equipe"])


    with st.expander("Cadastrar Novo Operador",expanded=False):
        c1,c2,c3=st.columns([3,1,1])
        with c1: nn=st.text_input("Nome",placeholder="Nome completo", key="op_nome_input")
        with c2: np_op=st.checkbox("Pleno", key="op_pleno_input")
        with c3:
            st.markdown("<div style='margin-top:28px'>",unsafe_allow_html=True)
            if st.button("Buscar / Cadastrar", use_container_width=True, key="op_buscar_btn"):
                if nn.strip():
                    primeiro = nn.strip().upper().split()[0]
                    encontrado = next((op for op in get_db().operadores.find({"equipeId": {"$ne": eq}}) 
                                      if op.get('nome','').strip().upper().split()[0] == primeiro), None)
                    if encontrado:
                        st.session_state['op_busca_result'] = {'op': encontrado, 'eq': eq, 'nome': nn.strip(), 'pleno': np_op}
                    else:
                        salvar_operador(eq, nn.strip(), np_op)
                        buscar_operadores.cache_clear()
                        st.session_state['op_busca_result'] = None
                        st.success(f"✅ {nn} cadastrado!")
                        st.rerun()
                else:
                    st.error("Digite o nome.")
            st.markdown("</div>",unsafe_allow_html=True)

        # Resultado da busca — dentro do expander, sem rerun
        if st.session_state.get('op_busca_result'):
            r = st.session_state['op_busca_result']
            st.warning(f"⚠️ **{r['op']['nome']}** já existe na equipe **{r['op'].get('equipeId','')}**.")
            st.markdown("Quer vincular o mesmo operador (mantém histórico de monitorias) ou criar um novo?")
            cv1, cv2, cv3 = st.columns(3)
            with cv1:
                if st.button("🔗 Vincular", key="op_vincular_btn", use_container_width=True):
                    novo_id = f"{r['op']['_id']}-{r['eq']}"
                    get_db().operadores.update_one(
                        {"_id": novo_id},
                        {"$set": {"nome": r['op']['nome'], "equipeId": r['eq'], "pleno": r['pleno'], "vinculadoA": r['op']['_id']}},
                        upsert=True
                    )
                    buscar_operadores.cache_clear()
                    st.session_state['op_busca_result'] = None
                    st.rerun()
            with cv2:
                if st.button("➕ Criar novo", key="op_criar_btn", use_container_width=True):
                    salvar_operador(r['eq'], r['nome'], r['pleno'])
                    buscar_operadores.cache_clear()
                    st.session_state['op_busca_result'] = None
                    st.rerun()
            with cv3:
                if st.button("❌ Cancelar", key="op_cancelar_btn", use_container_width=True):
                    st.session_state['op_busca_result'] = None
                    st.rerun()


    st.markdown("---")
    ops=buscar_operadores(eq)
    if not ops:
        st.info("Nenhum operador cadastrado.")
        padrao=OPERADORES_PADRAO.get(eq,[])
        if padrao:
            if st.button("Importar Operadores Padrão",use_container_width=True):
                for nome,pleno in padrao:
                    oid=re.sub(r'[^a-z0-9]','-',nome.lower().strip())
                    oid=re.sub(r'-+','-',oid).strip('-')
                    oid=f"{eq[:3]}-{oid}"[:40]
                    if not get_db().operadores.find_one({"_id":oid}):
                        get_db().operadores.insert_one({"_id":oid,"equipeId":eq,"nome":nome,"pleno":pleno,"criadoEm":datetime.now()})
                st.success("Importados!"); st.rerun()
        return
    for op in ops:
        c1,c2,c3,c4=st.columns([3,1,1,1])
        with c1: nn=st.text_input("n",value=op["nome"],label_visibility="collapsed",key=f"n_{op['_id']}")
        with c2: np=st.checkbox("Pleno",value=op.get("pleno",False),key=f"p_{op['_id']}")
        with c3:
            if st.button("Salvar",key=f"s_{op['_id']}"): atualizar_operador(op["_id"],nn,np); st.rerun()
        with c4:
            if st.button("Excluir",key=f"d_{op['_id']}"): excluir_operador(op["_id"]); st.rerun()


def ler_arquivo(arquivo, sheet_name=None):
    """Lê Excel ou CSV automaticamente."""
    import unicodedata
    nome = arquivo.name.lower()
    # Colunas que devem ser lidas como string (CPF, identificadores)
    dtype_str = {"cpf": str, "CPF": str, "uc_cpf": str, "identificador": str,
                 "numinstalacao": str, "idcliente": str, "conta_unica": str}
    if nome.endswith('.csv'):
        for sep in [',',';','|','	']:
            try:
                arquivo.seek(0)
                df = pd.read_csv(arquivo, sep=sep, header=0, low_memory=False, dtype=dtype_str)
                if len(df.columns) > 1:
                    return df
            except: pass
        arquivo.seek(0)
        return pd.read_csv(arquivo, header=0, low_memory=False, dtype=dtype_str)
    else:
        if sheet_name:
            return pd.read_excel(arquivo, sheet_name=sheet_name, header=0, dtype=dtype_str)
        return pd.read_excel(arquivo, header=0, dtype=dtype_str)

def mapear_colunas_pagos(df):
    """Mapeia colunas da base de pagos para nomes padrão."""
    import unicodedata
    def norm(s): return unicodedata.normalize('NFKD',str(s).upper().strip()).encode('ascii','ignore').decode()
    
    cols_norm = {norm(str(c)): c for c in df.columns}
    
    # CPF — match exato primeiro, depois parcial (nunca pega numinstalacao ou idcliente)
    col_cpf = cols_norm.get("CPF") or cols_norm.get("DOCUMENTO") or cols_norm.get("IDENTIFICADOR")
    if not col_cpf:
        for c in df.columns:
            cn = norm(str(c))
            if "CPF" in cn and "CLIENTE" not in cn and "NOME" not in cn:
                col_cpf = c; break

    # VALOR — prioriza valorapagar
    col_val = cols_norm.get("VALORAPAGAR") or cols_norm.get("VALOR_APAGAR") or cols_norm.get("VALORPAGAR")
    if not col_val:
        for c in df.columns:
            cn = norm(str(c))
            if any(x in cn for x in ["VALORAPAGAR","VALOR_APAGAR"]):
                col_val = c; break
    if not col_val:
        for c in df.columns:
            cn = norm(str(c))
            if any(x in cn for x in ["VALOR","VLR"]) and c != col_cpf:
                col_val = c; break

    # DATA PAGAMENTO
    col_dpag = cols_norm.get("DTPAGAMENTO") or cols_norm.get("DT_PAGAMENTO") or cols_norm.get("DATAPAGAMENTO")
    if not col_dpag:
        for c in df.columns:
            cn = norm(str(c))
            if any(x in cn for x in ["DTPAGAMENTO","DT_PAGAMENTO","DATAPAGAMENTO","DATA_PAGAMENTO","PAGAM","PAGTO","DT_PAG","BAIXA"]):
                col_dpag = c; break

    # DATA VENCIMENTO
    col_dvenc = cols_norm.get("DTVENCIMENTO") or cols_norm.get("DT_VENCIMENTO") or cols_norm.get("DATAVENCIMENTO")
    if not col_dvenc:
        for c in df.columns:
            cn = norm(str(c))
            if any(x in cn for x in ["DTVENCIMENTO","DT_VENCIMENTO","DATAVENCIMENTO","DATA_VENCIMENTO","VENCIMENTO"]) and c != col_dpag:
                col_dvenc = c; break

    # FORNECEDORA
    col_forn = cols_norm.get("FORNECEDORA") or cols_norm.get("FORNECEDOR")
    if not col_forn:
        for c in df.columns:
            cn = norm(str(c))
            if any(x in cn for x in ["FORNEC","FORN"]):
                col_forn = c; break

    # UF — manter coluna original no dataframe (não precisa renomear)

    mapa={}
    if col_cpf:  mapa[col_cpf]="uc_cpf"
    if col_val:  mapa[col_val]="valor"
    if col_dpag: mapa[col_dpag]="data_pagamento"
    if col_dvenc: mapa[col_dvenc]="data_vencimento"
    if col_forn: mapa[col_forn]="fornecedora"
    return df.rename(columns=mapa)

def processar_contatos(dc):
    """Extrai CPF e data de contato de um dataframe."""
    import unicodedata
    def norm(s): return unicodedata.normalize('NFKD',str(s).upper().strip()).encode('ascii','ignore').decode()
    # Remover espaços dos nomes das colunas
    dc = dc.copy()
    dc.columns = [str(c).strip() for c in dc.columns]
    cc=next((c for c in dc.columns if norm(str(c)) in ["CPF","IDENTIFICADOR","IDENTIF","IDENTIFICACAO"]),dc.columns[0])
    cd=next((c for c in dc.columns if any(x in norm(str(c)) for x in ["DATA","DT_","BAIXA","CONTATO","INTERAC","LIGAC","CHAT","DISPAR","PAGAM","DIA"])),dc.columns[1] if len(dc.columns)>1 else dc.columns[0])
    dd=pd.DataFrame({"uc_cpf":dc[cc].apply(normalizar_cpf),"data_contato":parse_data_inteligente(dc[cd])}).dropna(subset=["data_contato"])
    dd=dd[dd["uc_cpf"].str.len()>=8]
    dd=dd[~dd["uc_cpf"].str.match(r"^0+$")]
    dd=dd[dd["uc_cpf"]!="nan"]
    # Manter apenas a PRIMEIRA interação por CPF
    dd=dd.sort_values("data_contato").drop_duplicates(subset=["uc_cpf"],keep="first")
    return dd

def finalizar_processamento(df, contatos, abas_lidas, eq, ma):
    """Aplica elegibilidade e retorna df final."""
    if contatos:
        pc=pd.concat(contatos,ignore_index=True).groupby("uc_cpf",as_index=False)["data_contato"].min()
        df["primeiro_contato"]=df["uc_cpf"].map(dict(zip(pc["uc_cpf"],pc["data_contato"])))
    else: df["primeiro_contato"]=pd.NaT
    df=df.drop(columns=["_row_id"],errors="ignore").reset_index(drop=True)
    df["data_pagamento"]=pd.to_datetime(df["data_pagamento"],errors="coerce").dt.normalize()
    df["primeiro_contato"]=pd.to_datetime(df["primeiro_contato"],errors="coerce").dt.normalize()
    df["diferenca_dias"]=(df["data_pagamento"]-df["primeiro_contato"]).dt.days
    def classif(row):
        if pd.isna(row.get("primeiro_contato")): return "ND"
        d=row.get("diferenca_dias")
        if pd.isna(d): return "ND"
        # Elegível: contato ANTES ou NO DIA do pagamento (diferença >= 0)
        # Não Elegível: contato DEPOIS do pagamento (diferença < 0)
        return "Elegível" if int(d)>=0 else "Não Elegível"
    df["elegibilidade"]=df.apply(classif,axis=1)
    if "data_vencimento" in df.columns: df["dias_vencidos"]=(df["data_pagamento"]-df["data_vencimento"]).dt.days
    else: df["dias_vencidos"]=None
    df["aging"]=df["dias_vencidos"].apply(aging_faixa)
    for col in ["data_vencimento","data_pagamento","primeiro_contato"]:
        if col in df.columns:
            try: df[col]=pd.to_datetime(df[col],errors="coerce").dt.strftime("%Y-%m-%d").where(pd.to_datetime(df[col],errors="coerce").notna(),other=None)
            except: pass
    df["equipe"]=eq; df["mes_ano"]=ma
    return df

def processar_base_unica(arquivo, eq, ma):
    import unicodedata
    def norm(s): return unicodedata.normalize('NFKD',str(s).upper().strip()).encode('ascii','ignore').decode()

    nome_arq = arquivo.name.lower()
    is_csv = nome_arq.endswith('.csv')

    if is_csv:
        # CSV — base de pagos direto
        try:
            df = ler_arquivo(arquivo)
        except Exception as e: return None,[f"Erro ao ler CSV: {e}"],[]
        df = mapear_colunas_pagos(df)
        df["_row_id"]=df.index
        if "uc_cpf" in df.columns: df["uc_cpf"]=df["uc_cpf"].apply(normalizar_cpf)
        if "data_pagamento" in df.columns:
            df["data_pagamento"]=parse_data_inteligente(df["data_pagamento"])
        if "data_vencimento" in df.columns: df["data_vencimento"]=parse_data_inteligente(df["data_vencimento"])
        if "valor" in df.columns:
            def cv(v):
                s=str(v).strip().replace("R$","").replace(" ","")
                try: return float(s)
                except:
                    try: return float(s.replace(".","").replace(",","."))
                    except: return 0.0
            df["valor"]=df["valor"].apply(cv)
        # CSV não tem abas de contato — retorna sem interação
        df = finalizar_processamento(df, [], [], eq, ma)
        return df, [], ["(sem interação — use upload separado)"]
    else:
        # Excel — lógica original com abas
        try: xls=pd.ExcelFile(arquivo)
        except Exception as e: return None,[f"Erro: {e}"],[]
        abas_norm=[norm(a) for a in xls.sheet_names]; abas_orig=xls.sheet_names
        aba_pagos=None
        for i,a in enumerate(abas_norm):
            if any(p in a for p in ["PAGO","PAGAM","RECEB","BASE","BAIXA","PAGT","RESULT"]):
                aba_pagos=abas_orig[i]; break
        if not aba_pagos:
            for i,a in enumerate(abas_norm):
                if not any(x in a for x in ["CHAT","LIG","DISPAR","CONTATO"]):
                    aba_pagos=abas_orig[i]; break
        if not aba_pagos: aba_pagos=abas_orig[0]
        df=pd.read_excel(xls,sheet_name=aba_pagos,header=0).reset_index(drop=True)
        df["_row_id"]=df.index
        df = mapear_colunas_pagos(df)
        if "uc_cpf" in df.columns: df["uc_cpf"]=df["uc_cpf"].apply(normalizar_cpf)
        if "data_pagamento" in df.columns:
            df["data_pagamento"]=parse_data_inteligente(df["data_pagamento"])
        if "data_vencimento" in df.columns: df["data_vencimento"]=parse_data_inteligente(df["data_vencimento"])
        if "valor" in df.columns:
            def cv(v):
                s=str(v).strip().replace("R$","").replace(" ","")
                try: return float(s)
                except:
                    try: return float(s.replace(".","").replace(",","."))
                    except: return 0.0
            df["valor"]=df["valor"].apply(cv)
        contatos=[]; abas_lidas=[]

        # Regra: se existir aba INTERAÇÃO, usar EXCLUSIVAMENTE ela
        aba_interacao=next((abas_orig[i] for i,a in enumerate(abas_norm) if "INTERAC" in a or a=="INTERACAO" or a=="INTERAÇÃO"),None)

        if aba_interacao:
            try:
                dc=pd.read_excel(xls,sheet_name=aba_interacao,header=0)
                if not dc.empty and len(dc.columns)>=2:
                    dd=processar_contatos(dc)
                    if not dd.empty: contatos.append(dd); abas_lidas.append("Interação")
            except: pass
        else:
            # Sem aba INTERAÇÃO — consolidar todas disponíveis (Chat, Ligações, Disparo)
            for busca,nome in [("CHAT","CHAT"),("LIG","LIGAÇÕES"),("DISPAR","DISPAROS")]:
                aba=next((abas_orig[i] for i,a in enumerate(abas_norm) if busca in a),None)
                if not aba: continue
                try:
                    dc=pd.read_excel(xls,sheet_name=aba,header=0)
                    if dc.empty or len(dc.columns)<2: continue
                    dd=processar_contatos(dc)
                    if not dd.empty: contatos.append(dd); abas_lidas.append(nome)
                except: pass

        df = finalizar_processamento(df, contatos, abas_lidas, eq, ma)
        return df,[],abas_lidas

def pagina_metas(ma):
    u=st.session_state.usuario
    header_page("Metas",ma.replace("-"," "))
    if u["role"]=="diretor":
        eq_opts=list(EQUIPES.keys())
        eq_labels=[f"Equipe {EQUIPES[e]['nome']}" for e in eq_opts]
        eq_sel=st.selectbox("Selecionar equipe:",eq_labels,key="dir_eq_metas")
        eq=eq_opts[eq_labels.index(eq_sel)]
    else:
        eq=seletor_equipe(u.get("equipe") or "luciano")
    ops_todos=buscar_operadores(eq)
    # Luciano: exclui Meet Call das metas
    ops = [op for op in ops_todos if op["nome"] not in OPERADORES_MEETCALL] if eq=="luciano" else ops_todos
    if not ops: st.warning("Cadastre operadores primeiro."); return
    buscar_metas_equipe.clear()
    st.markdown("### Meta da Gestora")
    mg_doc=buscar_meta_gestora(ma,eq)
    c1,c2=st.columns([2,1])
    with c1: mg_val=st.number_input("Meta Base (R$)",min_value=0.0,step=1000.0,format="%.2f",value=float(mg_doc.get("metaGestora",0)),key="mg_val")
    tpct=100
    with c2: st.markdown(f"<div style='padding-top:28px;color:#2daf5c;font-weight:700;font-size:16px'>{fmt_brl(mg_val)}</div>",unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("### Metas por Operador")
    ms=buscar_metas_equipe(ma,eq); mn={}
    for op in ops:
        c1,c2=st.columns([3,2])
        with c1: st.markdown(f"<div style='padding-top:10px;color:#1a3a1a'>{'[P] ' if op.get('pleno') else ''}{op['nome']}</div>",unsafe_allow_html=True)
        with c2: mn[op["_id"]]=st.number_input("m",label_visibility="collapsed",min_value=0.0,step=100.0,format="%.2f",value=float(ms.get(op["_id"],0)),key=f"mg_{ma}_{op['_id']}")
    st.markdown("---")
    if st.button("Salvar Metas",use_container_width=True):
        for oid,v in mn.items(): salvar_meta_operador(ma,eq,oid,v)
        salvar_meta_gestora(ma,eq,mg_val,tpct)
        buscar_metas_equipe.clear()
        st.session_state["meta_salva_msg"] = f"✅ Metas da Equipe {EQUIPES.get(eq,{}).get('nome',eq)} salvas com sucesso!"
        st.rerun()
    if st.session_state.get("meta_salva_msg"):
        st.success(st.session_state.pop("meta_salva_msg"))

# ── LANÇAMENTO ─────────────────────────────────
def pagina_lancamento(ma):
    u=st.session_state.usuario
    header_page("Lançamento de Resultado",ma.replace("-"," "))
    eq=seletor_equipe(u["equipe"])
    # Se equipe metcool, usar tela Meet Call
    if eq=="metcool":
        pagina_meetcall(ma)
        return
    ops_todos=buscar_operadores(eq)
    ops=[op for op in ops_todos if op["nome"] not in OPERADORES_MEETCALL] if eq=="luciano" else ops_todos
    if not ops: st.warning("Cadastre operadores primeiro."); return
    ms=buscar_metas_equipe(ma,eq)
    if st.session_state.get("ultimo_salvo"):
        st.success(st.session_state.ultimo_salvo)
        st.session_state.ultimo_salvo=""
    st.markdown("### Configuração do Lançamento")
    c1,c2,c3=st.columns([2,1,1])
    with c1:
        hoje=date.today()
        data_sel=st.date_input("Data do Resultado *",value=hoje,min_value=date(hoje.year,1,1),max_value=date(hoje.year,12,31),key=f"data_{eq}_{ma}")
        eh_fech=st.checkbox("Fechamento do Mês",key=f"fech_{eq}_{ma}")
    with c2: dt=st.number_input("Dias Trabalhados *",min_value=0,max_value=31,value=0,key=f"dt_{eq}_{ma}")
    with c3: td=st.number_input("Total Dias do Mês *",min_value=0,max_value=31,value=0,key=f"td_{eq}_{ma}")
    up_atual=buscar_ultimo_processamento(ma,eq)
    rec_auto=float(up_atual.get("valorElegivel",0)) if up_atual else 0
    # rec_anterior = último recGeral salvo manualmente (prioridade) ou da base
    rec_anterior=0.0
    lancs_ant=buscar_lancamentos(ma,eq)
    for l in lancs_ant:
        if l.get("recGeral",0)>0:
            rec_anterior=float(l["recGeral"]); break
    if rec_anterior==0:
        rec_anterior=rec_auto
    usar_rec_manual=st.checkbox("Inserir Recebido Geral manualmente",key=f"rec_manual_chk_{eq}_{ma}")
    if usar_rec_manual:
        rec_geral_manual=st.number_input("Recebido Geral (R$)",min_value=0.0,step=100.0,format="%.2f",value=float(rec_anterior or 0),key=f"rec_geral_manual_{eq}_{ma}")
        if rec_anterior>0:
            st.caption(f"Último valor salvo: {fmt_brl(rec_anterior)}")
    else:
        rec_geral_manual=rec_anterior
    st.markdown("---")
    st.markdown("### Valores por Operador")
    mostrar_imp = st.checkbox("Importar via Excel", key=f"chk_imp_{eq}_{ma}")
    arq_imp = None
    if mostrar_imp:
        arq_imp = st.file_uploader("Planilha Excel (.xlsx)", type=["xlsx"], key=f"imp_{eq}_{ma}")
    if arq_imp:
        arq_imp.seek(0)
        resultados_imp, erro_imp = importar_excel_operadores(arq_imp, ops)
        if erro_imp:
            st.error(erro_imp)
        elif resultados_imp:
                st.markdown("**Prévia do lançamento:**")
                prev_rows = []
                for r in resultados_imp:
                    if r['status'] == 'ambiguo':
                        icone = " Ambíguo"
                    elif r['op'] is None:
                        icone = " Não encontrado"
                    else:
                        icone = "" + r['op']['nome']
                    row = {"Excel": r['nome_excel'], "Sistema": icone, "Valor": fmt_brl(r['valor'])}
                    if r['col_lig']: row["Ligações"] = r['ligacoes']
                    prev_rows.append(row)
                st.dataframe(pd.DataFrame(prev_rows), use_container_width=True, hide_index=True)
                pode_importar = all(r['op'] is not None and r['status'] != 'ambiguo' for r in resultados_imp if r['valor'] > 0)
                if not pode_importar:
                    st.warning("Alguns operadores não foram identificados. Corrija manualmente os campos abaixo.")
                if st.button("Confirmar Importação", use_container_width=True, key=f"imp_confirmar_{eq}_{ma}"):
                    for r in resultados_imp:
                        if r['op'] and r['status'] != 'ambiguo' and r['valor'] > 0:
                            # Ignorar operadores Meet Call no lançamento do Luciano
                            if r['op'].get('nome','') in OPERADORES_MEETCALL:
                                continue
                            st.session_state[f"op_{eq}_{ma}_{r['op']['_id']}"] = r['valor']
                            if r['col_lig']:
                                st.session_state[f"lig_{eq}_{ma}_{r['op']['_id']}"] = r['ligacoes']
                    st.success("Valores importados! Revise e salve.")
                    st.rerun()
    vi={}; lig_vi={}
    # No lançamento do Luciano: mostrar APENAS operadores iGreen (sem Meet Call)
    ops_igreen=[op for op in ops if op["nome"] not in OPERADORES_MEETCALL]
    grupos=[]
    grupos.append(("",None,ops_igreen))
    for grupo_nome,grupo_cor,grupo_ops in grupos:
        if grupo_nome:
            st.markdown(f"<p style='color:{grupo_cor};font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:1px;margin:12px 0 4px'>{grupo_nome}</p>",unsafe_allow_html=True)
        for op in grupo_ops:
            meta=float(ms.get(op["_id"],0))
            c1,c2,c3=st.columns([3,2,2])
            with c1: st.markdown(f"<div style='padding-top:10px;color:var(--text-color,#1a3a1a);font-weight:500'>{'★ ' if op.get('pleno') else ''}{op['nome']}</div>",unsafe_allow_html=True)
            with c2: st.markdown(f"<div style='padding-top:10px;color:#2e7d32;font-size:13px;font-weight:500'>{fmt_brl(meta) if meta>0 else '—'}</div>",unsafe_allow_html=True)
            with c3: vi[op["_id"]]=st.number_input("v",label_visibility="collapsed",min_value=0.0,step=100.0,format="%.2f",key=f"op_{eq}_{ma}_{op['_id']}")
            lig_vi[op["_id"]]=st.number_input("Ligações",label_visibility="visible",min_value=0,step=1,value=0,key=f"lig_{eq}_{ma}_{op['_id']}")
    tc=sum(vi.values())
    st.markdown("---")
    usar_manual=st.checkbox("Inserir valor total manualmente",key=f"manual_{eq}_{ma}")
    tc_manual=0.0
    if usar_manual:
        doc_tmp=get_db().temp_lancamento.find_one({"_id":f"tmp_{eq}_{ma}"}) or {}
        tc_anterior_tmp=float(doc_tmp.get("tc_manual",0))
        tc_manual_str=st.text_input("Valor Total com Interação (R$)",
            value=f"{tc_anterior_tmp:.2f}".replace(".",",") if tc_anterior_tmp>0 else "",
            placeholder="Ex: 2.061.583,17",key=f"tc_manual_txt_{eq}_{ma}")
        # Salvar no banco sempre que o valor mudar
        val_digitado=parse_brl(tc_manual_str)
        if val_digitado>0 and val_digitado!=tc_anterior_tmp:
            get_db().temp_lancamento.update_one(
                {"_id":f"tmp_{eq}_{ma}"},
                {"$set":{"tc_manual":val_digitado}},upsert=True)
            tc_manual=val_digitado
        else:
            tc_manual=tc_anterior_tmp
        if tc==0 and tc_manual>0:
            tc=tc_manual
    st.markdown(f"<div style='background:#0a2414;border-radius:8px;padding:12px 16px;margin-bottom:16px'><span style='color:#5a9a70;font-size:11px'>TOTAL COM INTERAÇÃO</span><br><span style='color:#2daf5c;font-size:20px;font-weight:700'>{fmt_brl(tc)}</span></div>",unsafe_allow_html=True)
    ja_salvando=st.session_state.get(f"salvando_{eq}_{ma}",False)
    if st.button("Salvar Lançamento",use_container_width=True,key=f"btn_{eq}_{ma}",disabled=ja_salvando):
        errs=[]
        if dt==0: errs.append("Dias Trabalhados é obrigatório.")
        if td==0: errs.append("Total de Dias do Mês é obrigatório.")
        if errs:
            for e in errs: st.error(e)
        else:
            st.session_state[f"salvando_{eq}_{ma}"]=True
            label="Fechamento do Mês" if eh_fech else data_sel.strftime("%d/%m/%Y")
            ag={op["_id"]:{"valorRecebido":vi.get(op["_id"],0),"nome":op["nome"],"ligacoes":lig_vi.get(op["_id"],0)} for op in ops if op["nome"] not in OPERADORES_MEETCALL}
            tc_ops=sum(float(v.get("valorRecebido",0)) for v in ag.values())
            # Ler tc_manual do banco temporário
            doc_tmp=get_db().temp_lancamento.find_one({"_id":f"tmp_{eq}_{ma}"}) or {}
            tc_manual_val=float(doc_tmp.get("tc_manual",0))
            # Se atendentes zerados usa manual, senão usa atendentes
            tc_real = tc_ops if tc_ops > 0 else tc_manual_val
            # Recebido Geral — preservar o do último lançamento, senão 0
            if usar_rec_manual:
                rg_salvar = rec_geral_manual
            else:
                _lancs_prev = buscar_lancamentos(ma, eq)
                rg_salvar = float(_lancs_prev[0].get('recGeral', 0)) if _lancs_prev else 0
            criar_lancamento(ma,eq,str(data_sel),label,ag,tc_real,0,dt,td,rg_salvar)
            buscar_lancamentos.clear()
            buscar_metas_equipe.clear()
            st.session_state[f"salvando_{eq}_{ma}"]=False
            st.session_state.ultimo_salvo=f"✅ Lançamento salvo! Geral: {fmt_brl(rg_salvar)} | Com Interação: {fmt_brl(tc_real)}"
            st.rerun()
    st.markdown("---")
    lancs=buscar_lancamentos(ma,eq)
    if lancs:
        # Detectar duplicatas por label
        from collections import Counter
        labels_count=Counter(l.get("label","") for l in lancs)
        duplicatas={label for label,cnt in labels_count.items() if cnt>1}

        st.markdown("<p style='color:#81c784;font-size:11px;text-transform:uppercase;letter-spacing:1px;margin-bottom:8px'>Lançamentos do mês</p>",unsafe_allow_html=True)

        if duplicatas:
            st.warning(f"⚠️ Encontrados lançamentos duplicados: {', '.join(duplicatas)}. Exclua os repetidos abaixo.")

        for lanc in reversed(lancs):
            soma=sum(float(v.get("valorRecebido",0) if isinstance(v,dict) else v) for v in lanc.get("agentes",{}).values())
            label_lanc=lanc.get("label","")
            is_dup = label_lanc in duplicatas
            cor_borda = "border-left:3px solid #e53935" if is_dup else ""
            titulo = f"⚠️ DUPLICADO — {label_lanc} — {fmt_brl(soma)}" if is_dup else f"{label_lanc} — {fmt_brl(soma)}"
            with st.expander(titulo):
                if is_dup:
                    st.error("Este lançamento está duplicado! Exclua se não for o correto.")
                rows=[{"Operador":op["nome"],"Valor":fmt_brl(get_val_op(lanc.get("agentes",{}),op["_id"],op["nome"]))} for op in ops]
                st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)
                if st.button("🗑️ Excluir",key=f"del_{lanc['_id']}",type="primary" if is_dup else "secondary"):
                    excluir_lancamento(lanc["_id"])
                    buscar_lancamentos.clear()
                    st.rerun()

# ── QUADRO DE RESULTADOS ───────────────────────
def pagina_quadro(ma):
    u=st.session_state.usuario
    is_dir=u["role"] in ["diretor","diretor_upload"]; is_adm=u["role"]=="admin"
    eqs_vis=["luciano","deborah","tamires"]
    # Para admin e diretor: mostrar Luciano, Déborah, Tamires + Meet Call separado
    # Ordem fixa: Luciano, Meet Call, Déborah, Tamires
    eqs_base=["luciano","metcool","deborah","tamires"]
    if is_adm or is_dir:
        eqs=eqs_base
    elif u.get("equipe")=="metcool": eqs=["metcool"]
    else: eqs=[u["equipe"]]
    header_page("Quadro de Resultados", ma.replace("-"," ").upper())
    if is_dir:
        tot_rec=tot_ci=tot_si=tot_meta=tot_proj=0
        equipes_data=[]
        for eq_r in ["luciano","metcool","deborah","tamires"]:
            try:
                lancs_r=buscar_lancamentos(ma,eq_r)
                ul_r=lancs_r[0] if lancs_r else {}
                if not lancs_r and not buscar_ultimo_processamento(ma,eq_r): continue
                mg_r=float(buscar_meta_gestora(ma,eq_r).get("metaGestora",0))
                # Buscar tc_r do lançamento mais recente que tem agentes com valor
                def _get_tc(lancs, excluir_nomes=None):
                    for l in lancs:
                        ag = l.get("agentes", {})
                        if not isinstance(ag, dict): continue
                        # Usar totalEquipe se disponível
                        te = float(l.get("totalEquipe", 0))
                        if te > 0:
                            return te
                        # Calcular da soma dos agentes
                        total = sum(float(v.get("valorRecebido",0)) for v in ag.values()
                                   if isinstance(v,dict) and (not excluir_nomes or v.get("nome","") not in excluir_nomes))
                        if total > 0:
                            return total
                    return 0.0
                if eq_r=="metcool":
                    tc_r=_get_tc(lancs_r)
                elif eq_r=="luciano":
                    tc_r=_get_tc(lancs_r, excluir_nomes=OPERADORES_MEETCALL)
                else:
                    tc_r=_get_tc(lancs_r)
                dt_r=int(ul_r.get("diasTrabalhados",0)); td_r=int(ul_r.get("totalDias",22))
                # Regra simples: pegar o valor mais recente de cada campo
                # recGeral = primeiro lançamento que tem recGeral > 0
                # totalEquipe = já calculado acima no tc_r
                rg_r = 0.0
                for l in lancs_r:
                    v = float(l.get("recGeral", 0))
                    if v > 0:
                        rg_r = v
                        break
                # Fallback para Meet Call legado
                if rg_r == 0 and eq_r == "metcool":
                    mc_doc_r = buscar_lancamento_meetcall(ma)
                    rg_r = float(mc_doc_r.get("recGeralTotal", mc_doc_r.get("recGeral", 0)))
                si_r=max(0,rg_r-tc_r); proj_r=calc_projecao(rg_r,dt_r,td_r)
                pct_r=(rg_r/mg_r*100) if mg_r>0 else 0
                tot_rec+=rg_r; tot_ci+=tc_r; tot_si+=si_r; tot_meta+=mg_r; tot_proj+=proj_r
                nome_eq = "Meet Call" if eq_r=="metcool" else EQUIPES[eq_r]["nome"]
                equipes_data.append({"nome":nome_eq,"rg":rg_r,"ci":tc_r,"si":si_r,"meta":mg_r,"proj":proj_r,"pct":pct_r})
            except: pass
        if equipes_data:
            pct_t=(tot_rec/tot_meta*100) if tot_meta>0 else 0
            cv_t=cor_pct(pct_t)
            # Mês anterior
            meses_map2={"Janeiro":1,"Fevereiro":2,"Março":3,"Abril":4,"Maio":5,"Junho":6,"Julho":7,"Agosto":8,"Setembro":9,"Outubro":10,"Novembro":11,"Dezembro":12}
            if "-" in ma:
                partes_ma=ma.split("-"); mes_r=meses_map2.get(partes_ma[0],1); ano_r=int(partes_ma[1])
            else:
                mes_r=1; ano_r=2026
            ma_ant=f"Dezembro-{ano_r-1}" if mes_r==1 else f"{MESES_NOMES[mes_r-2]}-{ano_r}"
            rg_ant={}
            for eq_ant in ["luciano","metcool","deborah","tamires"]:
                try:
                    nome_ant="Meet Call" if eq_ant=="metcool" else EQUIPES[eq_ant]["nome"]
                    if eq_ant=="metcool":
                        lancs_mc_ant=buscar_lancamentos(ma_ant,"metcool")
                        if lancs_mc_ant:
                            rg_ant[nome_ant]=float(lancs_mc_ant[0].get("recGeral",0)) or float(lancs_mc_ant[0].get("totalEquipe",0))
                        else:
                            mc_ant=buscar_lancamento_meetcall(ma_ant)
                            rg_ant[nome_ant]=float(mc_ant.get("recGeralTotal",mc_ant.get("recGeral",0)))
                    else:
                        lancs_ant=buscar_lancamentos(ma_ant,eq_ant); v=0.0
                        for l in (lancs_ant or []):
                            if l.get("recGeral",0)>0: v=float(l["recGeral"]); break
                        if v==0:
                            up_ant=buscar_ultimo_processamento(ma_ant,eq_ant)
                            v=float(up_ant.get("valorElegivel",0)) if up_ant else 0
                        rg_ant[nome_ant]=v
                except: pass

            # Build HTML table for perfect column alignment
            th = lambda t: f'<th style="color:#3a6a4a;font-size:10px;text-transform:uppercase;padding:8px 12px;font-weight:400;white-space:nowrap;text-align:left">{t}</th>'
            
            header = (
                '<thead><tr>'
                + '<th style="min-width:120px;padding:8px 12px"></th>'
                + th("RECEBIDO")
                + th("COM INTER.")
                + th("SEM INTER.")
                + th("META")
                + th("PROJEÇÃO")
                + th("% META")
                + th("MÊS ANT.")
                + th("CRESCIMENTO")
                + '</tr></thead>'
            )
            
            rows_html = ""
            for ed in equipes_data:
                cv_e = cor_pct(ed["pct"])
                rg_anterior = rg_ant.get(ed["nome"], 0)
                if rg_anterior > 0:
                    diff = ed["rg"] - rg_anterior
                    pct_cresc = (diff/rg_anterior*100) if rg_anterior > 0 else 0
                    seta = f"↑ +{pct_cresc:.1f}%" if diff > 0 else (f"↓ {pct_cresc:.1f}%" if diff < 0 else "→ 0%")
                    cor_seta = "#00c853" if diff > 0 else ("#e53935" if diff < 0 else "#8ab89a")
                    td_ant = f'<td style="padding:10px 12px;white-space:nowrap;color:#8ab89a;font-size:13px">{fmt_brl(rg_anterior)}</td>'
                    td_cresc = f'<td style="padding:10px 12px;white-space:nowrap;color:{cor_seta};font-size:13px;font-weight:700">{seta}</td>'
                else:
                    td_ant = '<td style="padding:10px 12px;color:#5a6a5a;font-size:13px">—</td>'
                    td_cresc = '<td style="padding:10px 12px;color:#5a6a5a;font-size:13px">—</td>'
                
                rows_html += (
                    f'<tr style="border-bottom:1px solid #1e3a1e">'
                    f'<td style="padding:10px 12px;color:#ffffff;font-weight:700;font-size:13px;white-space:nowrap">Equipe {ed["nome"]}</td>'
                    f'<td style="padding:10px 12px;color:#00c853;font-weight:700;font-size:14px;white-space:nowrap">{fmt_brl(ed["rg"])}</td>'
                    f'<td style="padding:10px 12px;color:#e8f5e9;font-size:13px;white-space:nowrap">{fmt_brl(ed["ci"])}</td>'
                    f'<td style="padding:10px 12px;color:#8ab89a;font-size:13px;white-space:nowrap">{fmt_brl(ed["si"])}</td>'
                    f'<td style="padding:10px 12px;color:#8ab89a;font-size:13px;white-space:nowrap">{fmt_brl(ed["meta"])}</td>'
                    f'<td style="padding:10px 12px;color:#8ab89a;font-size:13px;white-space:nowrap">{fmt_brl(ed["proj"])}</td>'
                    f'<td style="padding:10px 12px;color:{cv_e};font-size:14px;font-weight:800;white-space:nowrap">{ed["pct"]:.2f}%</td>'
                    f'{td_ant}'
                    f'{td_cresc}'
                    f'</tr>'
                )

            # Total row
            tot_ant = sum(rg_ant.get(ed["nome"], 0) for ed in equipes_data)
            tot_diff = tot_rec - tot_ant if tot_ant > 0 else 0
            tot_pct_cresc = (tot_diff/tot_ant*100) if tot_ant > 0 else 0
            tot_seta = f"↑ +{tot_pct_cresc:.1f}%" if tot_diff > 0 else (f"↓ {tot_pct_cresc:.1f}%" if tot_diff < 0 else "→ 0%")
            tot_cor = "#00c853" if tot_diff > 0 else ("#e53935" if tot_diff < 0 else "#8ab89a")
            
            total_html = (
                f'<tr style="background:#051005">'
                f'<td style="padding:12px;color:#ffffff;font-weight:800;font-size:14px;white-space:nowrap">TOTAL GERAL</td>'
                f'<td style="padding:12px;color:#00c853;font-weight:800;font-size:15px;white-space:nowrap">{fmt_brl(tot_rec)}</td>'
                f'<td style="padding:12px;color:#ffffff;font-weight:700;font-size:14px;white-space:nowrap">{fmt_brl(tot_ci)}</td>'
                f'<td style="padding:12px;color:#8ab89a;font-size:13px;white-space:nowrap">{fmt_brl(tot_si)}</td>'
                f'<td style="padding:12px;color:#8ab89a;font-size:13px;white-space:nowrap">{fmt_brl(tot_meta)}</td>'
                f'<td style="padding:12px;color:#8ab89a;font-size:13px;white-space:nowrap">{fmt_brl(tot_proj)}</td>'
                f'<td style="padding:12px;color:{cv_t};font-size:16px;font-weight:800;white-space:nowrap">{pct_t:.2f}%</td>'
                f'<td style="padding:12px;color:#8ab89a;font-size:13px;white-space:nowrap">{"—" if tot_ant==0 else fmt_brl(tot_ant)}</td>'
                f'<td style="padding:12px;color:{tot_cor};font-size:13px;font-weight:700;white-space:nowrap">{"—" if tot_ant==0 else tot_seta}</td>'
                f'</tr>'
            )

            table_html = (
                f'<div style="background:linear-gradient(135deg,#0a1f0a,#0d2a0d);border:1px solid #1e3a1e;border-radius:14px;padding:0;overflow:hidden">'
                f'<table style="width:100%;border-collapse:collapse">'
                f'{header}'
                f'<tbody>{rows_html}{total_html}</tbody>'
                f'</table></div>'
            )
            st.markdown(table_html, unsafe_allow_html=True)
        st.markdown("---")
    for eq in eqs:
        try:
            ops=buscar_operadores(eq)
            lancs=buscar_lancamentos(ma,eq)
        except Exception as _e_lanc:
            st.error(f"Erro ao conectar ao banco ({eq}): {_e_lanc}. Tente recarregar.")
            continue
        # Mostrar card se tiver lançamento OU base processada no mês
        up=buscar_ultimo_processamento(ma,eq)
        tem_base_mes = up and float(up.get("valorElegivel",0)) > 0 and up.get("mesAno")==ma
        if not lancs and not tem_base_mes: continue
        mg_doc=buscar_meta_gestora(ma,eq); mops=buscar_metas_equipe(ma,eq)
        mg=float(mg_doc.get("metaGestora",0))
        ul=lancs[0] if lancs else {}
        # Para Com Interação, usar o lançamento que tem agentes preenchidos
        ul_agentes=next((l for l in lancs if l.get("agentes")), ul)
        # Com Interação: para Luciano exclui Meet Call; para metcool só Meet Call
        if eq=="luciano":
            tc=sum(float(v.get("valorRecebido",0)) for v in ul_agentes.get("agentes",{}).values() if isinstance(v,dict) and v.get("nome","") not in OPERADORES_MEETCALL)
        elif eq=="metcool":
            tc=sum(float(v.get("valorRecebido",0)) for v in ul.get("agentes",{}).values() if isinstance(v,dict))
        else:
            tc=sum(float(v.get("valorRecebido",0)) for v in ul_agentes.get("agentes",{}).values() if isinstance(v,dict))
        dt=int(ul.get("diasTrabalhados",0)); td=int(ul.get("totalDias",22))
        # Usar o mais recente: comparar data do lançamento manual vs base processada
        rec_geral=0.0
        rec_manual=0.0; dt_manual=None
        for l in lancs:
            if l.get("recGeral",0)>0:
                rec_manual=float(l["recGeral"])
                _cm=l.get("criadoEm")
                dt_manual=_cm.isoformat() if hasattr(_cm,'isoformat') else str(_cm or "")
                break
        rec_base=float(up.get("valorElegivel",0)) if up else 0
        dt_base=up.get("atualizadoEm") if up else None
        # Usar o mais recente — converter tudo para string segura
        def _to_str_dt(v):
            if v is None: return ""
            try:
                if hasattr(v,'isoformat'): return v.isoformat()
                return str(v)
            except: return ""
        if rec_manual>0 and rec_base>0:
            rec_geral=rec_manual if _to_str_dt(dt_manual)>_to_str_dt(dt_base) else rec_base
        elif rec_manual>0:
            rec_geral=rec_manual
        elif rec_base>0:
            rec_geral=rec_base
        # Para metcool: recGeral vem do lancamento_meetcall (recGeralTotal)
        if eq=="metcool":
            mc_doc=buscar_lancamento_meetcall(ma)
            rg_mc=float(mc_doc.get("recGeralTotal",mc_doc.get("recGeral",0)))
            if rg_mc>0: rec_geral=rg_mc

        sem=max(0, rec_geral - tc)
        proj=calc_projecao(rec_geral, dt, td)
        pct=(rec_geral/mg*100) if mg>0 else 0
        cv=cor_pct(pct)
        rg_mc=ci_mc=mg_mc=si_mc=proj_mc=pct_mc=0.0; cv_mc="#e03c3c"; dt_mc=dt; td_mc=td
        if eq=="luciano":
            try:
                ci_mc=sum(float(v.get("valorRecebido",0)) for k,v in ul.get("agentes",{}).items() if isinstance(v,dict) and v.get("nome","") in OPERADORES_MEETCALL)
                lancs_mc=buscar_lancamentos(ma,"metcool")
                if lancs_mc:
                    ul_mc=lancs_mc[0]
                    rg_mc=float(ul_mc.get("totalEquipe",0))
                    dt_mc=int(ul_mc.get("diasTrabalhados",dt)); td_mc=int(ul_mc.get("totalDias",td))
                mg_mc=float(buscar_meta_gestora(ma,"metcool").get("metaGestora",0))
                si_mc=max(0,rg_mc-ci_mc); proj_mc=calc_projecao(rg_mc,dt_mc,td_mc)
                pct_mc=(rg_mc/mg_mc*100) if mg_mc>0 else 0; cv_mc=cor_pct(pct_mc)
            except: pass

        st.markdown(
            f"<div style='background:linear-gradient(135deg,#0a1f0a,#0d2a0d);border:1px solid #1e3a1e;"
            f"border-radius:14px;padding:20px 24px;margin-bottom:6px;border-left:3px solid #00c853'>"
            f"<div style='display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px;margin-bottom:14px'>"
            f"<div style='font-size:15px;font-weight:700;color:#ffffff'>Equipe {EQUIPES[eq]['nome']} · {ul.get('label','')}</div>"
            f"<div style='text-align:right'><div style='color:#3a6a4a;font-size:9px;text-transform:uppercase;letter-spacing:1.5px'>% META</div>"
            f"<div style='color:{cv};font-size:24px;font-weight:800'>{pct:.2f}%</div></div></div>"
            f"<div style='display:flex;gap:28px;flex-wrap:wrap'>"
            f"<div><div style='color:#3a6a4a;font-size:9px;text-transform:uppercase;letter-spacing:1.5px;margin-bottom:3px'>RECEBIDO GERAL</div>"
            f"<div style='color:#00c853;font-weight:700;font-size:16px'>{fmt_brl(rec_geral)}</div></div>"
            f"<div><div style='color:#3a6a4a;font-size:9px;text-transform:uppercase;letter-spacing:1.5px;margin-bottom:3px'>COM INTERAÇÃO</div>"
            f"<div style='color:#ffffff;font-weight:600;font-size:14px'>{fmt_brl(tc)}</div></div>"
            f"<div><div style='color:#3a6a4a;font-size:9px;text-transform:uppercase;letter-spacing:1.5px;margin-bottom:3px'>SEM INTERAÇÃO</div>"
            f"<div style='color:#8ab89a;font-weight:600;font-size:14px'>{fmt_brl(sem)}</div></div>"
            f"<div><div style='color:#3a6a4a;font-size:9px;text-transform:uppercase;letter-spacing:1.5px;margin-bottom:3px'>META</div>"
            f"<div style='color:#8ab89a;font-weight:600;font-size:14px'>{fmt_brl(mg)}</div></div>"
            f"<div><div style='color:#3a6a4a;font-size:9px;text-transform:uppercase;letter-spacing:1.5px;margin-bottom:3px'>PROJEÇÃO</div>"
            f"<div style='color:#8ab89a;font-weight:600;font-size:14px'>{fmt_brl(proj)}</div></div>"
            f"<div><div style='color:#3a6a4a;font-size:9px;text-transform:uppercase;letter-spacing:1.5px;margin-bottom:3px'>DIAS</div>"
            f"<div style='color:#8ab89a;font-weight:600;font-size:14px'>{dt}/{td}</div></div>"
            f"</div>"
            +
            f"</div>", unsafe_allow_html=True)
        # Gráfico de evolução — apenas admin (tamires)
        if u.get('id') == 'tamires':
            from datetime import datetime as _dtm
            import json as _json
            DATA_CORTE = _dtm(2026, 6, 24)
            _lancs_todos = buscar_lancamentos(ma, eq)
            _por_dia = {}
            for l in _lancs_todos:
                dt_lanc = l.get('criadoEm')
                if not dt_lanc: continue
                try:
                    dt_obj = dt_lanc if isinstance(dt_lanc, _dtm) else _dtm.fromisoformat(str(dt_lanc)[:19])
                except: continue
                if dt_obj < DATA_CORTE: continue
                rg = float(l.get('recGeral',0))
                if rg <= 0: continue
                dia = dt_obj.strftime('%d/%m')
                if dia not in _por_dia or dt_obj > _por_dia[dia]['dt']:
                    _por_dia[dia] = {'dt': dt_obj, 'valor': rg}
            _pontos = [{'data': dia, 'valor': v['valor']} for dia, v in sorted(_por_dia.items())]
            if len(_pontos) >= 2:
                _sk = f"chart_exp_{eq}"
                if _sk not in st.session_state: st.session_state[_sk] = False
                _col1, _col2 = st.columns([6,1])
                with _col1:
                    st.markdown("<p style='color:#3a6a4a;font-size:10px;text-transform:uppercase;letter-spacing:1.5px;margin:8px 0 4px'>Evolução dos lançamentos</p>", unsafe_allow_html=True)
                with _col2:
                    if st.button("⛶" if not st.session_state[_sk] else "✕", key=f"btn_exp_{eq}", help="Expandir gráfico"):
                        st.session_state[_sk] = not st.session_state[_sk]; st.rerun()
                _altura = 320 if st.session_state[_sk] else 200
                _labs = _json.dumps([p['data'] for p in _pontos])
                _vals = _json.dumps([p['valor'] for p in _pontos])
                _cid = f"chev_{eq}"
                import streamlit.components.v1 as _components
                _html = (
                    f'<div style="width:100%;height:{_altura}px;margin-bottom:8px;background:transparent">'
                    f'<canvas id="{_cid}"></canvas></div>'
                    '<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>'
                    '<script>(function(){'
                    f'const L={_labs},V={_vals};'
                    'const fmtBrl=v=>"R$ "+v.toLocaleString("pt-BR",{minimumFractionDigits:2,maximumFractionDigits:2});'
                    f'const canvas=document.getElementById("{_cid}");if(!canvas)return;'
                    'new Chart(canvas,{type:"line",data:{labels:L,datasets:[{data:V,'
                    'borderColor:"#2a78d6",backgroundColor:"rgba(42,120,214,0.08)",'
                    'borderWidth:2,pointRadius:6,pointBackgroundColor:"#2a78d6",'
                    'pointBorderColor:"#fff",pointBorderWidth:2,tension:0.3,fill:true}]},'
                    'plugins:[{id:"lbl",afterDatasetsDraw(chart){'
                    'const ctx=chart.ctx,meta=chart.getDatasetMeta(0);ctx.save();'
                    'meta.data.forEach((pt,i)=>{'
                    'const val=V[i],prev=i>0?V[i-1]:null,x=pt.x,y=pt.y;'
                    'ctx.textAlign="center";'
                    'if(prev!==null){'
                    'const diff=val-prev,seta=diff>=0?"\u2191":"\u2193",cor=diff>=0?"#1baf7a":"#e34948";'
                    'ctx.font="500 11px sans-serif";ctx.fillStyle=cor;'
                    'ctx.fillText(seta+" R$ "+Math.abs(diff).toLocaleString("pt-BR",{minimumFractionDigits:2,maximumFractionDigits:2}),x,y-26);}'
                    'ctx.font="500 11px sans-serif";ctx.fillStyle="#2a78d6";'
                    'ctx.fillText(fmtBrl(val),x,y-12);'
                    '});ctx.restore();}}],'
                    'options:{responsive:true,maintainAspectRatio:false,'
                    'layout:{padding:{top:44,left:90,right:90,bottom:4}},'
                    'plugins:{legend:{display:false},tooltip:{enabled:false}},'
                    'scales:{x:{grid:{display:false},border:{display:false},'
                    'ticks:{font:{size:12},color:"#555"}},'
                    'y:{display:false}}}});'
                    '})();</script>'
                )
                _components.html(_html, height=_altura + 10, scrolling=False)

        k=f"show_ops_{eq}"
        if k not in st.session_state: st.session_state[k]=False
        nome_eq_btn = "Meet Call" if eq=="metcool" else EQUIPES[eq]['nome']
        if st.button(f"{'Ocultar' if st.session_state[k] else 'Exibir'} Operadores — {nome_eq_btn}",key=f"btn_ops_{eq}",use_container_width=True):
            st.session_state[k]=not st.session_state[k]; st.rerun()

        show=st.session_state[k]
        if show and ops:
            # Filtrar operadores da equipe (Luciano sem Meet Call)
            ops_show=[op for op in ops if op["nome"] not in OPERADORES_MEETCALL] if eq=="luciano" else ops

            # Montar ranking por % meta
            rows_com_meta=[]
            rows_sem_meta=[]
            for op in ops_show:
                v=get_val_op(ul.get("agentes",{}),op["_id"],op["nome"])
                meta=float(mops.get(op["_id"],0))
                pc=(v/meta*100) if meta>0 else 0
                proj_op=calc_projecao(v,dt,td) if v>0 else 0
                lig_op=int(ul.get("agentes",{}).get(op["_id"],{}).get("ligacoes",0) if isinstance(ul.get("agentes",{}).get(op["_id"]),dict) else 0)
                row={"Operador":op["nome"]+(" ★" if op.get("pleno") else ""),
                     "Recebido":fmt_brl(v) if v>0 else "—",
                     "Meta":fmt_brl(meta) if meta>0 else "—",
                     "% Meta":f"{pc:.2f}%" if meta>0 else "—",
                     "Projeção":fmt_brl(proj_op) if v>0 else "—",
                     "_pc":pc,"_v":v}
                if lig_op>0: row["Ligações"]=lig_op
                if meta>0:
                    rows_com_meta.append(row)
                else:
                    rows_sem_meta.append(row)

            # Ordenar com meta por % desc, sem meta por valor desc
            rows_com_meta.sort(key=lambda x: x["_pc"], reverse=True)
            rows_sem_meta.sort(key=lambda x: x["_v"], reverse=True)
            todos=rows_com_meta+rows_sem_meta

            # Adicionar posição com medalhas
            rows_final=[]
            for i,r in enumerate(todos):
                pos = "🥇" if i==0 else "🥈" if i==1 else "🥉" if i==2 else f"{i+1}º"
                r2={k2:v2 for k2,v2 in r.items() if not k2.startswith("_")}
                r2={"#":pos,**r2}
                rows_final.append(r2)

            if rows_final:
                st.markdown(f"<p style='color:#81c784;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:1px;margin:10px 0 4px'>🏆 Ranking — {EQUIPES.get(eq,{}).get('nome','') if eq!='metcool' else 'Meet Call'}</p>",unsafe_allow_html=True)
                df=pd.DataFrame(rows_final).reset_index(drop=True)
                df.index=range(1,len(df)+1)
                st.dataframe(df,use_container_width=True,height=min(600,(len(df)+1)*38+40))
        if up and up.get("porFornecedora"):
            try:
                pf=up["porFornecedora"]
                if pf:
                    forn_rows=[]; total_forn=0
                    for forn,dados in sorted(pf.items(),key=lambda x:x[1].get("valor",0),reverse=True):
                        val=float(dados.get("valor",0))
                        if val>0:
                            forn_rows.append({"Fornecedora":forn,"Valor Recebido":fmt_brl(val),"Boletos":dados.get("boletos",0)})
                            total_forn+=val
                            # Se tiver breakdown por UF, exibir sub-linhas
                            por_uf=dados.get("porUF",{})
                            for uf,duf in sorted(por_uf.items(),key=lambda x:x[1].get("valor",0),reverse=True):
                                forn_rows.append({"Fornecedora":f"  └ {forn} — {uf}","Valor Recebido":fmt_brl(float(duf.get("valor",0))),"Boletos":duf.get("boletos",0)})
                    if forn_rows:
                        forn_rows.append({"Fornecedora":"TOTAL GERAL","Valor Recebido":fmt_brl(total_forn),"Boletos":""})
                        st.markdown("<p style='color:#3a6a4a;font-size:9px;text-transform:uppercase;letter-spacing:1.5px;margin:12px 0 6px'>POR FORNECEDORA</p>",unsafe_allow_html=True)
                        df_forn=pd.DataFrame(forn_rows); df_forn.index=range(1,len(df_forn)+1)
                        st.dataframe(df_forn,use_container_width=True,hide_index=False)
            except: pass

        st.markdown("---")



# ── MONITORIAS ─────────────────────────────────

def gerar_relatorio_monitorias(eq, ma):
    """Gera relatório de monitorias no formato: analista x semana."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ops = buscar_operadores(eq)
    monitorias = buscar_monitorias_equipe(eq, ma)

    # Buscar também monitorias de operadores vinculados (vinculadoA)
    for op in ops:
        vid = op.get('vinculadoA')
        if vid:
            mons_vinc = list(get_db().monitorias.find({"opId": vid, "mesAno": ma}))
            for m in mons_vinc:
                # Substituir opId/opNome pelo operador da equipe atual
                m['opId'] = op['_id']
                m['opNome'] = op['nome']
            monitorias.extend(mons_vinc)

    if not monitorias:
        return None

    # Organizar por operador e semana — usar exatamente os valores salvos no banco
    semanas = SEMANAS_MONITORIA  # ["1ª Semana — 1ª Monitoria", ...]

    # Mapear monitorias por opId e semana
    dados = {}
    for m in monitorias:
        oid = m.get('opId')
        nome = m.get('opNome','')
        sem = m.get('semana_mon','')
        nota = float(m.get('nota', 0))
        if oid not in dados:
            dados[oid] = {'nome': nome, 'semanas': {}}
        if sem not in dados[oid]['semanas']:
            dados[oid]['semanas'][sem] = []
        dados[oid]['semanas'][sem].append(nota)

    # Criar Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Monitorias {ma}"

    # Cores
    cores_semanas = ["DCE9FF","DCE9FF","FFEFD5","FFEFD5","E8FFE8","E8FFE8","FFE4FF","FFE4FF"]
    verde = "1A3D2B"
    branco = "FFFFFF"
    cinza = "F2F4F3"

    # Cabeçalho linha 1 — Semanas
    ws.merge_cells("A1:A2")
    ws["A1"] = "Analista"
    ws["A1"].font = Font(bold=True, color=branco)
    ws["A1"].fill = PatternFill("solid", start_color=verde)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 30

    semana_grupos = [("1ª Semana", 2, 3), ("2ª Semana", 4, 5), ("3ª Semana", 6, 7), ("4ª Semana", 8, 9)]
    for label, col_ini, col_fim in semana_grupos:
        letra_ini = get_column_letter(col_ini)
        letra_fim = get_column_letter(col_fim)
        ws.merge_cells(f"{letra_ini}1:{letra_fim}1")
        ws[f"{letra_ini}1"] = label
        ws[f"{letra_ini}1"].font = Font(bold=True, color=verde)
        ws[f"{letra_ini}1"].alignment = Alignment(horizontal="center")
        ws[f"{letra_ini}1"].fill = PatternFill("solid", start_color=cores_semanas[col_ini-2])

    # Cabeçalho linha 2 — 1ª Monitoria / 2ª Monitoria
    for i, col in enumerate(range(2, 10)):
        letra = get_column_letter(col)
        ws[f"{letra}2"] = "1ª Monitoria" if i % 2 == 0 else "2ª Monitoria"
        ws[f"{letra}2"].font = Font(bold=True)
        ws[f"{letra}2"].fill = PatternFill("solid", start_color=cores_semanas[i])
        ws[f"{letra}2"].alignment = Alignment(horizontal="center")
        ws.column_dimensions[letra].width = 12

    # MÉDIA
    ws.merge_cells("J1:J2")
    ws["J1"] = "MÉDIA"
    ws["J1"].font = Font(bold=True, color=branco)
    ws["J1"].fill = PatternFill("solid", start_color=verde)
    ws["J1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["J"].width = 10

    # Dados por operador
    row = 3
    medias_semana = {s: [] for s in semanas}
    todas_notas_equipe = []  # todas as notas brutas da equipe para média geral

    for oid, info in sorted(dados.items(), key=lambda x: x[1]['nome']):
        ws[f"A{row}"] = info['nome']
        ws[f"A{row}"].font = Font(bold=False)
        if row % 2 == 0:
            ws[f"A{row}"].fill = PatternFill("solid", start_color=cinza)

        notas_brutas_op = []  # notas brutas do operador para média real
        for i, sem in enumerate(semanas):
            col = i + 2
            letra = get_column_letter(col)
            notas = info['semanas'].get(sem, [])
            if notas:
                media = sum(notas) / len(notas)
                ws[f"{letra}{row}"] = f"{int(round(media))}%"
                ws[f"{letra}{row}"].alignment = Alignment(horizontal="center")
                ws[f"{letra}{row}"].fill = PatternFill("solid", start_color=cores_semanas[i])
                notas_brutas_op.extend(notas)
                medias_semana[sem].append(media)
            else:
                ws[f"{letra}{row}"] = "—"
                ws[f"{letra}{row}"].alignment = Alignment(horizontal="center")
                ws[f"{letra}{row}"].fill = PatternFill("solid", start_color=cores_semanas[i])

        # Média do operador — baseada em todas as notas brutas dele
        if notas_brutas_op:
            media_op = sum(notas_brutas_op) / len(notas_brutas_op)
            ws[f"J{row}"] = f"{int(round(media_op))}%"
            ws[f"J{row}"].font = Font(bold=True)
            ws[f"J{row}"].alignment = Alignment(horizontal="center")
            todas_notas_equipe.extend(notas_brutas_op)

        row += 1

    # Linha média equipe
    ws[f"A{row}"] = "Média Equipe"
    ws[f"A{row}"].font = Font(bold=True, color="2D6A4F")
    ws[f"A{row}"].fill = PatternFill("solid", start_color="D8F3DC")

    for i, sem in enumerate(semanas):
        col = i + 2
        letra = get_column_letter(col)
        if medias_semana[sem]:
            m = sum(medias_semana[sem]) / len(medias_semana[sem])
            ws[f"{letra}{row}"] = f"{int(round(m))}%"
            ws[f"{letra}{row}"].font = Font(bold=True, color="2D6A4F")
            ws[f"{letra}{row}"].fill = PatternFill("solid", start_color="D8F3DC")
            ws[f"{letra}{row}"].alignment = Alignment(horizontal="center")
        else:
            ws[f"{letra}{row}"] = "—"
            ws[f"{letra}{row}"].fill = PatternFill("solid", start_color="D8F3DC")
            ws[f"{letra}{row}"].alignment = Alignment(horizontal="center")

    # Média geral da equipe — média das médias dos operadores
    medias_ops = [sum(n for notas in info["semanas"].values() for n in notas) / max(1, sum(len(notas) for notas in info["semanas"].values())) for oid, info in dados.items() if any(info["semanas"].values())]
    if medias_ops:
        media_geral = sum(medias_ops) / len(medias_ops)
    if medias_ops:
        ws[f"J{row}"] = f"{int(round(media_geral))}%"
        ws[f"J{row}"].font = Font(bold=True, color="2D6A4F")
        ws[f"J{row}"].fill = PatternFill("solid", start_color="D8F3DC")
        ws[f"J{row}"].alignment = Alignment(horizontal="center")

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def pagina_monitorias(ma):
    u=st.session_state.usuario
    header_page("Monitorias","Avaliação de qualidade")
    if u["role"]=="diretor": pagina_monitorias_diretor(ma); return
    eq=seletor_equipe(u["equipe"])
    ops_todos=buscar_operadores(eq)
    # Luciano: mostrar só operadores iGreen (sem Meet Call) — igual Minha Conta
    ops=[op for op in ops_todos if op["nome"] not in OPERADORES_MEETCALL] if eq=="luciano" else ops_todos
    if not ops: st.warning("Cadastre operadores primeiro."); return

    # Botão download relatório — sempre visível no topo
    rel = gerar_relatorio_monitorias(eq, ma)
    if rel:
        st.download_button(
            "⬇️ Baixar Relatório de Monitorias (.xlsx)",
            rel,
            file_name=f"monitorias_{eq}_{ma}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_rel_mon_topo"
        )

    if "mon_op_sel" not in st.session_state: st.session_state.mon_op_sel=None
    if "mon_modo" not in st.session_state: st.session_state.mon_modo=None
    if st.session_state.mon_op_sel is None:
        ultimo=st.session_state.pop("mon_ultimo_salvo",None)
        if ultimo:
            st.success(f"Monitoria salva! {ultimo['nome']} — Nota: {ultimo['nota']:.0f}% | Média: {ultimo['media']:.2f}% | Pontos: {ultimo['pontos']}")
            st.markdown(f'<a href="data:text/html;base64,{ultimo["b64"]}" download="Mon_{ultimo["nome"].replace(" ","_")}.html" style="display:inline-block;background:#1a3a1a;color:#a0c4a0;border:1px solid #2a4a2a;padding:6px 14px;border-radius:6px;text-decoration:none;font-size:12px;margin-bottom:12px">Baixar PDF</a>',unsafe_allow_html=True)
        monts_eq=buscar_monitorias_equipe(eq,ma)
        if monts_eq:
            medias_eq=[calc_media_operador(op["_id"],ma)[0] for op in ops if calc_media_operador(op["_id"],ma)[1]>0]
            if medias_eq:
                me_eq=sum(medias_eq)/len(medias_eq)
                st_txt,st_cor,_=get_status_media(me_eq)
                st.markdown(
                    f"<div style='background:#0a1a0a;border:1px solid #1e3a1e;border-radius:10px;"
                    f"padding:12px 20px;margin-bottom:16px;display:flex;justify-content:space-between;align-items:center'>"
                    f"<div><div style='color:#3a6a4a;font-size:9px;text-transform:uppercase;letter-spacing:1.5px'>MEDIA DA EQUIPE — {ma.replace('-',' ').upper()}</div>"
                    f"<div style='color:{st_cor};font-size:22px;font-weight:800;margin-top:2px'>{me_eq:.2f}%</div></div>"
                    f"<div style='text-align:right'><div style='color:#3a6a4a;font-size:9px;text-transform:uppercase'>STATUS</div>"
                    f"<div style='color:{st_cor};font-size:13px;font-weight:600'>{st_txt}</div></div>"
                    f"</div>",unsafe_allow_html=True)
        st.markdown(f"<div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:16px'><div style='color:#81c784;font-size:12px;text-transform:uppercase;letter-spacing:1px;font-weight:600'>Selecione um operador</div><div style='color:#a5d6a7;font-size:12px'>{len(ops)} operadores · {ma.replace('-',' ')}</div></div>",unsafe_allow_html=True)
        for i in range(0,len(ops),4):
            cols=st.columns(4)
            for j,op in enumerate(ops[i:i+4]):
                media,n=calc_media_operador(op["_id"],ma)
                st_txt,st_cor,st_bg=get_status_media(media)
                ini=get_iniciais(op["nome"]); cini=get_cor_inicial(op["nome"])
                with cols[j]:
                    pontos_op=calc_pontos(media)
                    st.markdown(f"""<div style="background:#ffffff;border:1px solid #c8e0c8;border-radius:12px;padding:16px;text-align:center;margin-bottom:8px;box-shadow:0 1px 4px rgba(0,0,0,0.06)">
                        <div style="width:44px;height:44px;background:{cini};border-radius:50%;display:inline-flex;align-items:center;justify-content:center;color:white;font-weight:700;font-size:15px;margin-bottom:8px">{ini}</div>
                        <div style="color:#1a2e1a;font-weight:700;font-size:12px;margin-bottom:4px">{op['nome']}{'  ★' if op.get('pleno') else ''}</div>
                        <div style="color:{st_cor};font-size:20px;font-weight:800">{round(media)}%</div>
                        <div style="color:#5a8a5a;font-size:10px">{n} monitoria{'s' if n!=1 else ''}</div>
                        <div style="color:#2e7d32;font-size:11px;font-weight:600;margin-top:2px">{pontos_op} pts</div>
                    </div>""",unsafe_allow_html=True)
                    c1,c2=st.columns(2)
                    with c1:
                        if st.button("+ Nova",key=f"nova_{op['_id']}",use_container_width=True):
                            st.session_state.mon_op_sel=op; st.session_state.mon_modo="nova"; st.rerun()
                    with c2:
                        if st.button("Histórico",key=f"hist_{op['_id']}",use_container_width=True):
                            st.session_state.mon_op_sel=op; st.session_state.mon_modo="historico"; st.rerun()
        return
    op=st.session_state.mon_op_sel
    media_op,n_op=calc_media_operador(op["_id"],ma)
    if st.button("← Voltar"): st.session_state.mon_op_sel=None; st.session_state.mon_modo=None; st.rerun()
    st.markdown(f"<div style='background:#e8f5e9;border:1px solid #c8e0c8;border-radius:8px;padding:10px 16px;margin-bottom:12px'><span style='color:#2e7d32;font-weight:700;font-size:15px'>👤 {op['nome']}</span><span style='color:#5a8a5a;font-size:12px;margin-left:12px'>Média: {media_op:.0f}% · {n_op} monitoria{'s' if n_op!=1 else ''}</span></div>",unsafe_allow_html=True)
    st.markdown("---")
    t1,t2=st.tabs(["Nova Monitoria","Monitorias do Mês"])
    with t1:
        # Verificar quais semanas já foram registradas para este operador/mês
        monts_op_mes=[m for m in buscar_monitorias_equipe(eq,ma) if m["opId"]==op["_id"]]
        semanas_usadas={m.get("semana_mon","") for m in monts_op_mes}

        # Mostrar seletor com indicação de semanas já usadas
        semanas_opts=[]
        for s in SEMANAS_MONITORIA:
            if s in semanas_usadas:
                semanas_opts.append(f"🔴 {s} — JÁ REGISTRADA")
            else:
                semanas_opts.append(f"✅ {s}")

        semana_sel=st.selectbox("Qual monitoria é esta?",semanas_opts,key="semana_sel")
        semana=SEMANAS_MONITORIA[semanas_opts.index(semana_sel)]
        semana_bloqueada = semana in semanas_usadas

        if semana_bloqueada:
            st.error(f"⛔ A **{semana}** já foi registrada para {op['nome']} em {ma.replace('-',' ')}. Escolha outra semana ou edite a existente na aba 'Monitorias do Mês'.")

        tipo_mon=st.radio("Tipo de Monitoria",["📞 Ligação","💬 Chat"],horizontal=True,key="tipo_mon_sel")
        tipo_key="ligacao" if "Ligação" in tipo_mon else "chat"
        prot_label="Protocolo da Ligação" if tipo_key=="ligacao" else "ID/Protocolo do Chat"
        prot=st.text_input(prot_label,placeholder="Ex: 20260520-001",key="prot_input")
        obs=st.text_area("Observações",placeholder="Anotações...",height=70,key="obs_input")
        st.markdown("---")
        # Carregar critérios e erros conforme tipo
        crits_usar = get_criterios() if tipo_key=="ligacao" else get_criterios_chat()
        erros_usar = get_erros_criticos() if tipo_key=="ligacao" else get_erros_criticos_chat()
        erros_m=[]; c1,c2=st.columns(2)
        for i,ec in enumerate(erros_usar):
            with (c1 if i%2==0 else c2):
                if st.checkbox(f"{ec['nome']}",key=f"ec_{ec['id']}_{tipo_key}"): erros_m.append(ec)
        st.markdown("---")
        zerada=len(erros_m)>0; crits_r=[]; nota=0 if zerada else 100
        if zerada:
            st.error("MONITORIA ZERADA — Erro crítico marcado!")
            for c in crits_usar: crits_r.append({**c,"passou":False})
        else:
            st.markdown("<p style='color:#e53935;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:1px;margin-bottom:8px'>ITENS DE QUALIDADE — MARQUE O QUE NÃO FOI FEITO</p>",unsafe_allow_html=True)
            for crit in crits_usar:
                c1,c2=st.columns([8,1])
                with c1: nao_passou=st.checkbox(f"{crit['nome']}",key=f"cr_{crit['id']}_{tipo_key}",value=False)
                with c2: st.markdown(f"<div style='padding-top:6px;color:#e53935;font-size:12px;font-weight:600;text-align:right'>−{crit['peso']} pts</div>",unsafe_allow_html=True)
                if crit.get('itens'):
                    for it in crit['itens']:
                        cor_it = "#f87171" if "obrigatório" in it.lower() or "!" in it else "#34d399"
                        st.markdown(f"<div style='padding:3px 0 3px 24px;font-size:12px;color:{cor_it};line-height:1.5'>• {it}</div>",unsafe_allow_html=True)
                passou=not nao_passou
                if not passou: nota-=crit["peso"]
                crits_r.append({**crit,"passou":passou})
        nota=max(0,nota)
        pontos_perdidos=100-nota
        cn="#2e7d32" if nota>=80 else "#f57f17" if nota>=60 else "#c62828"
        st.markdown(
            f"<div style='background:#f0f7f0;border:1px solid #c8e0c8;border-radius:10px;padding:14px 20px;margin-top:16px;display:flex;justify-content:space-between;align-items:center'>"
            f"<div><div style='color:#5a8a5a;font-size:11px'>Pontuação final (máx. 100 pts)</div>"
            f"<div style='color:#5a8a5a;font-size:11px'>Pontos perdidos: {pontos_perdidos}</div></div>"
            f"<div style='color:{cn};font-size:36px;font-weight:800'>{round(nota)}</div>"
            f"</div>",unsafe_allow_html=True)
        sk_salvo=f"mon_salvo_{op['_id']}_{semana}_{ma}_{tipo_key}"
        if not st.session_state.get(sk_salvo):
            if st.button("Salvar Monitoria",use_container_width=True,key="btn_salvar_mon",disabled=semana_bloqueada):
                if not prot.strip(): st.error(f"Preencha o {prot_label}!")
                else:
                    salvar_monitoria(eq,op["_id"],op["nome"],prot,obs,crits_r,erros_m,nota,ma,semana=semana,tipo=tipo_key)
                    mm,nm=calc_media_operador(op["_id"],ma)
                    html=gerar_pdf_monitoria(op["nome"],prot,obs,crits_r,erros_m,nota,mm,nm,ma)
                    b64=base64.b64encode(html.encode()).decode()
                    st.session_state[sk_salvo]={"nome":op["nome"],"nota":nota,"media":mm,"pontos":calc_pontos(mm),"b64":b64,"prot":prot}
        else:
            salvo=st.session_state[sk_salvo]
            cn2="#2e7d32" if salvo['nota']>=80 else "#f57f17" if salvo['nota']>=60 else "#c62828"
            st.markdown(
                f"<div style='background:#0a1a0a;border:2px solid #00c853;border-radius:12px;padding:20px 24px;margin:16px 0'>"
                f"<div style='color:#00c853;font-weight:700;font-size:15px;margin-bottom:8px'>✓ Monitoria salva!</div>"
                f"<div style='color:#e8f5e9;font-size:13px'>Operador: <strong>{salvo['nome']}</strong></div>"
                f"<div style='color:#e8f5e9;font-size:13px'>Nota: <strong style='color:{cn2}'>{salvo['nota']:.0f}%</strong> | "
                f"Média: <strong>{salvo['media']:.2f}%</strong> | Pontos: <strong>{salvo['pontos']}</strong></div>"
                f"</div>",unsafe_allow_html=True)
            st.markdown(
                f'<a href="data:text/html;base64,{salvo["b64"]}" '
                f'download="Monitoria_{salvo["nome"].replace(" ","_")}_{salvo["prot"]}.html" '
                f'style="display:inline-block;background:#1a3a1a;color:#a0c4a0;border:1px solid #2a4a2a;'
                f'padding:10px 24px;border-radius:6px;text-decoration:none;font-weight:600;font-size:13px;margin-bottom:12px">'
                f'⬇ Baixar PDF da Monitoria</a>',
                unsafe_allow_html=True)
            st.markdown("<div style='height:8px'></div>",unsafe_allow_html=True)
            if st.button("Concluir e Voltar",use_container_width=True,key="btn_concluir_mon"):
                ultimo=st.session_state.pop(sk_salvo,None)
                st.session_state.mon_op_sel=None
                st.session_state.mon_modo=None
                if ultimo: st.session_state["mon_ultimo_salvo"]=ultimo
                st.rerun()
    with t2:
        monts2=buscar_monitorias_operador(op["_id"])
        monts2=[m for m in monts2 if m.get("mesAno")==ma]
        if not monts2: st.info(f"Nenhuma monitoria para {op['nome']} em {ma.replace('-',' ')}.")
        else:
            ordem_semanas={s:i for i,s in enumerate(SEMANAS_MONITORIA)}
            monts2=sorted(monts2,key=lambda x:ordem_semanas.get(x.get("semana_mon",""),99))
            for m in monts2:
                nm=float(m.get("nota",0)); cm="#2e7d32" if nm>=80 else "#f57f17" if nm>=60 else "#c62828"
                st.markdown(f"<div style='background:#f8fdf8;border:1px solid #c8e0c8;border-radius:10px;padding:14px 18px;margin-bottom:8px;border-left:3px solid {cm}'><div style='color:#1a2e1a;font-weight:600'>{m.get('semana_mon','—')}</div><div style='color:#5a8a5a;font-size:11px'>Protocolo: {m.get('protocolo','—')} · {str(m.get('criadoEm',''))[:10]}</div><div style='color:{cm};font-size:18px;font-weight:800'>{int(nm)}%</div></div>",unsafe_allow_html=True)
                with st.expander("Ver detalhes"):
                    for c in m.get("criterios",[]):
                        passou=c.get("passou",True); cc="#2e7d32" if passou else "#c62828"
                        st.markdown(f"<div style='display:flex;justify-content:space-between;padding:6px 12px;background:#f0f7f0;border-radius:6px;margin-bottom:4px;border-left:3px solid {cc}'><span style='color:#1a2e1a;font-size:12px'>{c.get('num','')} {c.get('nome','')}</span><span style='color:{cc};font-weight:600;font-size:12px'>{'Passou' if passou else 'Não passou'}</span></div>",unsafe_allow_html=True)
                    if m.get("observacao"): st.markdown(f"<div style='padding:8px 12px;background:#f0f7f0;border-radius:6px;border-left:3px solid #5a8a5a;color:#2d4a2d;font-size:12px'><strong>Obs:</strong> {m['observacao']}</div>",unsafe_allow_html=True)
                    mm2,nm2=calc_media_operador(op["_id"],ma)
                    hp=gerar_pdf_monitoria(op["nome"],m.get("protocolo",""),m.get("observacao",""),m.get("criterios",[]),m.get("errosCriticos",[]),nm,mm2,nm2,ma)
                    b64h=base64.b64encode(hp.encode()).decode()
                    st.markdown(f'<a href="data:text/html;base64,{b64h}" download="Mon_{op["nome"].replace(" ","_")}_{m.get("semana_mon","").replace(" ","_").replace("—","")}.html" style="display:inline-block;background:#1a3a1a;color:#a0c4a0;border:1px solid #2a4a2a;padding:5px 12px;border-radius:5px;text-decoration:none;font-size:12px;margin-top:6px">⬇ Baixar PDF</a>',unsafe_allow_html=True)

                    # Edição completa
                    st.markdown("---")
                    st.markdown("<p style='color:#5a8a5a;font-size:11px;font-weight:600'>✏️ EDITAR MONITORIA</p>",unsafe_allow_html=True)
                    ed_semana=st.selectbox("Semana",SEMANAS_MONITORIA,index=SEMANAS_MONITORIA.index(m.get("semana_mon",SEMANAS_MONITORIA[0])) if m.get("semana_mon") in SEMANAS_MONITORIA else 0,key=f"ed_sem_{m['_id']}")
                    ed_prot=st.text_input("Protocolo",value=m.get("protocolo",""),key=f"ed_prot_{m['_id']}")
                    ed_obs=st.text_area("Observações",value=m.get("observacao",""),height=60,key=f"ed_obs_{m['_id']}")

                    # Edição dos critérios
                    st.markdown("<p style='color:#e53935;font-size:11px;font-weight:600;margin-top:8px'>CRITÉRIOS — MARQUE O QUE NÃO FOI FEITO</p>",unsafe_allow_html=True)
                    tipo_ed = m.get("tipo","ligacao")
                    crits_usar_ed = get_criterios() if tipo_ed!="chat" else get_criterios_chat()
                    erros_usar_ed = get_erros_criticos() if tipo_ed!="chat" else get_erros_criticos_chat()

                    # Erros críticos
                    erros_ed=[]; c1e,c2e=st.columns(2)
                    for i,ec in enumerate(erros_usar_ed):
                        ja_marcado=any(e.get("id")==ec["id"] for e in m.get("errosCriticos",[]))
                        with (c1e if i%2==0 else c2e):
                            if st.checkbox(f"{ec['nome']}",value=ja_marcado,key=f"ed_ec_{m['_id']}_{ec['id']}"): erros_ed.append(ec)

                    zerada_ed=len(erros_ed)>0
                    crits_ed=[]; nota_ed=0 if zerada_ed else 100
                    if zerada_ed:
                        st.error("MONITORIA ZERADA — Erro crítico marcado!")
                        for c in crits_usar_ed: crits_ed.append({**c,"passou":False})
                    else:
                        for crit in crits_usar_ed:
                            ja_passou=next((c.get("passou",True) for c in m.get("criterios",[]) if c.get("id")==crit["id"]),True)
                            c1c,c2c=st.columns([8,1])
                            with c1c: nao_passou_ed=st.checkbox(f"{crit['nome']}",value=not ja_passou,key=f"ed_cr_{m['_id']}_{crit['id']}")
                            with c2c: st.markdown(f"<div style='padding-top:6px;color:#e53935;font-size:12px;font-weight:600;text-align:right'>−{crit['peso']}</div>",unsafe_allow_html=True)
                            passou_ed=not nao_passou_ed
                            if not passou_ed: nota_ed-=crit["peso"]
                            crits_ed.append({**crit,"passou":passou_ed})
                    nota_ed=max(0,nota_ed)
                    cn_ed="#2e7d32" if nota_ed>=80 else "#f57f17" if nota_ed>=60 else "#c62828"
                    st.markdown(f"<div style='background:#f0f7f0;border-radius:8px;padding:10px 16px;display:flex;justify-content:space-between;align-items:center;margin-top:8px'><span style='color:#5a8a5a;font-size:12px'>Nova nota</span><span style='color:{cn_ed};font-size:28px;font-weight:800'>{round(nota_ed)}</span></div>",unsafe_allow_html=True)

                    if st.button("💾 Salvar Edição Completa",key=f"ed_save_{m['_id']}",use_container_width=True):
                        get_db().monitorias.update_one(
                            {"_id":m["_id"]},
                            {"$set":{
                                "semana_mon":ed_semana,
                                "protocolo":ed_prot,
                                "observacao":ed_obs,
                                "criterios":crits_ed,
                                "errosCriticos":erros_ed,
                                "nota":nota_ed
                            }}
                        )
                        st.success(f"✅ Monitoria atualizada! Nova nota: {round(nota_ed)}%")
                        st.rerun()

                c1x,c2x=st.columns(2)
                with c2x:
                    if st.button("Excluir",key=f"del_op_{m['_id']}"): excluir_monitoria(m["_id"]); st.rerun()

def pagina_monitorias_diretor(ma):
    if "dir_op_sel" not in st.session_state: st.session_state.dir_op_sel=None
    if "dir_eq_sel" not in st.session_state: st.session_state.dir_eq_sel=None
    if st.session_state.dir_op_sel:
        op=st.session_state.dir_op_sel; eq=st.session_state.dir_eq_sel
        media_op,n_op=calc_media_operador(op["_id"],ma)
        st_txt,st_cor,_=get_status_media(media_op)
        if st.button("← Voltar"):
            st.session_state.dir_op_sel=None; st.session_state.dir_eq_sel=None; st.rerun()
        st.markdown(f"<div style='background:#0a1a0a;border:1px solid #1e3a1e;border-radius:12px;padding:16px 20px;margin-bottom:16px'><div style='color:#fff;font-weight:700;font-size:16px'>{op['nome']}</div><div style='color:#3a6a4a;font-size:12px'>Equipe {EQUIPES.get(eq,{}).get('nome','—')} · {ma.replace('-',' ')} · Média: <strong style='color:{st_cor}'>{media_op:.2f}%</strong></div></div>",unsafe_allow_html=True)
        monts_op=[m for m in buscar_monitorias_equipe(eq,ma) if m["opId"]==op["_id"]]
        if not monts_op: st.info("Nenhuma monitoria registrada neste mês.")
        else:
            for m in monts_op:
                nm=float(m.get("nota",0)); cm="#2e7d32" if nm>=80 else "#f57f17" if nm>=60 else "#c62828"
                st.markdown(f"<div style='background:#0a1a0a;border:1px solid #1e3a1e;border-radius:10px;padding:14px 18px;margin-bottom:8px;border-left:3px solid {cm}'><div style='color:#fff;font-weight:600'>{m.get('semana_mon','—')}</div><div style='color:#3a6a4a;font-size:11px'>Protocolo: {m.get('protocolo','—')} · {str(m.get('criadoEm',''))[:10]}</div><div style='color:{cm};font-size:18px;font-weight:800'>{nm:.0f}%</div></div>",unsafe_allow_html=True)
                with st.expander("Ver detalhes"):
                    for c in m.get("criterios",[]):
                        passou=c.get("passou",True); cc="#2e7d32" if passou else "#c62828"
                        st.markdown(f"<div style='display:flex;justify-content:space-between;padding:6px 12px;background:#0a1a0a;border-radius:6px;margin-bottom:4px;border-left:3px solid {cc}'><span style='color:#e8f5e9;font-size:12px'>{c.get('num','')} {c.get('nome','')}</span><span style='color:{cc};font-weight:600;font-size:12px'>{'Passou' if passou else 'Nao passou'}</span></div>",unsafe_allow_html=True)
                    if m.get("observacao"): st.markdown(f"<div style='padding:8px 12px;background:#0a1a0a;border-radius:6px;border-left:3px solid #3a6a4a;color:#8ab89a;font-size:12px'><strong>Obs:</strong> {m['observacao']}</div>",unsafe_allow_html=True)
                    hp=gerar_pdf_monitoria(m["opNome"],m.get("protocolo",""),m.get("observacao",""),m.get("criterios",[]),m.get("errosCriticos",[]),nm,media_op,n_op,ma)
                    b64=base64.b64encode(hp.encode()).decode()
                    st.markdown(f'<a href="data:text/html;base64,{b64}" download="Mon_{m["opNome"].replace(" ","_")}.html" style="display:inline-block;background:#1a3a1a;color:#a0c4a0;border:1px solid #2a4a2a;padding:5px 12px;border-radius:5px;text-decoration:none;font-size:12px">Baixar PDF</a>',unsafe_allow_html=True)
        return
    # Calcular médias por equipe para mostrar no header
    linhas_eq=""; todas_medias_top=[]; eqs_mon=["luciano","deborah","tamires"]
    eq_medias_list=[]
    for eq_pre in eqs_mon:
        ops_pre=buscar_operadores(eq_pre)
        medias_pre=[calc_media_operador(op["_id"],ma)[0] for op in ops_pre if calc_media_operador(op["_id"],ma)[1]>0]
        if medias_pre:
            me_pre=sum(medias_pre)/len(medias_pre)
            nome_eq=EQUIPES.get(eq_pre,{}).get("nome",eq_pre)
            eq_medias_list.append((nome_eq,me_pre))
            todas_medias_top.append(me_pre)
    # Ordenar por melhor média
    eq_medias_list.sort(key=lambda x: x[1], reverse=True)
    for nome_eq,me_pre in eq_medias_list:
        cor_me=cor_pct(me_pre)
        linhas_eq+=f"<div style='display:flex;justify-content:space-between;align-items:center;padding:4px 0;border-bottom:1px solid #e8f0e8'><span style='color:#2d4a2d;font-size:13px;font-weight:500'>{nome_eq}</span><span style='color:{cor_me};font-size:15px;font-weight:700'>{me_pre:.2f}%</span></div>"

    if todas_medias_top:
        mg_top=sum(todas_medias_top)/len(todas_medias_top)
        cor_mg=cor_pct(mg_top)
        linhas_eq+=f"<div style='border-top:1px solid #c8e0c8;margin-top:6px;padding-top:8px;display:flex;justify-content:space-between;align-items:center'><span style='color:#2d4a2d;font-size:13px;font-weight:700;text-transform:uppercase;letter-spacing:1px'>Média Geral</span><span style='color:{cor_mg};font-size:22px;font-weight:800'>{mg_top:.2f}%</span></div>"

    if linhas_eq:
        st.markdown(
            f"<div style='background:#ffffff;border:1px solid #c8e0c8;border-radius:12px;padding:16px 24px;"
            f"margin-bottom:20px;border-left:4px solid #2e7d32;box-shadow:0 2px 8px rgba(0,0,0,0.06)'>"
            f"{linhas_eq}</div>",
            unsafe_allow_html=True)

    st.markdown("### Visão Geral — Monitorias por Equipe")
    todas_medias=[]
    for eq in EQUIPES:
        ops=buscar_operadores(eq)
        if not ops: continue
        monts=buscar_monitorias_equipe(eq,ma)
        if not monts: continue
        medias={op["nome"]:(op,calc_media_operador(op["_id"],ma)) for op in ops}
        medias={k:v for k,v in medias.items() if v[1][1]>0}
        if not medias: continue
        me=sum(v[1][0] for v in medias.values())/len(medias)
        todas_medias.append(me)
        st.markdown(f"<div style='background:#0a1a0a;border:1px solid #1e3a1e;border-radius:12px;padding:16px 20px;margin-bottom:8px;border-left:3px solid #00c853'><div style='display:flex;justify-content:space-between;align-items:center'><div style='font-size:15px;font-weight:700;color:#fff'>Equipe {EQUIPES[eq]['nome']}</div><div style='text-align:right'><div style='color:#3a6a4a;font-size:10px;text-transform:uppercase'>MEDIA DA EQUIPE</div><div style='color:{cor_pct(me)};font-size:24px;font-weight:800'>{me:.2f}%</div></div></div></div>",unsafe_allow_html=True)
        cols_op=st.columns(4)
        for idx_op,(nome,(op_obj,(media,n))) in enumerate(sorted(medias.items(),key=lambda x:-x[1][1][0])):
            st_txt,st_cor,_=get_status_media(media)
            with cols_op[idx_op%4]:
                st.markdown(f"<div style='background:#0d1a0d;border:1px solid #1e3a1e;border-radius:10px;padding:12px;text-align:center;margin-bottom:8px'><div style='color:#fff;font-weight:600;font-size:12px'>{nome}</div><div style='color:{st_cor};font-size:18px;font-weight:800'>{media:.2f}%</div><div style='color:#3a6a4a;font-size:10px'>{n} monitoria{'s' if n!=1 else ''}</div></div>",unsafe_allow_html=True)
                if st.button("Ver detalhes",key=f"dir_op_{op_obj['_id']}",use_container_width=True):
                    st.session_state.dir_op_sel=op_obj; st.session_state.dir_eq_sel=eq; st.rerun()
        st.markdown("---")


# ── ANÁLISE DOS OPERADORES ─────────────────────
def pagina_analise_operadores(ma):
    u=st.session_state.usuario
    eqs=list(EQUIPES.keys()) if u["role"] in ["diretor","diretor_upload","admin"] else [u["equipe"]]
    header_page("Análise dos Operadores",f"Resultado comparativo · {ma.replace('-',' ')}")
    idx=MESES_NOMES.index(ma.split("-")[0]); ano=int(ma.split("-")[1])
    ma_ant=f"{MESES_NOMES[11]}-{ano-1}" if idx==0 else f"{MESES_NOMES[idx-1]}-{ano}"
    st.markdown("---")
    for eq in eqs:
        ops_todos=buscar_operadores(eq)
        # Luciano: excluir Meet Call da análise
        ops=[op for op in ops_todos if op["nome"] not in OPERADORES_MEETCALL] if eq=="luciano" else ops_todos
        if not ops: continue
        lat=buscar_lancamentos(ma,eq); lan=buscar_lancamentos(ma_ant,eq)
        if not lat: continue
        ul=lat[0]; ul_an=lan[0] if lan else None
        mops=buscar_metas_equipe(ma,eq)
        rows=[]
        for op in ops:
            vat=get_val_op(ul.get("agentes",{}),op["_id"],op["nome"])
            van=get_val_op(ul_an.get("agentes",{}),op["_id"],op["nome"]) if ul_an else 0
            meta=float(mops.get(op["_id"],0))
            pct=(vat/meta*100) if meta>0 else 0
            var_op=calc_variacao(vat,van)
            sv="↑" if (var_op or 0)>=0 else "↓"
            rows.append({"Operador":("★ " if op.get("pleno") else "")+op["nome"],"Recebido":fmt_brl(vat) if vat>0 else "—","Meta":fmt_brl(meta) if meta>0 else "—","% Meta":f"{pct:.2f}%" if meta>0 else "—","Projeção":fmt_brl(calc_projecao(vat,int(ul.get("diasTrabalhados",0)),int(ul.get("totalDias",22)))) if vat>0 else "—","Mês Ant.":fmt_brl(van) if van>0 else "—","Variação":f"{sv} {abs(var_op):.2f}%" if var_op is not None else "—","Monitoria":f"{calc_media_operador(op['_id'],ma)[0]:.1f}%" if calc_media_operador(op['_id'],ma)[1]>0 else "—","_v":vat})
        st.markdown(f"**Equipe {EQUIPES[eq]['nome']}**")
        df=pd.DataFrame(rows).sort_values("_v",ascending=False).drop(columns=["_v"]).reset_index(drop=True)
        df.index=range(1,len(df)+1)
        st.dataframe(df,use_container_width=True)
        st.markdown("---")

# ── VISUALIZAÇÃO RCA ───────────────────────────
def pagina_dashboard_executivo():
    header_page("Visualização RCA","Gestão de Inadimplência Comercial")
    mp=listar_meses_processados()
    if not mp: st.info("Nenhuma base processada ainda."); return
    c1,c2,c3=st.columns(3)
    with c1: mf=st.selectbox("Mês",["Todos"]+mp)
    with c2: ef=st.selectbox("Equipe",["Todas","luciano","deborah","tamires"])
    df=buscar_processamentos(None if mf=="Todos" else mf, None if ef=="Todas" else ef)
    if df.empty: st.warning("Nenhum dado."); return
    df["valor"]=pd.to_numeric(df["valor"],errors="coerce").fillna(0)
    elig=df[df["elegibilidade"]=="Elegível"] if "elegibilidade" in df.columns else df
    with c3:
        forns=["Todas"]+sorted(df["fornecedora"].dropna().unique().tolist())
        ff=st.selectbox("Fornecedora",forns)
    if ff!="Todas":
        df=df[df["fornecedora"]==ff]
        elig=elig[elig["fornecedora"]==ff] if "fornecedora" in elig.columns else elig
    st.markdown("---")
    val_rec=float(elig["valor"].sum()) if not elig.empty else 0
    cli_unic=int(elig["uc_cpf"].nunique()) if "uc_cpf" in elig.columns and not elig.empty else 0
    tot_bol=len(elig)
    c1,c2,c3=st.columns(3)
    c1.metric("Valor Recuperado",fmt_brl(val_rec))
    c2.metric("Clientes Únicos",f"{cli_unic:,}")
    c3.metric("Total Boletos",f"{tot_bol:,}")
    st.markdown("---")
    t1,t2,t3,t4=st.tabs(["Aging","Fornecedoras","Evolução","Por Equipe"])
    with t1:
        if "aging" in df.columns:
            ag=df.groupby("aging").agg(Boletos=("uc_cpf","count"),Valor=("valor","sum")).reset_index()
            ag["Valor"]=ag["Valor"].apply(fmt_brl)
            st.dataframe(ag.rename(columns={"aging":"Faixa"}),use_container_width=True,hide_index=True)
    with t2:
        if "fornecedora" in df.columns:
            fdf=df.groupby("fornecedora").agg(Boletos=("uc_cpf","count"),Valor=("valor","sum")).reset_index()
            fdf=fdf.sort_values("Valor",ascending=False)
            fdf["Valor"]=fdf["Valor"].apply(fmt_brl)
            st.dataframe(fdf.rename(columns={"fornecedora":"Fornecedora"})[["Fornecedora","Boletos","Valor"]],use_container_width=True,hide_index=True)
    with t3:
        da=buscar_processamentos()
        if not da.empty:
            da["valor"]=pd.to_numeric(da["valor"],errors="coerce").fillna(0)
            da_elig=da[da["elegibilidade"]=="Elegível"] if "elegibilidade" in da.columns else da
            if "_mes_ano" in da_elig.columns:
                ev=da_elig.groupby("_mes_ano")["valor"].sum().reset_index()
                ev.columns=["Mês","Valor"]
                st.bar_chart(ev.sort_values("Mês").set_index("Mês"),color="#2daf5c")
    with t4:
        if "_equipe" in df.columns:
            ed=df.groupby("_equipe").agg(Boletos=("uc_cpf","count"),Valor=("valor","sum")).reset_index()
            ed["Equipe"]=ed["_equipe"].map(lambda x:EQUIPES.get(x,{}).get("nome",x))
            ed["Valor"]=ed["Valor"].apply(fmt_brl)
            st.dataframe(ed[["Equipe","Boletos","Valor"]],use_container_width=True,hide_index=True)



def calcular_divisao_proporcional_luciano(df_eleg, arq_interacoes, ma):
    """Calcula divisão proporcional Luciano vs Amitycall usando df já processado."""
    import unicodedata
    def norm(s): return unicodedata.normalize('NFKD',str(s).upper().strip()).encode('ascii','ignore').decode()

    AGENTES_LUCIANO = {n.upper().strip() for n in [
        'JHENIFFER SANTOS','MARCOS MARTINS','JUNIOR OTAIDES','CAMILA NARA',
        'MICHELLE BATISTA','LORENZZO PEREIRA','EDUARDA SANQUETA','MARIA CLARA',
        'HEVERTON TAVARES','DIOGO OLIVEIRA','GRASIELLE DA SILVA SANTOS',
        'EMANUEL FERREIRA','KETLE SILVA','CAUA ALVES','VICTORIA SILVA',
        'PAULO ROBERTO','GABRIELLE MARTINS','JENNIFER ARIELLE','SAMIRES BARROS',
        'LUCIANO','JENNIFER SILVEIRA','MAYCOW GABRIEL','LAURA SILVA'
    ]}

    try:
        # 1. df_eleg já vem processado com elegíveis — apenas pegar CPF e valor
        # Colunas esperadas: uc_cpf, valor, data_pagamento
        df_pagos = df_eleg[df_eleg['elegibilidade']=='Elegível'].copy() if 'elegibilidade' in df_eleg.columns else df_eleg.copy()
        if 'uc_cpf' not in df_pagos.columns:
            return None, "Coluna uc_cpf não encontrada no resultado processado."
        if 'valor' not in df_pagos.columns:
            return None, "Coluna valor não encontrada no resultado processado."

        # 2. Ler interações — pode ter abas CHAT, LIGACOES, DISPAROS
        arq_interacoes.seek(0)
        try:
            xls_int = pd.ExcelFile(arq_interacoes)
            abas_norm = {norm(a): a for a in xls_int.sheet_names}
        except:
            arq_interacoes.seek(0)
            xls_int = None
            abas_norm = {}

        contatos = []

        def tempo_para_segundos(t):
            try:
                import datetime as _dt
                import pandas as _pd
                # Se é Timedelta
                if isinstance(t, _pd.Timedelta):
                    return max(1, int(t.total_seconds()))
                # Se é time
                if isinstance(t, _dt.time):
                    return t.hour*3600 + t.minute*60 + t.second
                # Se é Timestamp ou datetime — pegar só a parte de hora
                if isinstance(t, (_dt.datetime, _pd.Timestamp)):
                    return max(1, t.hour*3600 + t.minute*60 + t.second)
                s = str(t).strip()
                # Formato "0 days HH:MM:SS"
                if 'days' in s:
                    s = s.split(' ')[-1]
                # Formato "1900-01-01 HH:MM:SS" (datetime do Excel)
                if len(s) > 8 and ' ' in s:
                    s = s.split(' ')[-1]
                if ':' in s:
                    partes = s.split(':')
                    if len(partes) == 3:
                        return max(1, int(partes[0])*3600 + int(partes[1])*60 + int(float(partes[2])))
                    elif len(partes) == 2:
                        return max(1, int(partes[0])*60 + int(partes[1]))
                f = float(s)
                if 0 < f < 1:
                    return max(1, int(round(f * 86400)))
                return max(1, int(f))
            except:
                return 1

        def extrair_cpf_agente_data(df_raw, meio, segundos_fixos=None):
            """Extrai CPF, agente, data e segundos de um dataframe."""
            cols_norm = {norm(str(c)): c for c in df_raw.columns}

            # CPF — primeira coluna que pareça CPF
            col_cpf = next((cols_norm[k] for k in cols_norm if any(x in k for x in ['CPF','IDENTIFICAD','CLIENTE'])), df_raw.columns[0])

            # Data — qualquer coluna com data
            col_data = next((cols_norm[k] for k in cols_norm if any(x in k for x in ['DATA','DT','DATE','DIA'])), None)
            if col_data is None:
                # Tentar segunda coluna como data
                if len(df_raw.columns) > 1:
                    col_data = df_raw.columns[1]
                else:
                    return pd.DataFrame()

            # Agente — opcional (disparo e chat não têm)
            col_agente = next((cols_norm[k] for k in cols_norm if any(x in k for x in ['AGENTE','ATENDENTE','OPERADOR','COLABORADOR','USUARIO','USER','AGENT'])), None)

            # Tempo — só ligações
            col_tempo = next((cols_norm[k] for k in cols_norm if any(x in k for x in ['TEMPO','DURACAO','DURATION','TIME','MINUTO','SEGUNDO'])), None)

            df_out = pd.DataFrame()
            df_out['uc_cpf'] = df_raw[col_cpf].apply(normalizar_cpf)
            df_out['data_contato'] = parse_data_inteligente(df_raw[col_data])
            df_out['meio'] = meio

            # Agente: disparo e chat = LUCIANO, ligação = nome do agente
            if col_agente is not None:
                df_out['agente'] = df_raw[col_agente].astype(str).str.strip().str.upper()
            else:
                df_out['agente'] = 'LUCIANO'

            # Segundos
            if segundos_fixos is not None:
                df_out['segundos'] = segundos_fixos
            elif col_tempo is not None:
                df_out['segundos'] = df_raw[col_tempo].apply(tempo_para_segundos)
            else:
                df_out['segundos'] = 1

            df_out = df_out.dropna(subset=['data_contato'])
            df_out = df_out[df_out['uc_cpf'].str.len() >= 8]
            df_out = df_out[df_out['uc_cpf'] != 'nan']
            return df_out

        def tem_tempo_real(df_raw, col_tempo):
            if col_tempo is None:
                return False
            vals = df_raw[col_tempo].dropna()
            return len(vals) > 0

        # Processar cada aba
        if xls_int:
            for aba_orig in xls_int.sheet_names:
                aba_n = norm(aba_orig)
                df_raw = pd.read_excel(xls_int, sheet_name=aba_orig)
                # Detectar coluna de tempo
                cols_norm_tmp = {norm(str(c)): c for c in df_raw.columns}
                col_tempo_tmp = next((cols_norm_tmp[k] for k in cols_norm_tmp if any(x in k for x in ['TEMPO','DURACAO','DURATION','TIME','MINUTO','SEGUNDO'])), None)
                tem_tempo = tem_tempo_real(df_raw, col_tempo_tmp)
                if 'DISPAR' in aba_n:
                    df_c = extrair_cpf_agente_data(df_raw, 'DISPARO', segundos_fixos=None if tem_tempo else 1)
                elif 'CHAT' in aba_n:
                    df_c = extrair_cpf_agente_data(df_raw, 'CHAT', segundos_fixos=None if tem_tempo else 5)
                elif 'LIG' in aba_n or 'LIGAC' in aba_n:
                    df_c = extrair_cpf_agente_data(df_raw, 'LIGACAO', segundos_fixos=None)
                else:
                    df_c = extrair_cpf_agente_data(df_raw, 'OUTRO', segundos_fixos=None if tem_tempo else 1)
                if not df_c.empty:
                    contatos.append(df_c)
        else:
            # Arquivo único — usar tempo real se disponível
            arq_interacoes.seek(0)
            df_raw = ler_arquivo(arq_interacoes)
            df_c = extrair_cpf_agente_data(df_raw, 'OUTRO', segundos_fixos=None)
            if not df_c.empty:
                contatos.append(df_c)

        if not contatos:
            # Debug: mostrar abas encontradas
            abas_encontradas = xls_int.sheet_names if xls_int else "arquivo não é Excel"
            return None, f"Nenhuma interação encontrada. Abas no arquivo: {abas_encontradas}. Verifique se as abas têm colunas de CPF e Agente."

        df_int = pd.concat(contatos, ignore_index=True)

        # 3. Elegíveis já vêm do df processado — apenas garantir que temos dados
        if df_pagos.empty:
            return None, "Nenhum registro elegível encontrado na base processada."

        # 4. Divisão proporcional — linhas individuais (sem groupby), igual ao Excel manual
        # Normalizar CPFs dos dois lados para garantir que o merge não perde registros
        df_pagos['uc_cpf'] = df_pagos['uc_cpf'].apply(normalizar_cpf)
        df_int['uc_cpf'] = df_int['uc_cpf'].apply(normalizar_cpf)

        # Converter data_pagamento para datetime
        df_pagos['data_pagamento'] = pd.to_datetime(df_pagos['data_pagamento'], errors='coerce').dt.normalize()
        df_pagos['valor'] = pd.to_numeric(df_pagos['valor'], errors='coerce').fillna(0)

        # PASSO 1: ID único por pagamento (essencial para boletos com mesmo CPF/data/valor)
        df_pagos = df_pagos.reset_index(drop=True)
        df_pagos['id_pagamento'] = df_pagos.index

        df_int_eleg = df_int[df_int['uc_cpf'].isin(df_pagos['uc_cpf'].unique())].copy()
        df_int_eleg['agente'] = df_int_eleg['agente'].fillna('DESCONHECIDO').replace('', 'DESCONHECIDO').replace('NAN', 'DESCONHECIDO')
        df_int_eleg['data_contato'] = pd.to_datetime(df_int_eleg['data_contato'], errors='coerce').dt.normalize()
        df_int_eleg['segundos'] = pd.to_numeric(df_int_eleg['segundos'], errors='coerce').fillna(0)
        # Remover duplicatas — mesmo CPF + agente + data_contato + segundos
        df_int_eleg = df_int_eleg.drop_duplicates(subset=['uc_cpf','agente','data_contato','segundos']).copy()

        # PASSO 2: Mesclar por CPF — cada par (pagamento x interacao) vira uma linha
        df_cross = df_pagos[['id_pagamento','uc_cpf','valor','data_pagamento']].merge(
            df_int_eleg, on='uc_cpf', how='inner'
        )

        # PASSO 3: Filtro — manter so contatos no dia ou antes do pagamento
        df_cross = df_cross[df_cross['data_contato'] <= df_cross['data_pagamento']].copy()

        # PASSO 5: Seg Totais por ID Pagamento (nao por CPF+Data!)
        seg_tot = df_cross.groupby('id_pagamento')['segundos'].sum().reset_index()
        seg_tot.columns = ['id_pagamento','total_seg']
        df_cross = df_cross.merge(seg_tot, on='id_pagamento', how='left')
        df_cross = df_cross[df_cross['total_seg'] > 0].copy()

        # PASSO 6: Valor Proporcional = valor_boleto * (seg_cada / seg_totais)
        df_cross['valor_proporcional'] = df_cross['valor'] * (df_cross['segundos'] / df_cross['total_seg'])
        df_seg = df_cross

                # 5. Separar Luciano vs Amitycall
        def match_dois_nomes(agente, lista):
            a = str(agente).upper().strip()
            palavras_a = a.split()
            dois_a = ' '.join(palavras_a[:2]) if len(palavras_a) >= 2 else a
            for nome in lista:
                palavras_n = nome.upper().strip().split()
                dois_n = ' '.join(palavras_n[:2]) if len(palavras_n) >= 2 else nome.upper()
                if dois_a == dois_n:
                    return True
            return False

        df_seg['equipe'] = df_seg['agente'].apply(
            lambda a: 'luciano' if match_dois_nomes(a, AGENTES_LUCIANO) else 'metcool'
        )

        # Debug — agentes únicos e classificação
        agentes_unicos = df_seg[['agente','equipe']].drop_duplicates().to_dict('records')

        total_luciano = float(df_seg[df_seg['equipe']=='luciano']['valor_proporcional'].sum())
        total_amitycall = float(df_seg[df_seg['equipe']=='metcool']['valor_proporcional'].sum())
        total_geral = total_luciano + total_amitycall
        n_eleg = len(df_pagos)
        n_boletos = len(df_eleg)

        # Breakdown por UF para Luciano e Meet Call
        por_uf_luciano = {}
        por_uf_metcool = {}
        if 'uf' in df_pagos.columns:
            # Merge df_seg com df_pagos para pegar UF
            df_seg_uf = df_seg.merge(df_pagos[['uc_cpf','uf','valor']].drop_duplicates('uc_cpf'), on='uc_cpf', how='left', suffixes=('','_pag'))
            for equipe, por_uf in [('luciano', por_uf_luciano), ('metcool', por_uf_metcool)]:
                df_eq = df_seg_uf[df_seg_uf['equipe']==equipe]
                for uf, grp in df_eq.groupby('uf'):
                    por_uf[str(uf)] = {
                        'valor': float(grp['valor_proporcional'].sum()),
                        'boletos': int(grp['uc_cpf'].nunique())
                    }

        # Detalhe — usa df_seg (linhas individuais) expandido por boleto
        # df_seg já tem valor_proporcional por linha de contato
        # Para cada linha de contato x cada boleto do CPF:
        # val_prop_linha = valor_proporcional_linha * (valor_boleto / valor_total_cpf)

        def seg_to_hms_div(s):
            try:
                s=int(s); return f"{s//3600:02d}:{(s%3600)//60:02d}:{s%60:02d}"
            except: return "00:00:00"

        _ags_luc_det = {n.upper().strip() for n in ['JHENIFFER SANTOS','MARCOS MARTINS','JUNIOR OTAIDES','CAMILA NARA','HEVERTON TAVARES','MARIA CLARA','LORENZZO PEREIRA','GRASIELLE DA SILVA SANTOS','DIOGO OLIVEIRA','MICHELLE BATISTA','KETLE SILVA','EMANUEL FERREIRA','EDUARDA SANQUETA','GABRIELLE MARTINS','VICTORIA SILVA','CAUA ALVES','PAULO ROBERTO','SAMIRES BARROS','JENNIFER ARIELLE','LUCIANO','MAYCOW GABRIEL','LAURA SILVA']}

        # Detalhe — usar df_seg diretamente (já tem valor_proporcional por boleto)
        df_seg_det = df_seg.copy()
        df_seg_det['Empresa'] = df_seg_det['agente'].apply(
            lambda a: 'iGreen' if str(a).upper().strip() in _ags_luc_det else 'Meet Call'
        )
        df_seg_det['Tempo'] = df_seg_det['segundos'].apply(seg_to_hms_div)
        # Adicionar fornecedora e UF do df_pagos
        cols_extra = ['uc_cpf','data_pagamento']
        if 'fornecedora' in df_pagos.columns: cols_extra.append('fornecedora')
        if 'uf' in df_pagos.columns: cols_extra.append('uf')
        df_seg_det = df_seg_det.merge(
            df_pagos[cols_extra].drop_duplicates(['uc_cpf','data_pagamento']),
            on=['uc_cpf','data_pagamento'], how='left'
        )
        rename_map = {
            'uc_cpf':'CPF','agente':'Agente','data_contato':'Data Contato',
            'data_pagamento':'Data Pagamento','valor':'Valor Boleto',
            'segundos':'Segundos Cada','total_seg':'Segundos Totais',
            'valor_proporcional':'Valor Proporcional'
        }
        if 'fornecedora' in df_seg_det.columns: rename_map['fornecedora'] = 'Fornecedora'
        if 'uf' in df_seg_det.columns: rename_map['uf'] = 'UF'
        df_seg_det = df_seg_det.rename(columns=rename_map)
        cols_det = ['CPF','Agente']
        if 'Fornecedora' in df_seg_det.columns: cols_det.append('Fornecedora')
        if 'UF' in df_seg_det.columns: cols_det.append('UF')
        cols_det += ['Empresa','Data Contato','Data Pagamento','Valor Boleto','Tempo','Segundos Cada','Segundos Totais','Valor Proporcional']
        df_seg_det = df_seg_det[[c for c in cols_det if c in df_seg_det.columns]]

        resultado = {
            'total_luciano': total_luciano,
            'total_metcool': total_amitycall,
            'total_geral': total_geral,
            'n_elegivel': n_eleg,
            'n_boletos': n_boletos,
            'df_agentes': df_seg,
            'df_agentes_det': df_seg_det,
            'agentes_debug': agentes_unicos,
            'df_eleg': df_eleg,
            'por_uf_luciano': por_uf_luciano,
            'por_uf_metcool': por_uf_metcool,
            'ma': ma
        }
        return resultado, None

    except Exception as e:
        import traceback
        return None, f"Erro: {e}\n{traceback.format_exc()}"


def calcular_resultado_atendentes(arq_pagos, arq_interacoes, eq, ma):
    """Calcula resultado por atendente — usa o mesmo fluxo do processamento normal."""
    import unicodedata
    def norm(s): return unicodedata.normalize('NFKD',str(s).upper().strip()).encode('ascii','ignore').decode()

    try:
        # 1. Processar base igual ao fluxo normal (pagos + elegibilidade)
        arq_pagos.seek(0)
        df_res, erros, abas = processar_base_unica(arq_pagos, eq, ma)

        if df_res is None or df_res.empty:
            return None, f"Erro no processamento: {erros}"

        # 2. Cruzar com interações para elegibilidade — igual ao fluxo normal
        arq_interacoes.seek(0)
        try:
            xls_int = pd.ExcelFile(arq_interacoes)
            contatos_tmp = []
            for aba in xls_int.sheet_names:
                aba_n = norm(aba)
                if any(x in aba_n for x in ['PAGO','PAGAM','RECEB','BASE','BAIXA']):
                    continue
                df_aba = pd.read_excel(xls_int, sheet_name=aba)
                dd = processar_contatos(df_aba)
                if not dd.empty:
                    contatos_tmp.append(dd)
            if contatos_tmp:
                pc = pd.concat(contatos_tmp, ignore_index=True).groupby('uc_cpf', as_index=False)['data_contato'].min()
                df_res['primeiro_contato'] = df_res['uc_cpf'].map(dict(zip(pc['uc_cpf'], pc['data_contato'])))
                df_res['primeiro_contato'] = pd.to_datetime(df_res['primeiro_contato'], errors='coerce').dt.normalize()
                df_res['data_pagamento'] = pd.to_datetime(df_res['data_pagamento'], errors='coerce').dt.normalize()
                df_res['diferenca_dias'] = (df_res['data_pagamento'] - df_res['primeiro_contato']).dt.days
                def classif(row):
                    if pd.isna(row.get('primeiro_contato')): return 'ND'
                    d = row.get('diferenca_dias')
                    if pd.isna(d): return 'ND'
                    return 'Elegível' if int(d) >= 0 else 'Não Elegível'
                df_res['elegibilidade'] = df_res.apply(classif, axis=1)
        except Exception as e:
            return None, f"Erro ao processar interações: {e}"

        # 3. Filtrar elegíveis
        df_eleg = df_res[df_res['elegibilidade'] == 'Elegível'].copy()
        if df_eleg.empty:
            return None, "Nenhum registro elegível encontrado."

        # 4. Ler interações com agente para divisão proporcional
        arq_interacoes.seek(0)
        contatos_ag = []
        try:
            arq_interacoes.seek(0)
            xls_int2 = pd.ExcelFile(arq_interacoes)
            for aba in xls_int2.sheet_names:
                aba_n = norm(aba)
                if any(x in aba_n for x in ['PAGO','PAGAM','RECEB','BASE','BAIXA']):
                    continue
                df_aba = pd.read_excel(xls_int2, sheet_name=aba)
                cols_n = [norm(str(c)) for c in df_aba.columns]
                tem_agente = any(any(x in c for x in ['AGENTE','ATENDENTE','OPERADOR','COLABORADOR','USER']) for c in cols_n)
                if not tem_agente:
                    continue
                if 'DISPAR' in aba_n:
                    dd = processar_contatos_com_agente(df_aba, segundos_fixos=1)
                elif 'CHAT' in aba_n:
                    dd = processar_contatos_com_agente(df_aba, segundos_fixos=5)
                else:
                    dd = processar_contatos_com_agente(df_aba)
                if not dd.empty:
                    contatos_ag.append(dd)
        except Exception as e:
            return None, f"Erro ao ler interações com agente: {e}"

        if not contatos_ag:
            return None, "Nenhuma interação com agente encontrada."

        df_int = pd.concat(contatos_ag, ignore_index=True)

        # 5. Divisão proporcional com ID único por pagamento (igual ao Excel manual)
        df_eleg_at = df_eleg.copy()
        df_eleg_at['valor'] = pd.to_numeric(df_eleg_at['valor'], errors='coerce').fillna(0)
        df_eleg_at['data_pagamento'] = pd.to_datetime(df_eleg_at['data_pagamento'], errors='coerce').dt.normalize()
        df_eleg_at = df_eleg_at.reset_index(drop=True)
        df_eleg_at['id_pagamento'] = df_eleg_at.index

        df_int_todos = df_int[df_int['uc_cpf'].isin(df_eleg_at['uc_cpf'].unique())].copy()
        df_int_todos['data_contato'] = pd.to_datetime(df_int_todos['data_contato'], errors='coerce').dt.normalize()
        df_int_todos['segundos'] = pd.to_numeric(df_int_todos['segundos'], errors='coerce').fillna(0)

        # Mesclar por CPF — cada par (pagamento x interacao) vira uma linha
        df_cross_at = df_eleg_at[['id_pagamento','uc_cpf','valor','data_pagamento']].merge(
            df_int_todos, on='uc_cpf', how='inner'
        )
        # Filtro: manter so contatos no dia ou antes do pagamento
        df_cross_at = df_cross_at[df_cross_at['data_contato'] <= df_cross_at['data_pagamento']].copy()

        # Seg Totais por ID Pagamento (nao por CPF+Data!)
        seg_tot_at = df_cross_at.groupby('id_pagamento')['segundos'].sum().reset_index()
        seg_tot_at.columns = ['id_pagamento','total_seg']
        df_cross_at = df_cross_at.merge(seg_tot_at, on='id_pagamento', how='left')
        df_cross_at = df_cross_at[df_cross_at['total_seg'] > 0].copy()
        df_cross_at['valor_proporcional'] = df_cross_at['valor'] * (df_cross_at['segundos'] / df_cross_at['total_seg'])

        if df_cross_at.empty:
            return None, "Nenhum resultado calculado."

        # 6. Somar por agente
        df_result = df_cross_at.groupby('agente')['valor_proporcional'].sum().reset_index()
        df_result = df_result.sort_values('valor_proporcional', ascending=False)
        df_result.columns = ['agente', 'valor']

        total = float(df_result['valor'].sum())

        def seg_to_hms(s):
            try:
                s=int(s); return f"{s//3600:02d}:{(s%3600)//60:02d}:{s%60:02d}"
            except: return "00:00:00"

        # Detalhe — usa df_cross_at que já tem id_pagamento e valor_proporcional correto
        _ags_luc_full = {n.upper().strip() for n in ['JHENIFFER SANTOS','MARCOS MARTINS','JUNIOR OTAIDES','CAMILA NARA','HEVERTON TAVARES','MARIA CLARA','LORENZZO PEREIRA','GRASIELLE DA SILVA SANTOS','DIOGO OLIVEIRA','MICHELLE BATISTA','KETLE SILVA','EMANUEL FERREIRA','EDUARDA SANQUETA','GABRIELLE MARTINS','VICTORIA SILVA','CAUA ALVES','PAULO ROBERTO','SAMIRES BARROS','JENNIFER ARIELLE','LUCIANO','MAYCOW GABRIEL','LAURA SILVA']}
        df_det = df_cross_at.copy()
        df_det['Tempo'] = df_det['segundos'].apply(seg_to_hms)
        df_det['Empresa'] = df_det['agente'].str.upper().str.strip().apply(
            lambda a: 'iGreen' if a in _ags_luc_full else 'Meet Call'
        )
        # Adicionar fornecedora e UF do df_eleg_at via id_pagamento
        cols_extra = ['id_pagamento']
        if 'fornecedora' in df_eleg_at.columns: cols_extra.append('fornecedora')
        if 'uf' in df_eleg_at.columns: cols_extra.append('uf')
        df_det = df_det.merge(df_eleg_at[cols_extra], on='id_pagamento', how='left')
        rename_map = {
            'uc_cpf':'CPF','agente':'Agente','data_contato':'Data Contato',
            'data_pagamento':'Data Pagamento','valor':'Valor Boleto',
            'segundos':'Segundos Cada','total_seg':'Segundos Totais',
            'valor_proporcional':'Valor Proporcional','id_pagamento':'ID Pagamento'
        }
        if 'fornecedora' in df_det.columns: rename_map['fornecedora'] = 'Fornecedora'
        if 'uf' in df_det.columns: rename_map['uf'] = 'UF'
        df_det = df_det.rename(columns=rename_map)
        cols_det = ['ID Pagamento','CPF','Agente','Data Contato','Data Pagamento']
        if 'Fornecedora' in df_det.columns: cols_det.append('Fornecedora')
        if 'UF' in df_det.columns: cols_det.append('UF')
        cols_det += ['Empresa','Valor Boleto','Tempo','Segundos Cada','Segundos Totais','Valor Proporcional']
        df_detalhe = df_det[[c for c in cols_det if c in df_det.columns]]

        # Aba 2 — Pagos elegíveis
        df_eleg_out = df_eleg.copy()
        if 'primeiro_contato' not in df_eleg_out.columns:
            pc_map = df_int.groupby('uc_cpf')['data_contato'].min()
            df_eleg_out['primeiro_contato'] = df_eleg_out['uc_cpf'].map(pc_map)
        df_eleg_out['diferenca_dias'] = (
            pd.to_datetime(df_eleg_out['data_pagamento'], errors='coerce').dt.normalize() -
            pd.to_datetime(df_eleg_out['primeiro_contato'], errors='coerce').dt.normalize()
        ).dt.days
        df_eleg_out = df_eleg_out.rename(columns={'uc_cpf':'CPF','valor':'Valor','data_pagamento':'Data Pagamento','primeiro_contato':'Primeiro Contato','diferenca_dias':'Diferença Dias'})

        return {
            'df_result': df_result,
            'df_detalhe': df_detalhe,
            'df_eleg_out': df_eleg_out,
            'total': total,
            'n_elegivel': len(df_eleg),
            'n_boletos': len(df_res),
            'eq': eq,
            'ma': ma
        }, None

    except Exception as e:
        import traceback
        return None, f"Erro: {e}\n{traceback.format_exc()}"



def processar_contatos_com_agente(df_raw, segundos_fixos=None):
    """Extrai CPF, data, agente e segundos — igual ao processar_contatos mas com agente."""
    import unicodedata, datetime
    def norm(s): return unicodedata.normalize('NFKD',str(s).upper().strip()).encode('ascii','ignore').decode()

    df_raw = df_raw.reset_index(drop=True)
    cols_norm = {norm(str(c)): c for c in df_raw.columns}

    # CPF
    col_cpf = next((cols_norm[k] for k in cols_norm if k in ['CPF','IDENTIFICADOR','IDENTIF','IDENTIFICACAO']), df_raw.columns[0])
    # Data
    col_data = next((cols_norm[k] for k in cols_norm if any(x in k for x in ['DATA','DT_','BAIXA','CONTATO','INTERAC','LIGAC','CHAT','DISPAR','PAGAM','DIA'])), df_raw.columns[1] if len(df_raw.columns)>1 else df_raw.columns[0])
    # Agente
    col_agente = next((cols_norm[k] for k in cols_norm if any(x in k for x in ['AGENTE','ATENDENTE','OPERADOR','COLABORADOR','USUARIO','USER','AGENT'])), None)
    # Tempo
    col_tempo = next((cols_norm[k] for k in cols_norm if any(x in k for x in ['TEMPO','DURACAO','DURATION','TIME','MINUTO','SEGUNDO'])), None)

    def t2s(t):
        try:
            import datetime as _dt
            if isinstance(t, _dt.time): return t.hour*3600+t.minute*60+t.second
            if isinstance(t, _dt.datetime): return t.hour*3600+t.minute*60+t.second
            s=str(t).strip()
            if ':' in s:
                p=s.split(':')
                if len(p)==3: return int(p[0])*3600+int(p[1])*60+int(p[2])
                if len(p)==2: return int(p[0])*60+int(p[1])
            f=float(s)
            # Sempre pegar só a parte decimal (fração do dia = hora)
            f = f - int(f)
            return max(1,int(round(f*86400)))
        except: return 1

    dd = pd.DataFrame()
    dd['uc_cpf'] = df_raw[col_cpf].apply(normalizar_cpf)
    dd['data_contato'] = parse_data_inteligente(df_raw[col_data])
    dd['agente'] = df_raw[col_agente].astype(str).str.strip().str.upper() if col_agente else 'DESCONHECIDO'

    if segundos_fixos is not None:
        dd['segundos'] = segundos_fixos
    elif col_tempo:
        dd['segundos'] = df_raw[col_tempo].apply(t2s)
    else:
        dd['segundos'] = 1

    dd = dd.dropna(subset=['data_contato'])
    dd = dd[dd['uc_cpf'].str.len()>=8]
    dd = dd[dd['uc_cpf']!='nan']
    return dd

def pagina_upload(ma):
    u=st.session_state.usuario
    header_page('Upload de Bases Mensais','Processamento automatico')
    eq=seletor_equipe(u['equipe'] or 'tamires')
    col_up,col_hist=st.columns([1,1])

    with col_up:
        st.markdown('<p style="color:#3a6a4a;font-size:10px;text-transform:uppercase;letter-spacing:1.5px;margin-bottom:8px">PROCESSAR BASE</p>',unsafe_allow_html=True)
        st.markdown('<div style="background:#0d1a0d;border:1px solid #1e3a1e;border-radius:10px;padding:12px 16px;margin-bottom:12px;font-size:12px;color:#5a9a70;line-height:1.8">'
            'Aceita <strong style="color:#e8f5e9">.xlsx</strong> ou <strong style="color:#e8f5e9">.csv</strong><br>'
            '<strong>Opção 1:</strong> Excel único com abas PAGOS | CHAT | LIGACOES | DISPAROS<br>'
            '<strong>Opção 2:</strong> CSV/Excel de Pagos + arquivo separado de Interações'
            '</div>',unsafe_allow_html=True)

        modo_upload=st.radio("Modo de upload:",["📄 Arquivo único","📂 Dois arquivos (Pagos + Interações)"],horizontal=True,key="modo_upload")
        
        if modo_upload=="📄 Arquivo único":
            arq=st.file_uploader('Base (.xlsx ou .csv)',type=['xlsx','csv'],label_visibility='collapsed',key='base_unica')
            arq_interacoes=None
        else:
            col_a,col_b=st.columns(2)
            with col_a:
                st.markdown("<p style='color:#5a9a70;font-size:11px;margin-bottom:4px'>BASE DE PAGOS (.xlsx ou .csv)</p>",unsafe_allow_html=True)
                arq=st.file_uploader('Pagos',type=['xlsx','csv'],label_visibility='collapsed',key='base_pagos')
            with col_b:
                st.markdown("<p style='color:#5a9a70;font-size:11px;margin-bottom:4px'>BASE DE INTERAÇÕES (.xlsx ou .csv)</p>",unsafe_allow_html=True)
                arq_interacoes=st.file_uploader('Interações',type=['xlsx','csv'],label_visibility='collapsed',key='base_interacoes')
        if arq:
            try:
                xls=pd.ExcelFile(arq)
                ah=' | '.join(xls.sheet_names)
                arq.seek(0)
                st.markdown(f'<div style="background:#0a1a0a;border:1px solid #1e3a1e;border-radius:6px;padding:8px 12px;margin:8px 0;color:#5a9a70;font-size:12px"><strong style="color:#e8f5e9">{arq.name}</strong><br>Abas: {ah}</div>',unsafe_allow_html=True)
            except: pass
        # Seletor tipo de processamento
        tipo_proc = st.radio("Tipo de processamento:",
            ["Recebido Geral", "Resultado por Atendente"],
            key=f"tipo_proc_{eq}_{ma}", horizontal=True)

        # Dias trabalhados — só para Resultado por Atendente
        if tipo_proc == "Resultado por Atendente":
            col_dt1, col_dt2 = st.columns(2)
            with col_dt1:
                dias_trab_at = st.number_input("Dias Trabalhados", min_value=0, max_value=31, value=0, key=f"dt_at_{eq}_{ma}")
            with col_dt2:
                total_dias_at = st.number_input("Total Dias do Mês", min_value=0, max_value=31, value=21, key=f"td_at_{eq}_{ma}")

        if st.button('PROCESSAR',use_container_width=True):
            if not arq: st.error('Selecione a base de pagos antes de processar!'); return

            # Resultado por Atendente
            if tipo_proc == "Resultado por Atendente":
                if dias_trab_at == 0:
                    st.error("⚠️ Preencha os Dias Trabalhados antes de processar.")
                    st.stop()
                if total_dias_at == 0:
                    st.error("⚠️ Preencha o Total Dias do Mês antes de processar.")
                    st.stop()
                arq.seek(0)
                if arq_interacoes is not None:
                    arq_interacoes.seek(0)
                    _arq_int_at = arq_interacoes
                else:
                    arq.seek(0)
                    _arq_int_at = arq
                with st.spinner('Calculando resultado por atendente...'):
                    res_at, erro_at = calcular_resultado_atendentes(arq, _arq_int_at, eq, ma)
                if erro_at:
                    st.error(erro_at)
                else:
                    res_at['dias_trab'] = dias_trab_at
                    res_at['total_dias'] = total_dias_at
                    st.session_state['resultado_atendentes'] = res_at
                    st.session_state['div_prop_resultado'] = None
                    st.rerun()
                st.stop()



            arq.seek(0)
            with st.spinner('Processando...'):
                df_res,erros,abas=processar_base_unica(arq,eq,ma)
                # Se tiver arquivo de interações separado, processar e cruzar
                if arq_interacoes is not None and df_res is not None:
                    try:
                        arq_interacoes.seek(0)
                        # Ler todas as abas do Excel de interações
                        try:
                            xls_tmp=pd.ExcelFile(arq_interacoes)
                            dfs_tmp=[]
                            for aba in xls_tmp.sheet_names:
                                df_aba=pd.read_excel(xls_tmp,sheet_name=aba,dtype={"cpf":str,"CPF":str})
                                dd_aba=processar_contatos(df_aba)
                                if not dd_aba.empty:
                                    dfs_tmp.append(dd_aba)
                            dd_int=pd.concat(dfs_tmp,ignore_index=True) if dfs_tmp else pd.DataFrame()
                        except Exception as e_int:
                            st.warning(f"Erro ao ler interações: {e_int}")
                            arq_interacoes.seek(0)
                            df_int=ler_arquivo(arq_interacoes)
                            dd_int=processar_contatos(df_int)
                        if not dd_int.empty:
                            # Garantir CPF normalizado nos dois lados
                            dd_int["uc_cpf"]=dd_int["uc_cpf"].apply(normalizar_cpf)
                            df_res["uc_cpf"]=df_res["uc_cpf"].apply(normalizar_cpf)
                            # Normalizar datas (remover hora)
                            dd_int["data_contato"]=pd.to_datetime(dd_int["data_contato"],errors="coerce").dt.normalize()
                            pc=dd_int.groupby("uc_cpf",as_index=False)["data_contato"].min()

                            df_res["primeiro_contato"]=pd.to_datetime(
                                df_res["uc_cpf"].map(dict(zip(pc["uc_cpf"],pc["data_contato"]))),
                                errors="coerce").dt.normalize()
                            df_res["data_pagamento"]=pd.to_datetime(df_res["data_pagamento"],errors="coerce").dt.normalize()
                            df_res["diferenca_dias"]=(df_res["data_pagamento"]-df_res["primeiro_contato"]).dt.days
                            def classif2(row):
                                if pd.isna(row.get("primeiro_contato")): return "ND"
                                d=row.get("diferenca_dias")
                                if pd.isna(d): return "ND"
                                return "Elegível" if int(d)>=0 else "Não Elegível"
                            df_res["elegibilidade"]=df_res.apply(classif2,axis=1)
                            abas=["Interações"]
                            # Formatar datas
                            for col in ["data_pagamento","primeiro_contato"]:
                                if col in df_res.columns:
                                    try: df_res[col]=pd.to_datetime(df_res[col],errors="coerce").dt.strftime("%Y-%m-%d").where(pd.to_datetime(df_res[col],errors="coerce").notna(),other=None)
                                    except: pass
                    except Exception as e:
                        st.warning(f"Erro ao processar interações: {e}")
            for e in erros: st.error(e)
            if df_res is not None and not df_res.empty:
                # Luciano com interações: calcular divisão proporcional
                if eq=='luciano' and arq_interacoes is not None:
                    arq_interacoes.seek(0)
                    with st.spinner('Calculando divisão proporcional...'):
                        res_div, erro_div = calcular_divisao_proporcional_luciano(df_res, arq_interacoes, ma)
                    if erro_div:
                        st.error(erro_div)
                    else:
                        st.session_state['div_prop_resultado'] = res_div
                        st.rerun()
                    st.stop()
                st.session_state['df_proc_temp']=df_res
                st.session_state['proc_eq']=eq
                st.session_state['proc_ma']=ma
                st.session_state['proc_abas']=abas
                st.rerun()
        if (st.session_state.get('df_proc_temp') is not None and
            st.session_state.get('proc_eq')==eq and
            st.session_state.get('proc_ma')==ma):
            df_res=st.session_state['df_proc_temp']
            abas=st.session_state.get('proc_abas',[])
            elig=df_res[df_res['elegibilidade']=='Elegível'] if 'elegibilidade' in df_res.columns else df_res
            ve=elig['valor'].sum() if 'valor' in elig.columns else 0
            ce=elig['uc_cpf'].nunique() if 'uc_cpf' in elig.columns else 0
            if abas: st.info(f"Abas: {', '.join(abas)}")
            st.success(f"{len(df_res):,} registros processados!")
            c1,c2,c3=st.columns(3)
            c1.metric('Valor Recebido',fmt_brl(ve))
            c2.metric('Boletos',f'{len(elig):,}')
            c3.metric('Clientes',f'{ce:,}')
            st.markdown('---')
            col1,col2=st.columns(2)
            with col1:
                # Verificar se já existe processamento salvo
                proc_existente = buscar_ultimo_processamento(ma, eq)
                label_btn = '🔄 Substituir Resultado' if proc_existente else 'Salvar Resultado'
                if st.button(label_btn, use_container_width=True, key='btn_salvar_proc',
                             disabled=st.session_state.get('salvando_proc', False)):
                    st.session_state['salvando_proc'] = True
                    salvar_processamento(ma,eq,df_res,st.session_state.usuario.get('nome',''))
                    st.session_state['df_proc_temp']=None
                    st.session_state['salvando_proc'] = False
                    st.success('✅ Resultado salvo com sucesso!')
                    st.rerun()
            with col2:
                if st.button('Descartar',use_container_width=True,key='btn_desc'):
                    st.session_state['df_proc_temp']=None; st.rerun()
            cols_show=[c for c in ['uc_cpf','data_pagamento','valor','fornecedora','elegibilidade','aging'] if c in df_res.columns]
            st.dataframe(df_res[cols_show].head(30) if cols_show else df_res.head(30),use_container_width=True)

            # Botão para baixar apenas Elegíveis em CSV — todas as colunas originais
            st.markdown("---")
            elig_download=df_res[df_res['elegibilidade']=='Elegível'] if 'elegibilidade' in df_res.columns else df_res


    # Limpar resultados de outras equipes se equipe mudou
    if st.session_state.get('div_prop_resultado') and eq != 'luciano':
        st.session_state['div_prop_resultado'] = None
    if st.session_state.get('resultado_atendentes') and st.session_state['resultado_atendentes'].get('eq') != eq:
        st.session_state['resultado_atendentes'] = None

    # Mostrar resultado por atendente (aguarda confirmação)
    if 'resultado_atendentes' in st.session_state and st.session_state['resultado_atendentes']:
        res_at = st.session_state['resultado_atendentes']
        st.markdown("---")
        st.markdown("### 👥 Resultado por Atendente")
        st.markdown(f"**Mês:** {res_at['ma']} | **Boletos:** {res_at['n_boletos']} | **Elegíveis:** {res_at['n_elegivel']} | **Total:** {fmt_brl(res_at['total'])}")

        # Para Luciano — mostrar separado por equipe
        if res_at.get('eq') == 'luciano':
            _ags_luc_full = {n.upper().strip() for n in ['JHENIFFER SANTOS','MARCOS MARTINS','JUNIOR OTAIDES','CAMILA NARA','HEVERTON TAVARES','MARIA CLARA','LORENZZO PEREIRA','GRASIELLE DA SILVA SANTOS','DIOGO OLIVEIRA','MICHELLE BATISTA','KETLE SILVA','EMANUEL FERREIRA','EDUARDA SANQUETA','GABRIELLE MARTINS','VICTORIA SILVA','CAUA ALVES','PAULO ROBERTO','SAMIRES BARROS','JENNIFER ARIELLE','LUCIANO','MAYCOW GABRIEL','LAURA SILVA']}
            df_result = res_at['df_result'].copy()
            df_luc = df_result[df_result['agente'].str.upper().str.strip().isin(_ags_luc_full)].copy()
            df_mc = df_result[~df_result['agente'].str.upper().str.strip().isin(_ags_luc_full)].copy()
            col_l, col_m = st.columns(2)
            with col_l:
                tc_luc_show = df_luc['valor'].sum()
                st.markdown(f"**🟢 Equipe Luciano** — Com Interação: {fmt_brl(tc_luc_show)}")
                df_luc['valor'] = df_luc['valor'].apply(fmt_brl)
                df_luc.columns = ['Atendente', 'Valor']
                df_luc = df_luc.reset_index(drop=True); df_luc.index += 1
                st.dataframe(df_luc, use_container_width=True)
            with col_m:
                tc_mc_show = df_mc['valor'].sum()
                st.markdown(f"**🔵 Meet Call** — Com Interação: {fmt_brl(tc_mc_show)}")
                df_mc['valor'] = df_mc['valor'].apply(fmt_brl)
                df_mc.columns = ['Atendente', 'Valor']
                df_mc = df_mc.reset_index(drop=True); df_mc.index += 1
                st.dataframe(df_mc, use_container_width=True)
        else:
            df_show = res_at['df_result'].copy()
            df_show['valor'] = df_show['valor'].apply(fmt_brl)
            df_show.columns = ['Atendente', 'Valor Proporcional']
            df_show = df_show.reset_index(drop=True); df_show.index += 1
            st.dataframe(df_show, use_container_width=True)
        # Download detalhe — duas abas
        if res_at.get('df_detalhe') is not None:
            import io
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                res_at['df_detalhe'].to_excel(writer, index=False, sheet_name='Todos os Contatos')
                if res_at.get('df_eleg_out') is not None:
                    res_at['df_eleg_out'].to_excel(writer, index=False, sheet_name='Pagos Elegíveis')
            buf.seek(0)
            st.download_button("⬇️ Baixar Detalhamento", buf.getvalue(),
                file_name=f"detalhe_atendentes_{res_at['ma']}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_detalhe_at")

        st.markdown("**⚠️ Confirme os valores antes de salvar:**")
        col_ok2, col_cancel2 = st.columns(2)
        with col_ok2:
            if st.button("✅ Confirmar e Salvar", use_container_width=True, key="btn_confirmar_at"):
                from datetime import datetime as _dt
                _ts = _dt.now().strftime("%Y%m%d%H%M%S%f")
                eq_at = res_at['eq']
                ma_at = res_at['ma']
                from collections import Counter
                from datetime import datetime as _dt2

                # Palavras ignoradas no match
                _STOP = {'DE','DA','DO','DAS','DOS','DI','DU'}

                def _palavras_sig(nome):
                    """Retorna palavras significativas (sem DE/DA/DOS etc)"""
                    return [p for p in nome.strip().upper().split() if p not in _STOP]

                def match_operador(nome_base, ops):
                    from collections import Counter as _Counter
                    # Usar primeiro nome para detectar duplicados
                    primeiros = [_palavras_sig(op['nome'])[0] if _palavras_sig(op['nome']) else '' for op in ops]
                    duplicados = {n for n, c in _Counter(primeiros).items() if c > 1}
                    
                    palavras_base = _palavras_sig(nome_base)
                    primeiro_base = palavras_base[0] if palavras_base else ""
                    
                    for op in ops:
                        palavras_op = _palavras_sig(op['nome'])
                        primeiro_op = palavras_op[0] if palavras_op else ""
                        
                        if primeiro_op in duplicados:
                            # Duplicado — bate as duas primeiras palavras significativas
                            segundo_base = palavras_base[1] if len(palavras_base) > 1 else ""
                            segundo_op = palavras_op[1] if len(palavras_op) > 1 else ""
                            if primeiro_base == primeiro_op and segundo_base == segundo_op:
                                return op['_id'], op['nome']
                        else:
                            # Único — bate só primeiro nome significativo
                            if primeiro_base == primeiro_op:
                                return op['_id'], op['nome']
                    return None, nome_base

                # Para Luciano — separar agentes entre Luciano e Meet Call
                if eq_at == 'luciano':
                    ops_luc = buscar_operadores('luciano')
                    ops_mc = buscar_operadores('metcool')
                    agentes_luc = {}
                    agentes_mc = {}
                    tc_luc = 0.0
                    tc_mc = 0.0
                    for _, row in res_at['df_result'].iterrows():
                        nome = str(row['agente']).strip()
                        valor = float(row['valor'])
                        nome_upper = nome.upper().strip()
                        # Verificar se é agente do Luciano pela lista
                        _ags_luc_save = {n.upper().strip() for n in ['JHENIFFER SANTOS','MARCOS MARTINS','JUNIOR OTAIDES','CAMILA NARA','HEVERTON TAVARES','MARIA CLARA','LORENZZO PEREIRA','GRASIELLE DA SILVA SANTOS','DIOGO OLIVEIRA','MICHELLE BATISTA','KETLE SILVA','EMANUEL FERREIRA','EDUARDA SANQUETA','GABRIELLE MARTINS','VICTORIA SILVA','CAUA ALVES','PAULO ROBERTO','SAMIRES BARROS','JENNIFER ARIELLE','LUCIANO','MAYCOW GABRIEL','LAURA SILVA']}
                        is_luciano = nome_upper in _ags_luc_save
                        if is_luciano:
                            op_id, op_nome = match_operador(nome, ops_luc)
                            if op_id is None:
                                op_id = f"auto-{nome.lower().replace(' ','-')}"
                                op_nome = nome
                            if op_id in agentes_luc:
                                agentes_luc[op_id]["valorRecebido"] += valor
                            else:
                                agentes_luc[op_id] = {"valorRecebido": valor, "nome": op_nome}
                            tc_luc += valor
                        else:
                            op_id, op_nome = match_operador(nome, ops_mc)
                            if op_id is None:
                                op_id = f"auto-{nome.lower().replace(' ','-')}"
                                op_nome = nome
                            if op_id in agentes_mc:
                                agentes_mc[op_id]["valorRecebido"] += valor
                            else:
                                agentes_mc[op_id] = {"valorRecebido": valor, "nome": op_nome}
                            tc_mc += valor
                    agentes_dict = agentes_luc
                    tc_total = tc_luc
                    agentes_dict_mc = agentes_mc
                    tc_total_mc = tc_mc
                else:
                    # Outras equipes — não divide
                    ops_eq = buscar_operadores(eq_at)
                    agentes_dict = {}
                    tc_total = 0.0
                    agentes_dict_mc = {}
                    tc_total_mc = 0.0
                    for _, row in res_at['df_result'].iterrows():
                        nome = str(row['agente']).strip()
                        valor = float(row['valor'])
                        op_id, op_nome = match_operador(nome, ops_eq)
                        if op_id is None:
                            op_id = f"auto-{nome.lower().replace(' ','-')}"
                            op_nome = nome
                        if op_id in agentes_dict:
                            agentes_dict[op_id]["valorRecebido"] += valor
                        else:
                            agentes_dict[op_id] = {"valorRecebido": valor, "nome": op_nome}
                        tc_total += valor
                # Salvar lançamento com resultado por atendente
                # Buscar recGeral do lançamento mais recente da equipe
                dt_eq = int(res_at.get('dias_trab', 0))
                td_eq = int(res_at.get('total_dias', 21))
                _lancs_eq = buscar_lancamentos(ma_at, eq_at)
                _rec_geral_eq = next((float(l.get('recGeral',0)) for l in _lancs_eq if float(l.get('recGeral',0)) > 0), res_at.get('n_elegivel_valor', 0))
                get_db().lancamentos.insert_one({
                    "_id": f"lanc__{ma_at}__{eq_at}__{_ts}",
                    "mesAno": ma_at, "equipeId": eq_at,
                    "dataRef": str(_dt.now().date()),
                    "label": _dt.now().strftime("%d/%m/%Y"),
                    "agentes": agentes_dict,
                    "totalEquipe": tc_total,
                    "semInteracao": 0,
                    "diasTrabalhados": dt_eq, "totalDias": td_eq,
                    "recGeral": _rec_geral_eq,
                    "criadoEm": _dt.now().isoformat()
                })
                # Se for Luciano, salvar Meet Call também
                if eq_at == 'luciano' and tc_total_mc > 0:
                    _ts_mc = _dt.now().strftime("%Y%m%d%H%M%S%f") + "at"
                    _lancs_mc = buscar_lancamentos(ma_at, 'metcool')
                    _rec_geral_mc = next((float(l.get('recGeral',0)) for l in _lancs_mc if float(l.get('recGeral',0)) > 0), 0)
                    get_db().lancamentos.insert_one({
                        "_id": f"lanc__{ma_at}__metcool__{_ts_mc}",
                        "mesAno": ma_at, "equipeId": "metcool",
                        "dataRef": str(_dt.now().date()),
                        "label": _dt.now().strftime("%d/%m/%Y"),
                        "agentes": agentes_dict_mc,
                        "totalEquipe": tc_total_mc,
                        "semInteracao": 0,
                        "diasTrabalhados": dt_eq, "totalDias": td_eq,
                        "recGeral": _rec_geral_mc,
                        "criadoEm": _dt.now().isoformat()
                    })
                buscar_lancamentos.clear()
                buscar_metas_equipe.clear()
                st.session_state['resultado_atendentes'] = None
                if eq_at == 'luciano':
                    st.success(f"✅ Salvo! Luciano Com Interação: {fmt_brl(tc_total)} | Meet Call Com Interação: {fmt_brl(tc_total_mc)}")
                else:
                    st.success(f"✅ Resultado salvo! Com Interação: {fmt_brl(tc_total)}")
                st.rerun()
        with col_cancel2:
            if st.button("❌ Descartar", use_container_width=True, key="btn_descartar_at"):
                st.session_state['resultado_atendentes'] = None
                st.rerun()

    # Mostrar resultado da divisão proporcional Luciano (aguarda confirmação)
    if 'div_prop_resultado' in st.session_state and st.session_state['div_prop_resultado']:
        res = st.session_state['div_prop_resultado']
        st.markdown("---")
        st.markdown("### 📊 Resultado da Divisão Proporcional")
        st.markdown(f"**Mês:** {res['ma']} | **Boletos:** {res['n_boletos']} | **Elegíveis:** {res['n_elegivel']}")
        c1, c2, c3 = st.columns(3)
        with c1:
            st.metric("Equipe Luciano", fmt_brl(res['total_luciano']))
        with c2:
            st.metric("Meet Call", fmt_brl(res['total_metcool']))
        with c3:
            st.metric("Total Geral", fmt_brl(res['total_geral']))

        # Download detalhamento + elegíveis na mesma planilha
        if res.get('df_agentes_det') is not None and not res['df_agentes_det'].empty:
            try:
                import io
                buf2 = io.BytesIO()
                with pd.ExcelWriter(buf2, engine='openpyxl') as writer:
                    res['df_agentes_det'].to_excel(writer, index=False, sheet_name='Divisão Proporcional')
                    if res.get('df_eleg') is not None and not res['df_eleg'].empty:
                        df_eleg_exp = res['df_eleg'].copy()
                        cols_excluir = ['equipe','mes_ano','_row_id','dias_vencidos']
                        cols_ok = [c for c in df_eleg_exp.columns if c not in cols_excluir]
                        df_eleg_exp[cols_ok].to_excel(writer, index=False, sheet_name='Elegíveis')
                buf2.seek(0)
                st.download_button("⬇️ Baixar Detalhamento", buf2.getvalue(),
                    file_name=f"detalhe_divisao_{res['ma']}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_det_div")
            except: pass

        # Tabelas por agente separadas
        if 'df_agentes' in res and res['df_agentes'] is not None:
            df_ag = res['df_agentes']
            col_l, col_m = st.columns(2)
            with col_l:
                st.markdown(f"**🟢 Equipe Luciano** — {fmt_brl(res['total_luciano'])}")
                df_luc = df_ag[df_ag['equipe']=='luciano'].groupby('agente')['valor_proporcional'].sum().reset_index()
                df_luc = df_luc.sort_values('valor_proporcional', ascending=False)
                df_luc['valor_proporcional'] = df_luc['valor_proporcional'].apply(fmt_brl)
                df_luc.columns = ['Agente', 'Total Recebido']
                df_luc = df_luc.reset_index(drop=True)
                df_luc.index += 1
                st.dataframe(df_luc, use_container_width=True)
            with col_m:
                st.markdown(f"**🔵 Meet Call** — {fmt_brl(res['total_metcool'])}")
                df_mc = df_ag[df_ag['equipe']=='metcool'].groupby('agente')['valor_proporcional'].sum().reset_index()
                df_mc = df_mc.sort_values('valor_proporcional', ascending=False)
                df_mc['valor_proporcional'] = df_mc['valor_proporcional'].apply(fmt_brl)
                df_mc.columns = ['Agente', 'Total Recebido']
                df_mc = df_mc.reset_index(drop=True)
                df_mc.index += 1
                st.dataframe(df_mc, use_container_width=True)

        st.markdown("**⚠️ Confirme os valores antes de salvar no Quadro:**")
        col_ok, col_cancel = st.columns(2)
        with col_ok:
            if st.button("✅ Confirmar e Salvar", use_container_width=True, key="btn_confirmar_div"):
                from datetime import datetime as _dt
                _ts = _dt.now().strftime("%Y%m%d%H%M%S%f")

                # Salvar Luciano — novo lançamento com recGeral
                lancs_luc = buscar_lancamentos(res['ma'], 'luciano')
                dt_luc = int(lancs_luc[0].get('diasTrabalhados',0)) if lancs_luc else 0
                td_luc = int(lancs_luc[0].get('totalDias',21)) if lancs_luc else 21
                get_db().lancamentos.insert_one({
                    "_id": f"lanc__{res['ma']}__luciano__{_ts}",
                    "mesAno": res['ma'], "equipeId": "luciano",
                    "dataRef": str(_dt.now().date()),
                    "label": _dt.now().strftime("%d/%m/%Y"),
                    "agentes": {}, "totalEquipe": 0, "semInteracao": 0,
                    "diasTrabalhados": dt_luc, "totalDias": td_luc,
                    "recGeral": res['total_luciano'],
                    "criadoEm": _dt.now().isoformat()
                })
                # Salvar breakdown por UF do Luciano no processamento
                if res.get('por_uf_luciano'):
                    get_db().processamentos.update_one(
                        {"_id": f"proc__{res['ma']}__luciano"},
                        {"$set": {"porFornecedora": {"COMERC": {"valor": res['total_luciano'], "boletos": res['n_elegivel'], "porUF": res['por_uf_luciano']}}, "valorElegivel": res['total_luciano'], "mesAno": res['ma'], "equipeId": "luciano", "atualizadoEm": _dt.now()}},
                        upsert=True
                    )

                # Salvar Meet Call — novo lançamento + lancamento_meetcall
                _ts2 = _dt.now().strftime("%Y%m%d%H%M%S%f") + "mc"
                lancs_mc = buscar_lancamentos(res['ma'], 'metcool')
                dt_mc = int(lancs_mc[0].get('diasTrabalhados',0)) if lancs_mc else 0
                td_mc = int(lancs_mc[0].get('totalDias',21)) if lancs_mc else 21
                get_db().lancamentos.insert_one({
                    "_id": f"lanc__{res['ma']}__metcool__{_ts2}",
                    "mesAno": res['ma'], "equipeId": "metcool",
                    "dataRef": str(_dt.now().date()),
                    "label": _dt.now().strftime("%d/%m/%Y"),
                    "agentes": {}, "totalEquipe": 0, "semInteracao": 0,
                    "diasTrabalhados": dt_mc, "totalDias": td_mc,
                    "recGeral": res['total_metcool'],
                    "criadoEm": _dt.now().isoformat()
                })
                # Salvar breakdown por UF da Meet Call
                if res.get('por_uf_metcool'):
                    get_db().processamentos.update_one(
                        {"_id": f"proc__{res['ma']}__metcool"},
                        {"$set": {"porFornecedora": {"COMERC": {"valor": res['total_metcool'], "boletos": res['n_elegivel'], "porUF": res['por_uf_metcool']}}, "valorElegivel": res['total_metcool'], "mesAno": res['ma'], "equipeId": "metcool", "atualizadoEm": _dt.now()}},
                        upsert=True
                    )
                try:
                    # Salvar na coleção correta que o Quadro lê
                    get_db().metas.update_one(
                        {"_id": f"meetcall__{res['ma']}"},
                        {"$set": {
                            "mesAno": res['ma'],
                            "recGeralTotal": res['total_metcool'],
                            "recGeral": res['total_metcool'],
                            "atualizadoEm": _dt.now().isoformat()
                        }},
                        upsert=True
                    )
                except Exception as e_mc:
                    st.error(f"Erro ao salvar Meet Call: {e_mc}")
                buscar_lancamentos.clear()
                buscar_metas_equipe.clear()
                st.session_state['div_prop_resultado'] = None
                st.success("✅ Salvo! Luciano: " + fmt_brl(res['total_luciano']) + " | Meet Call: " + fmt_brl(res['total_metcool']))
                st.rerun()
        with col_cancel:
            if st.button("❌ Descartar", use_container_width=True, key="btn_descartar_div"):
                st.session_state['div_prop_resultado'] = None
                st.rerun()

    with col_hist:
        st.markdown('<p style="color:#3a6a4a;font-size:10px;text-transform:uppercase;letter-spacing:1.5px;margin-bottom:8px">HISTORICO DE BASES PROCESSADAS</p>',unsafe_allow_html=True)
        u_hist=st.session_state.usuario
        try:
            if u_hist['role'] in ['admin','diretor']:
                hist_geral=buscar_historico_geral(mes_ano=ma)
            else:
                hist_geral=buscar_historico_geral(mes_ano=ma, equipe_id=u_hist.get('equipe'))
        except Exception as e:
            hist_geral=[]
            st.warning(f"Não foi possível carregar o histórico: {e}")
        if not hist_geral:
            st.info('Nenhuma base processada ainda.')
        else:
            for h in hist_geral:
                forns=h.get('fornecedoras',[])
                forn_str=', '.join(forns[:3])+('...' if len(forns)>3 else '') if forns else '---'
                equipe_nome=EQUIPES.get(h.get('equipeId',''),{}).get('nome','---')
                usuario_nome=h.get('usuarioNome') or equipe_nome
                data_str=str(h.get('criadoEm',''))[:16]
                val=float(h.get('valorElegivel',0))
                c_hist1,c_hist2=st.columns([5,1])
                with c_hist1:
                    st.markdown(
                        f"<div style='background:#0a1a0a;border:1px solid #1e3a1e;border-radius:10px;"
                        f"padding:12px 16px;margin-bottom:6px;border-left:3px solid #00c853'>"
                        f"<div style='display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:4px'>"
                        f"<div>"
                        f"<div style='color:#ffffff;font-weight:600;font-size:13px'>{equipe_nome} -- {h.get('mesAno','').replace('-',' ')}</div>"
                        f"<div style='color:#3a6a4a;font-size:11px;margin-top:2px'>Por: {usuario_nome} | {data_str}</div>"
                        f"<div style='color:#3a6a4a;font-size:11px;margin-top:2px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:220px'>Fornecedoras: {forn_str}</div>"
                        f"</div>"
                        f"<div style='text-align:right'>"
                        f"<div style='color:#00c853;font-weight:700;font-size:14px'>{fmt_brl(val)}</div>"
                        f"<div style='color:#3a6a4a;font-size:11px'>{h.get('boletosElegiveis',0):,} boletos</div>"
                        f"</div></div></div>",
                        unsafe_allow_html=True)
                with c_hist2:
                    st.markdown("<div style='margin-top:8px'></div>",unsafe_allow_html=True)
                    if u_hist['role'] in ['admin','diretor']:
                        if st.button("🗑️",key=f"del_hist_{h['_id']}",help="Excluir este processamento"):
                            try:
                                get_db().processamentos.delete_one({"_id":h["_id"]})
                                get_db().historico_processamentos.delete_one({"_id":h["_id"]})
                                st.success("✅ Processamento excluído!")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Erro: {e}")
            st.markdown('<div style="height:8px"></div>',unsafe_allow_html=True)
            if st.button('Exportar Historico Excel',use_container_width=True,key='btn_exp_hist'):
                rows=[{'Gestor':h.get('usuarioNome','---'),'Equipe':EQUIPES.get(h.get('equipeId',''),{}).get('nome','---'),'Mes':h.get('mesAno','---'),'Fornecedoras':', '.join(h.get('fornecedoras',[])),'Valor Recebido':fmt_brl(h.get('valorElegivel',0)),'Boletos':h.get('boletosElegiveis',0),'Data':str(h.get('criadoEm',''))[:16]} for h in hist_geral]
                out=io.BytesIO()
                with pd.ExcelWriter(out,engine='xlsxwriter') as w: pd.DataFrame(rows).to_excel(w,index=False)
                st.download_button('Baixar Excel',data=out.getvalue(),file_name='historico_bases.xlsx',mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',key='dl_hist_exp')

# ── ANÁLISE DE INADIMPLÊNCIA ───────────────────
def pagina_inadimplencia(ma):
    u=st.session_state.usuario
    is_dir=u["role"]=="diretor"; is_adm=u["role"]=="admin"
    header_page("Análise de Inadimplência","Taxa de recuperação por faixa e fornecedora")
    FAIXAS=["D0-30","D31-60","D61-90","D90+"]
    c1,c2,c3=st.columns(3)
    with c1:
        md=listar_meses_inadimplencia() or [ma]
        if ma not in md: md=[ma]+md
        ms=st.selectbox("Mês",md,key="inad_mes")
    with c3:
        eq=seletor_equipe(u.get("equipe") or "tamires",key_suffix="_inad") if (is_adm or is_dir) else u["equipe"]
    if is_dir or is_adm:
        forns_disp = ["Todas"] + FORNECEDORAS_TODAS
    else:
        forns_disp = ["Todas"] + FORNECEDORAS_POR_GESTOR.get(eq or u.get("equipe",""), FORNECEDORAS_TODAS)
    with c2:
        fs=st.selectbox("Fornecedora", forns_disp, key="inad_forn")
    st.markdown("---")
    doc=buscar_inadimplencia(ms,eq or "tamires")
    dados=doc.get("dados",{}) if doc else {}
    with st.expander("📂 Subir base de inadimplência",expanded=not bool(dados)):
        st.markdown(
            "<div style='background:#0d1a0d;border:1px solid #1e3a1e;border-radius:8px;padding:10px 14px;"
            "margin-bottom:10px;font-size:12px;color:#5a9a70;line-height:1.8'>"
            "Aceita <strong style='color:#e8f5e9'>.xlsx</strong> ou <strong style='color:#e8f5e9'>.csv</strong><br>"
            "Colunas necessárias: <strong>fornecedora, dtvencimento, dtpagamento, valor</strong><br>"
            "Outras colunas são ignoradas. Data de referência: último dia do mês selecionado."
            "</div>", unsafe_allow_html=True)
        arq_i=st.file_uploader("Base (.xlsx ou .csv)",type=["xlsx","csv"],label_visibility="collapsed",key=f"arq_inad_{ms}_{eq}")
        col_proc1, col_proc2 = st.columns(2)
        with col_proc1:
            if st.button("🗑️ Limpar dados do mês",use_container_width=True,key=f"btn_limpar_inad_{ms}_{eq}"):
                salvar_inadimplencia(ms, eq or u.get("equipe","tamires"), {})
                st.success("✅ Dados do mês limpos!")
                st.rerun()
        if arq_i:
            with col_proc2:
                if st.button("⚙️ Processar Base",use_container_width=True,key=f"btn_proc_inad_{ms}_{eq}"):
                    arq_i.seek(0)
                    with st.spinner("Processando base..."):
                        resultado_inad,erro_inad=processar_base_inadimplencia(arq_i,eq or u.get("equipe","tamires"),ms)
                    if erro_inad:
                        st.error(erro_inad)
                    elif resultado_inad:
                        # Substituir completamente — não mesclar com dados antigos
                        salvar_inadimplencia(ms,eq or u.get("equipe","tamires"),resultado_inad)
                        st.success(f"✅ Base processada! {len(resultado_inad)} fornecedoras encontradas.")
                        st.rerun()
    st.markdown("---")
    edit=st.checkbox("Editar manualmente",key="edit_inad")
    if is_dir or is_adm:
        forns_usuario = FORNECEDORAS_TODAS
    else:
        forns_usuario = FORNECEDORAS_POR_GESTOR.get(u.get('equipe',''), FORNECEDORAS_TODAS)
    lista = forns_usuario if fs=='Todas' else [fs]
    st.markdown("""<style>
    .it{width:100%;border-collapse:collapse;font-size:11px}
    .it th{background:#1b5e20;color:#fff;padding:6px 8px;text-align:center;border:1px solid #145214;white-space:nowrap}
    .it th.thl{text-align:left}.it th.ths{background:#2e7d32;font-size:10px}
    .it td{border:1px solid #c8e6c9;padding:5px 8px;text-align:right;font-size:11px;color:#1b5e20;background:#fff}
    .it td.tdn{text-align:left;font-weight:600;background:#f1f8f1}
    .it td.tdp{color:#e53935;font-weight:600}.it td.tdf{color:#1565c0;font-weight:500}
    .it tr.trt td{background:#e8f5e9;font-weight:700}
    </style>""",unsafe_allow_html=True)
    html='<div style="overflow-x:auto"><table class="it"><thead>'
    html+='<tr><th class="thl" rowspan="2">Fornecedoras</th>'
    for f in FAIXAS: html+=f'<th colspan="4">{f}</th>'
    html+='<th colspan="3">Total</th></tr><tr>'
    for _ in FAIXAS: html+='<th class="ths">Pagos R$</th><th class="ths">Vencidos R$</th><th class="ths">%Faixa</th><th class="ths">%Inad</th>'
    html+='<th class="ths">Pagos R$</th><th class="ths">Vencidos R$</th><th class="ths">%Inad Geral</th>'
    html+='</tr></thead><tbody>'
    tots={f:{"p":0,"v":0} for f in FAIXAS}; tots["T"]={"p":0,"v":0}
    for forn in lista:
        cor=CORES_FORN.get(forn,"#333"); fd=dados.get(forn,{})
        html+=f'<tr><td class="tdn"><span style="display:inline-block;width:8px;height:8px;border-radius:2px;background:{cor};margin-right:6px"></span>{forn}</td>'
        tp=tv=0
        # Calcular total geral da fornecedora (todas as faixas)
        total_geral_forn = sum(
            float(fd.get(fx,{}).get("pagos",0)) + float(fd.get(fx,{}).get("vencidos",0))
            for fx in FAIXAS)
        for faixa in FAIXAS:
            p=float(fd.get(faixa,{}).get("pagos",0)); v=float(fd.get(faixa,{}).get("vencidos",0))
            # %Faixa = Vencidos / Total da faixa
            pf=(v/(p+v)*100) if (p+v)>0 else 0
            # %Inad = Vencidos da faixa / Total geral da fornecedora
            pi=(v/total_geral_forn*100) if total_geral_forn>0 else 0
            tots[faixa]["p"]+=p; tots[faixa]["v"]+=v; tp+=p; tv+=v
            html+=f'<td>{fmt_brl_td(p)}</td><td>{fmt_brl_td(v)}</td><td class="tdf">{pf:.2f}%</td><td class="tdp">{pi:.2f}%</td>'
        tots["T"]["p"]+=tp; tots["T"]["v"]+=tv
        pg=(tv/(tp+tv)*100) if (tp+tv)>0 else 0
        html+=f'<td>{fmt_brl_td(tp)}</td><td>{fmt_brl_td(tv)}</td><td class="tdp">{pg:.2f}%</td></tr>'
    html+='<tr class="trt"><td class="tdn">TOTAL</td>'
    for f in FAIXAS:
        tp2=tots[f]["p"]; tv2=tots[f]["v"]; pf2=(tp2/(tp2+tv2)*100) if (tp2+tv2)>0 else 0; pi2=(tv2/(tp2+tv2)*100) if (tp2+tv2)>0 else 0
        html+=f'<td>{fmt_brl_td(tp2)}</td><td>{fmt_brl_td(tv2)}</td><td class="tdf">{pf2:.2f}%</td><td class="tdp">{pi2:.2f}%</td>'
    tpt=tots["T"]["p"]; tvt=tots["T"]["v"]; pgt=(tvt/(tpt+tvt)*100) if (tpt+tvt)>0 else 0
    html+=f'<td>{fmt_brl_td(tpt)}</td><td>{fmt_brl_td(tvt)}</td><td class="tdp">{pgt:.2f}%</td></tr>'
    html+='</tbody></table></div>'
    st.markdown(html,unsafe_allow_html=True)
    st.markdown("<div style='height:12px'></div>",unsafe_allow_html=True)
    cols_r=st.columns(4)
    for idx,f in enumerate(FAIXAS):
        tp2=tots[f]["p"]; tv2=tots[f]["v"]
        pf2=(tp2/(tp2+tv2)*100) if (tp2+tv2)>0 else 0; pi2=(tv2/(tp2+tv2)*100) if (tp2+tv2)>0 else 0
        with cols_r[idx]:
            st.markdown(f"""<div style="background:#e8f5e9;border:1px solid #a5d6a7;border-radius:8px;padding:10px;text-align:center">
                <div style="font-size:11px;font-weight:600;color:#2e7d32;margin-bottom:6px">{f}</div>
                <div style="display:flex;justify-content:center;gap:16px">
                    <div><div style="font-size:9px;color:#555;text-transform:uppercase">%Faixa</div><div style="font-size:14px;font-weight:700;color:#1565c0">{pf2:.2f}%</div></div>
                    <div><div style="font-size:9px;color:#555;text-transform:uppercase">%Inad</div><div style="font-size:14px;font-weight:700;color:#e53935">{pi2:.2f}%</div></div>
                </div></div>""",unsafe_allow_html=True)
    if edit:
        st.markdown("---")
        st.markdown("### Edição Manual")
        nd={}
        for forn in lista:
            st.markdown(f"**{forn}**"); fd=dados.get(forn,{}); nd[forn]={}
            cols_e=st.columns(4)
            for idx_f,f in enumerate(FAIXAS):
                with cols_e[idx_f]:
                    st.markdown(f"<div style='font-size:11px;color:#2e7d32;font-weight:600;margin-bottom:4px'>{f}</div>",unsafe_allow_html=True)
                    p=st.number_input(f"Pagos {f}",min_value=0.0,step=100.0,format="%.2f",value=float(fd.get(f,{}).get("pagos",0)),key=f"ip_{forn}_{f}",label_visibility="collapsed")
                    st.markdown("<div style='font-size:10px;color:#555;margin-bottom:2px'>Pagos R$</div>",unsafe_allow_html=True)
                    v=st.number_input(f"Vencidos {f}",min_value=0.0,step=100.0,format="%.2f",value=float(fd.get(f,{}).get("vencidos",0)),key=f"iv_{forn}_{f}",label_visibility="collapsed")
                    st.markdown("<div style='font-size:10px;color:#e53935;margin-bottom:2px'>Vencidos R$</div>",unsafe_allow_html=True)
                    nd[forn][f]={"pagos":p,"vencidos":v}
            st.markdown("---")
        if st.button("Salvar Dados de Inadimplência",use_container_width=True):
            salvar_inadimplencia(ms,eq or "tamires",nd); st.success("Dados salvos!"); st.rerun()
    st.markdown("---")
    if st.button("Exportar Excel"):
        rows=[{"Fornecedora":f,**{f"{faixa} Pagos":float(dados.get(f,{}).get(faixa,{}).get("pagos",0)) for faixa in FAIXAS},**{f"{faixa} Vencidos":float(dados.get(f,{}).get(faixa,{}).get("vencidos",0)) for faixa in FAIXAS}} for f in lista]
        out=io.BytesIO()
        with pd.ExcelWriter(out,engine="xlsxwriter") as w: pd.DataFrame(rows).to_excel(w,index=False)
        st.download_button("Baixar Excel",data=out.getvalue(),file_name=f"Inadimplencia_{ms}.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── CRITÉRIOS ─────────────────────────────────
def pagina_criterios():
    header_page("Critérios de Monitoria","Configure os critérios de avaliação")
    crits=get_criterios(); erros=get_erros_criticos()
    t1,t2=st.tabs(["Critérios de Avaliação","Erros Críticos"])
    with t1:
        st.markdown("**Alterações valem apenas para novas monitorias.**"); st.markdown("---")
        ce=[]
        for i,c in enumerate(crits):
            with st.expander(f"{c['num']} {c['nome']} — Peso {c['peso']}",expanded=False):
                col1,col2,col3=st.columns([3,1,1])
                with col1: nm=st.text_input("Nome",value=c["nome"],key=f"cn_{i}")
                with col2: ps=st.number_input("Peso",min_value=1,max_value=100,value=int(c["peso"]),key=f"cp_{i}")
                with col3: ob=st.checkbox("Obrigatório",value=c.get("obrigatorio",False),key=f"co_{i}")
                it=st.text_area("Itens (um por linha)",value="\n".join(c.get("itens",[])),height=100,key=f"ci_{i}")
                ce.append({"id":c["id"],"num":c["num"],"nome":nm,"peso":ps,"obrigatorio":ob,"itens":[x.strip() for x in it.split("\n") if x.strip()]})
        st.markdown("---")
        if st.button("Salvar Critérios",use_container_width=True): salvar_criterios(ce); st.success("Critérios salvos!"); st.rerun()
    with t2:
        st.markdown("**Erros que zeram a monitoria automaticamente.**"); st.markdown("---")
        ee=[]
        for i,e in enumerate(erros):
            col1,col2=st.columns([2,3])
            with col1: ne=st.text_input("Nome",value=e["nome"],key=f"en_{i}")
            with col2: de=st.text_input("Descrição",value=e["desc"],key=f"ed_{i}")
            ee.append({"id":e["id"],"nome":ne,"desc":de})
        st.markdown("---")
        col1,col2=st.columns(2)
        with col1:
            if st.button("Salvar Erros Críticos",use_container_width=True): salvar_erros_criticos(ee); st.success("Salvo!"); st.rerun()
        with col2:
            if st.button("Adicionar Erro",use_container_width=True):
                ee.append({"id":f"e{len(erros)+1}","nome":"Novo erro","desc":"Descrição"}); salvar_erros_criticos(ee); st.rerun()

# ── MINHA CONTA ──────────────────────────────
def pagina_minha_conta():
    u=st.session_state.usuario
    header_page('Minha Conta', u['nome'])
    t1,t2,t3=st.tabs(['🔒  Senha','👥  Operadores','📋  Critérios'])
    with t1:
        st.markdown('<p style="color:#5a9a70;font-size:13px;margin-bottom:20px">Altere sua senha de acesso</p>',unsafe_allow_html=True)
        sa =st.text_input('Senha atual',    type='password',placeholder='senha atual',   key='mc_sa')
        sn =st.text_input('Nova senha',     type='password',placeholder='mín. 8 caracteres',key='mc_sn')
        sc2=st.text_input('Confirmar senha',type='password',placeholder='repita a nova senha',key='mc_sc')
        if st.button('Salvar Senha',key='mc_btn_senha',use_container_width=True):
            uid=u['id']; sc=u.get('senha')
            try:
                doc=get_db().usuarios_senhas.find_one({'_id':uid})
                if doc and doc.get('senha'): sc=doc['senha']
            except: pass
            if not sa: st.error('Digite a senha atual.')
            elif sa!=sc: st.error('Senha atual incorreta.')
            elif len(sn)<8: st.error('Mínimo 8 caracteres.')
            elif sn!=sc2: st.error('Confirmação não confere.')
            else: salvar_senha_usuario(uid,sn); buscar_senha_usuario.clear(); st.success('✅ Senha alterada com sucesso!')
    with t2:
        # Diretor pode editar operadores de qualquer equipe
        if u["role"]=="diretor":
            eq_opts=list(EQUIPES.keys())
            eq_labels=[f"Equipe {EQUIPES[e]['nome']}" for e in eq_opts]
            eq_sel=st.selectbox("Selecionar equipe:",eq_labels,key="dir_eq_ops")
            eq=eq_opts[eq_labels.index(eq_sel)]
        else:
            eq=u.get('equipe')
        if not eq:
            st.info('Gestão de operadores disponível apenas para gestores.')
        else:
            st.markdown(f'**Equipe {EQUIPES.get(eq,{}).get("nome",eq)}**')
            st.markdown('---')
            c1,c2,c3=st.columns([3,1,1])
            with c1: nn=st.text_input('Nome completo',placeholder='Nome do operador',key='mc_op_novo')
            with c2: np=st.checkbox('Pleno',key='mc_op_pleno')
            with c3:
                st.markdown("<div style='margin-top:28px'>",unsafe_allow_html=True)
                if st.button('Buscar / Cadastrar',use_container_width=True,key='mc_op_add'):
                    if nn.strip():
                        primeiro = nn.strip().upper().split()[0]
                        encontrado = next((op for op in get_db().operadores.find({"equipeId": {"$ne": eq}})
                                          if op.get('nome','').strip().upper().split()[0] == primeiro), None)
                        if encontrado:
                            st.session_state['mc_op_busca_result'] = {'op': encontrado, 'eq': eq, 'nome': nn.strip(), 'pleno': np}
                        else:
                            salvar_operador(eq, nn.strip(), np)
                            buscar_operadores.clear()
                            st.session_state['mc_op_busca_result'] = None
                            st.success(f"✅ {nn} cadastrado!")
                            st.rerun()
                    else: st.error('Digite o nome.')
                st.markdown('</div>',unsafe_allow_html=True)

            # Resultado da busca — vinculação
            if st.session_state.get('mc_op_busca_result'):
                r = st.session_state['mc_op_busca_result']
                st.warning(f"⚠️ **{r['op']['nome']}** já existe na equipe **{r['op'].get('equipeId','')}**.")
                st.markdown("Quer vincular o mesmo operador (mantém histórico) ou criar novo?")
                cv1, cv2, cv3 = st.columns(3)
                with cv1:
                    if st.button("🔗 Vincular", key="mc_op_vincular", use_container_width=True):
                        novo_id = f"{r['op']['_id']}-{r['eq']}"
                        get_db().operadores.update_one(
                            {"_id": novo_id},
                            {"$set": {"nome": r['op']['nome'], "equipeId": r['eq'], "pleno": r['pleno'], "vinculadoA": r['op']['_id']}},
                            upsert=True)
                        buscar_operadores.clear()
                        st.session_state['mc_op_busca_result'] = None
                        st.rerun()
                with cv2:
                    if st.button("➕ Criar novo", key="mc_op_criar", use_container_width=True):
                        salvar_operador(r['eq'], r['nome'], r['pleno'])
                        buscar_operadores.clear()
                        st.session_state['mc_op_busca_result'] = None
                        st.rerun()
                with cv3:
                    if st.button("❌ Cancelar", key="mc_op_cancelar", use_container_width=True):
                        st.session_state['mc_op_busca_result'] = None
                        st.rerun()
            st.markdown('---')
            # Luciano: excluir Meet Call da lista
            ops_todos=buscar_operadores(eq)
            ops=[op for op in ops_todos if op["nome"] not in OPERADORES_MEETCALL] if eq=="luciano" else ops_todos
            if not ops: st.info('Nenhum operador cadastrado.')
            else:
                for op in ops:
                    c1,c2,c3,c4=st.columns([3,1,1,1])
                    with c1: ne=st.text_input('',value=op['nome'],key=f'mc_n_{op["_id"]}',label_visibility='collapsed')
                    with c2: npl=st.checkbox('Pleno',value=op.get('pleno',False),key=f'mc_pl_{op["_id"]}')
                    with c3:
                        if st.button('Salvar',key=f'mc_s_{op["_id"]}',use_container_width=True):
                            atualizar_operador(op['_id'],ne,npl)
                            buscar_operadores.clear()
                            st.success(f"✅ {ne} atualizado com sucesso!"); st.rerun()
                    with c4:
                        if st.button('Excluir',key=f'mc_d_{op["_id"]}',use_container_width=True):
                            excluir_operador(op['_id'])
                            buscar_operadores.clear()
                            st.success(f"✅ {op['nome']} excluído!")
                            st.rerun()
    with t3:
        crits=get_criterios()
        st.markdown('**Critérios de avaliação das monitorias**')
        st.markdown('---')
        ce=[]
        for i,c in enumerate(crits):
            with st.expander(f"{c['num']} {c['nome']} — Peso {c['peso']}",expanded=False):
                c1,c2,c3=st.columns([3,1,1])
                with c1: nm=st.text_input('Nome',value=c['nome'],key=f'mc_cn_{i}')
                with c2: ps=st.number_input('Peso',min_value=1,max_value=100,value=int(c['peso']),key=f'mc_cp_{i}')
                with c3: ob=st.checkbox('Obrigatório',value=c.get('obrigatorio',False),key=f'mc_co_{i}')
                it=st.text_area('Itens (um por linha)',value='\n'.join(c.get('itens',[])),height=80,key=f'mc_ci_{i}')
                ce.append({'id':c['id'],'num':c['num'],'nome':nm,'peso':ps,'obrigatorio':ob,'itens':[x.strip() for x in it.split('\n') if x.strip()]})
        st.markdown('---')
        if st.button('Salvar Critérios',use_container_width=True,key='mc_crit_save'):
            salvar_criterios(ce); st.success('Critérios salvos!'); st.rerun()

# ── MEET CALL ─────────────────────────────────
def pagina_meetcall(ma):
    u=st.session_state.usuario
    if u.get("equipe") not in ["metcool","luciano"] and u["role"] not in ["admin","diretor"]:
        st.warning("Acesso restrito."); return
    header_page("Meet Call","Lançamento da equipe Meet Call")
    ops_mc=buscar_operadores("metcool")
    ms_mc=buscar_metas_equipe(ma,"metcool")
    mg_mc=float(buscar_meta_gestora(ma,"metcool").get("metaGestora",0))
    if st.session_state.get("mc_ultimo_salvo"):
        st.success(st.session_state.mc_ultimo_salvo)
        st.session_state.mc_ultimo_salvo=""

    st.markdown("### Configuração")
    c1,c2,c3=st.columns([2,1,1])
    with c1:
        hoje=date.today()
        data_sel=st.date_input("Data *",value=hoje,key=f"mc_data_{ma}")
        eh_fech=st.checkbox("Fechamento do Mês",key=f"mc_fech_{ma}")
    with c2: dt_mc=st.number_input("Dias Trabalhados *",min_value=0,max_value=31,value=0,key=f"mc_dt_{ma}")
    with c3: td_mc=st.number_input("Total Dias *",min_value=0,max_value=31,value=0,key=f"mc_td_{ma}")

    # Recebido Geral
    mc_doc=buscar_lancamento_meetcall(ma)
    rg_atual=float(mc_doc.get("recGeralTotal",mc_doc.get("recGeral",0)))
    st.markdown("---")
    usar_rg_manual=st.checkbox("Inserir Recebido Geral manualmente",key=f"mc_rg_chk_{ma}")
    if usar_rg_manual:
        rg_str=st.text_input("Recebido Geral (R$)",value=fmt_brl(rg_atual) if rg_atual>0 else "",placeholder="R$ 0,00",key=f"mc_rg_manual_{ma}")
        rg_total=parse_brl(rg_str)
    else:
        rg_total=rg_atual

    # Com Interação manual (caso não queira lançar por operador)
    usar_ci_manual=st.checkbox("Inserir Com Interação manualmente",key=f"mc_ci_chk_{ma}")
    ci_manual=0.0
    if usar_ci_manual:
        ci_atual=float(mc_doc.get("recGeral",0))
        ci_str=st.text_input("Com Interação (R$)",value=fmt_brl(ci_atual) if ci_atual>0 else "",placeholder="R$ 0,00",key=f"mc_ci_manual_{ma}")
        ci_manual=parse_brl(ci_str)

    st.markdown("---")

    # Sempre mostrar campos manuais de Recebido Geral e Com Interação
    usar_ci_manual_mc=st.checkbox("Inserir Com Interação manualmente",key=f"mc_ci_manual_chk_{ma}")
    ci_manual_mc=0.0
    if usar_ci_manual_mc:
        ci_ant=float(mc_doc.get("recGeral",0))
        ci_str=st.text_input("Com Interação (R$)",value=fmt_brl(ci_ant) if ci_ant>0 else "",placeholder="R$ 0,00",key=f"mc_ci_txt_{ma}")
        ci_manual_mc=parse_brl(ci_str) if ci_str else 0.0
        if ci_manual_mc>0:
            get_db().temp_lancamento.update_one({"_id":f"tmp_mc_{ma}"},{"$set":{"ci_manual":ci_manual_mc}},upsert=True)
        else:
            ci_manual_mc=float((get_db().temp_lancamento.find_one({"_id":f"tmp_mc_{ma}"}) or {}).get("ci_manual",0))

    if ops_mc:
        st.markdown("### Valores por Operador")
    # Importar via Excel
    mostrar_imp_mc=st.checkbox("Importar via Excel",key=f"chk_imp_mc_{ma}") if ops_mc else False
    if mostrar_imp_mc:
        arq_imp_mc=st.file_uploader("Planilha Excel (.xlsx)",type=["xlsx"],key=f"imp_mc_{ma}")
        if arq_imp_mc:
            arq_imp_mc.seek(0)
            res_mc,err_mc=importar_excel_operadores(arq_imp_mc,ops_mc)
            if err_mc:
                st.error(err_mc)
            elif res_mc:
                prev_mc=[]
                for r in res_mc:
                    icone=r['op']['nome'] if r['op'] and r['status']!='ambiguo' else ("Ambíguo" if r['status']=='ambiguo' else "Não encontrado")
                    row={"Excel":r['nome_excel'],"Sistema":icone,"Valor":fmt_brl(r['valor'])}
                    prev_mc.append(row)
                st.dataframe(pd.DataFrame(prev_mc),use_container_width=True,hide_index=True)
                if st.button("Confirmar Importação",use_container_width=True,key=f"imp_mc_confirmar_{ma}"):
                    for r in res_mc:
                        if r['op'] and r['status']!='ambiguo' and r['valor']>0:
                            st.session_state[f"mc_op_{ma}_{r['op']['_id']}"]=r['valor']
                    st.success("Valores importados! Revise e salve.")
                    st.rerun()

    vi_mc={}
    for op in ops_mc:
        meta=float(ms_mc.get(op["_id"],0))
        c1,c2,c3=st.columns([3,2,2])
        with c1: st.markdown(f"<div style='padding-top:10px;color:#1a3a1a;font-weight:500'>{op['nome']}</div>",unsafe_allow_html=True)
        with c2: st.markdown(f"<div style='padding-top:10px;color:#2e7d32;font-size:13px'>{fmt_brl(meta) if meta>0 else '—'}</div>",unsafe_allow_html=True)
        with c3:
            mc_op_key=f"mc_op_{ma}_{op['_id']}"
            val_mc=st.session_state.get(mc_op_key,0.0)
            vi_mc[op["_id"]]=st.number_input("v",label_visibility="collapsed",min_value=0.0,step=100.0,format="%.2f",value=float(val_mc) if val_mc else 0.0,key=mc_op_key)

    tc_mc=sum(vi_mc.values())
    sem_mc=max(0,rg_total-tc_mc) if rg_total>0 else 0
    pct_mc=(rg_total/mg_mc*100) if mg_mc>0 else 0

    st.markdown("---")
    c1,c2,c3=st.columns(3)
    c1.metric("Recebido Geral",fmt_brl(rg_total))
    c2.metric("Com Interação",fmt_brl(tc_mc))
    c3.metric("Sem Interação",fmt_brl(sem_mc))
    st.markdown(f"<div style='background:#0a2414;border-radius:8px;padding:12px 16px;margin:12px 0 16px'><span style='color:#5a9a70;font-size:11px'>META: {fmt_brl(mg_mc)} | % META: {pct_mc:.2f}%</span></div>",unsafe_allow_html=True)

    ja_salvando_mc = st.session_state.get("salvando_mc", False)
    if st.button("Salvar Lançamento Meet Call",use_container_width=True,key="mc_salvar", disabled=ja_salvando_mc):
        errs=[]
        if dt_mc==0: errs.append("Dias Trabalhados é obrigatório.")
        if td_mc==0: errs.append("Total de Dias é obrigatório.")
        if tc_mc==0: errs.append("Preencha pelo menos um valor de operador.")
        if errs:
            for e in errs: st.error(e)
        else:
            st.session_state["salvando_mc"] = True
            label="Fechamento do Mês" if eh_fech else data_sel.strftime("%d/%m/%Y")
            ag_mc={op["_id"]:{"valorRecebido":vi_mc[op["_id"]],"nome":op["nome"]} for op in ops_mc}
            tc_ops_mc=sum(vi_mc.values()) if vi_mc else 0
            # Prioridade: operadores > ci_manual_mc > ci_manual antigo
            if tc_ops_mc>0:
                tc_final=tc_ops_mc
            elif usar_ci_manual_mc and ci_manual_mc>0:
                tc_final=ci_manual_mc
            else:
                tc_final=float((get_db().temp_lancamento.find_one({"_id":f"tmp_mc_{ma}"}) or {}).get("ci_manual",0))
            criar_lancamento(ma,"metcool",str(data_sel),label,ag_mc,tc_final,0,dt_mc,td_mc)
            salvar_lancamento_meetcall(ma,0,tc_final,rg_total)
            buscar_lancamentos.clear()
            st.session_state["salvando_mc"] = False
            st.session_state.mc_ultimo_salvo=f"✅ Lançamento de {label} salvo com sucesso! Total: {fmt_brl(tc_mc)}"
            st.rerun()

    # Histórico
    lancs_mc=buscar_lancamentos(ma,"metcool")
    if lancs_mc:
        st.markdown("---")
        st.markdown("<p style='color:#5a8a5a;font-size:11px;text-transform:uppercase;margin-bottom:8px'>Lançamentos do mês</p>",unsafe_allow_html=True)
        for lanc in reversed(lancs_mc):
            soma=sum(float(v.get("valorRecebido",0) if isinstance(v,dict) else v) for v in lanc.get("agentes",{}).values())
            with st.expander(f"{lanc.get('label','')} — {fmt_brl(soma)}"):
                rows=[{"Operador":op["nome"],"Valor":fmt_brl(get_val_op(lanc.get("agentes",{}),op["_id"],op["nome"]))} for op in ops_mc]
                st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)
                if st.button("Excluir",key=f"mc_del_{lanc['_id']}"): excluir_lancamento(lanc["_id"]); st.rerun()

# ── MAIN ───────────────────────────────────────
def main():
    if "usuario" not in st.session_state:
        tela_login()
        return
    if "ids_corrigidos" not in st.session_state:
        st.session_state.ids_corrigidos=True
    if 'meetcall_limpo' not in st.session_state:
        st.session_state.meetcall_limpo = True
        try:
            db = get_db()
            for nome in OPERADORES_MEETCALL:
                ops_dup = list(db.operadores.find({"equipeId":"metcool","nome":nome}).sort("criadoEm",1))
                if len(ops_dup) > 1:
                    for op_dup in ops_dup[1:]: db.operadores.delete_one({"_id":op_dup["_id"]})
                ops_dup2 = list(db.operadores.find({"equipeId":"luciano","nome":nome}).sort("criadoEm",1))
                if len(ops_dup2) > 1:
                    for op_dup2 in ops_dup2[1:]: db.operadores.delete_one({"_id":op_dup2["_id"]})
        except: pass
    if 'meetcall_migrado' not in st.session_state:
        st.session_state.meetcall_migrado=True
        try: migrar_meetcall_para_luciano()
        except: pass
    ma,pag=render_sidebar()
    u=st.session_state.usuario
    area = st.empty()
    with area.container():
        if u["role"]=="diretor":
            if   "Quadro"        in pag: pagina_quadro(ma)
            elif "Visualização"  in pag: pagina_dashboard_executivo()
            elif "Operadores"    in pag: pagina_analise_operadores(ma)
            elif "Monitorias"    in pag: pagina_monitorias(ma)
            elif "Inadimplência" in pag: pagina_inadimplencia(ma)
            elif "Metas"         in pag: pagina_metas(ma)
            elif "Minha Conta"   in pag: pagina_minha_conta()
        elif u["role"]=="admin":
            if   "Quadro"        in pag: pagina_quadro(ma)
            elif "Lançamento"    in pag: pagina_lancamento(ma)
            elif "Visualização"  in pag: pagina_dashboard_executivo()
            elif "Operadores"    in pag: pagina_analise_operadores(ma)
            elif "Monitorias"    in pag: pagina_monitorias(ma)
            elif "Meet Call"     in pag: pagina_meetcall(ma)
            elif "Upload"        in pag: pagina_upload(ma)
            elif "Inadimplência" in pag: pagina_inadimplencia(ma)
            elif "Metas"         in pag: pagina_metas(ma)
            elif "Minha Conta"   in pag: pagina_minha_conta()
        elif u.get("equipe")=="metcool":
            if   "Meet Call"     in pag: pagina_meetcall(ma)
            elif "Operadores"    in pag: pagina_analise_operadores(ma)
            elif "Monitorias"    in pag: pagina_monitorias(ma)
            elif "Minha Conta"   in pag: pagina_minha_conta()
        else:
            if   "Quadro"        in pag: pagina_quadro(ma)
            elif "Lançamento"    in pag: pagina_lancamento(ma)
            elif "Operadores"    in pag: pagina_analise_operadores(ma)
            elif "Monitorias"    in pag: pagina_monitorias(ma)
            elif "Meet Call"     in pag: pagina_meetcall(ma)
            elif "Upload"        in pag: pagina_upload(ma)
            elif "Inadimplência" in pag: pagina_inadimplencia(ma)
            elif "Metas"         in pag: pagina_metas(ma)
            elif "Minha Conta"   in pag: pagina_minha_conta()

# deploy: 2026-06-05-v3
if __name__=="__main__":
    main()

